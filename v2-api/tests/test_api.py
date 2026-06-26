import html
import importlib.util
import time
from datetime import datetime
from io import BytesIO
from uuid import uuid4
from pathlib import Path
from types import SimpleNamespace

from fastapi.testclient import TestClient

import app.main as main_module
from app.main import create_app
from app.api.routes import auth
from app.core.config import settings
from app.core import security
from app.services.ezcodes_scheduler import sync_manager
from app.services import account_store, local_simulation
from app.services.photo_storage import resolve_photo_for_response


client = TestClient(create_app())


def build_api_workbook(rows: list[list[str]]) -> bytes:
    from io import BytesIO

    import pytest

    pytest.importorskip("openpyxl")
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    for row in rows:
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def load_static_page_verifier():
    verifier_path = Path(__file__).resolve().parents[2] / "scripts" / "verify-static-pages.py"
    spec = importlib.util.spec_from_file_location("verify_static_pages", verifier_path)
    assert spec and spec.loader
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def assert_success_shape(payload: dict) -> None:
    assert "data" in payload
    assert payload["error"] is None
    assert isinstance(payload["request_id"], str)


def assert_active_nav(html: str, href: str) -> None:
    assert f'href="{href}"' in html
    assert (
        f'class="button active" href="{href}"' in html
        or f'class="button secondary active" href="{href}"' in html
        or f'class="nav-link active" href="{href}"' in html
    )


def assert_vue_shell_response(response) -> None:
    assert response.status_code == 200
    assert "text/html" in response.headers["content-type"]
    assert '<div id="app"></div>' in response.text
    assert 'type="module"' in response.text
    assert "/vue/assets/" in response.text
    assert 'id="workFrame"' not in response.text
    assert 'embedded: "1"' not in response.text
    assert "LegacyStaticPageView" not in response.text


def test_health_check() -> None:
    response = client.get("/health", headers={"x-request-id": "test-request"})

    assert response.status_code == 200
    assert response.headers["x-request-id"] == "test-request"
    payload = response.json()
    assert payload["data"] == {"status": "ok"}
    assert payload["request_id"] == "test-request"


def test_system_status_requires_admin_and_reports_runtime_state() -> None:
    admin = client.post("/auth/login", json={"username": "admin", "password": "admin123"})
    reviewer = client.post("/auth/login", json={"username": "reviewer", "password": "review123"})
    admin_headers = {"Authorization": f"bearer {admin.json()['data']['access_token']}"}
    reviewer_headers = {"Authorization": f"bearer {reviewer.json()['data']['access_token']}"}

    denied = client.get("/local-test/system/status", headers=reviewer_headers)
    response = client.get("/local-test/system/status", headers=admin_headers)

    assert denied.status_code == 403
    assert response.status_code == 200
    data = response.json()["data"]
    assert data["version"] == "3.0.33"
    assert {"disk", "state_file", "uploads", "storage", "backups", "teams", "warnings"}.issubset(data)
    assert "used_percent" in data["disk"]
    assert "warn_bytes" in data["uploads"]
    assert data["storage"]["backend"] in {"local", "oss"}


def test_oss_photo_response_resolves_preview_url_without_rewriting_canonical(monkeypatch) -> None:
    monkeypatch.setattr(settings, "oss_public_base_url", "https://oss-preview.example.test/base")
    monkeypatch.setattr(settings, "oss_thumbnail_process", "image/resize,w_120")
    monkeypatch.setattr(settings, "oss_preview_process", "image/resize,w_800")
    photo = {
        "id": "p-oss",
        "image_url": "oss://bucket-a/path/to/photo 1.jpg",
        "storage_type": "oss",
        "storage_key": "path/to/photo 1.jpg",
        "storage_bucket": "bucket-a",
    }

    resolved = resolve_photo_for_response(photo)

    assert resolved["canonical_image_url"] == "oss://bucket-a/path/to/photo 1.jpg"
    assert resolved["image_url"] == "https://oss-preview.example.test/base/path/to/photo%201.jpg"
    assert resolved["thumbnail_url"] == "https://oss-preview.example.test/base/path/to/photo%201.jpg?x-oss-process=image/resize,w_120"
    assert resolved["preview_url"] == "https://oss-preview.example.test/base/path/to/photo%201.jpg?x-oss-process=image/resize,w_800"
    assert photo["image_url"] == "oss://bucket-a/path/to/photo 1.jpg"


def test_oss_photo_response_keeps_raw_reference_when_signing_config_missing(monkeypatch) -> None:
    monkeypatch.setattr(settings, "oss_public_base_url", "")
    monkeypatch.setattr(settings, "oss_endpoint", "")
    monkeypatch.setattr(settings, "oss_internal_endpoint", "")
    monkeypatch.setattr(settings, "oss_bucket", "")
    monkeypatch.setattr(settings, "oss_access_key_id", "")
    monkeypatch.setattr(settings, "oss_access_key_secret", "")
    photo = {
        "id": "p-oss",
        "image_url": "oss://bucket-a/path/to/photo.jpg",
        "storage_type": "oss",
        "storage_key": "path/to/photo.jpg",
        "storage_bucket": "bucket-a",
    }

    resolved = resolve_photo_for_response(photo)

    assert resolved["canonical_image_url"] == "oss://bucket-a/path/to/photo.jpg"
    assert resolved["image_url"] == "oss://bucket-a/path/to/photo.jpg"
    assert resolved["thumbnail_url"] == "oss://bucket-a/path/to/photo.jpg"
    assert resolved["preview_url"] == "oss://bucket-a/path/to/photo.jpg"


def test_broken_oss_processed_preview_falls_back_to_original(monkeypatch) -> None:
    pytest = importlib.import_module("pytest")
    pytest.importorskip("PIL")
    from PIL import Image, ImageDraw

    from app.api.routes import local_test

    broken = Image.new("RGB", (320, 180), (128, 128, 128))
    draw = ImageDraw.Draw(broken)
    palette = [(20, 20, 20), (236, 236, 236), (70, 120, 60), (160, 120, 170)]
    for x in range(84, 236):
        draw.line((x, 0, x, 180), fill=palette[x % len(palette)])
    broken_buffer = BytesIO()
    broken.save(broken_buffer, format="JPEG", quality=95)

    original = Image.new("RGB", (320, 180), (220, 230, 240))
    original_buffer = BytesIO()
    original.save(original_buffer, format="JPEG", quality=90)
    original_bytes = original_buffer.getvalue()

    def fake_sign_oss_server_url(_key: str, process: str = "") -> str:
        return f"signed://{process or 'original'}"

    def fake_read_remote_image(url: str, *, max_bytes: int = 30 * 1024 * 1024):
        if url.endswith("original"):
            return original_bytes, "image/jpeg"
        return broken_buffer.getvalue(), "image/jpeg"

    monkeypatch.setattr(local_test, "sign_oss_server_url", fake_sign_oss_server_url)
    monkeypatch.setattr(local_test, "_read_remote_image", fake_read_remote_image)

    content, media_type = local_test._read_oss_photo_or_repair(
        "g-1",
        {},
        "module-manager-v2/default-team/photos/aa/photo.jpg",
        "",
        "image/resize,w_1280",
    )

    assert media_type == "image/jpeg"
    assert content == original_bytes


def test_login_page_and_demo_auth_are_available() -> None:
    page = client.get("/login")
    config = client.get("/auth/config")
    admin = client.post("/auth/login", json={"username": "admin", "password": "admin123"})
    reviewer = client.post("/auth/login", json={"username": "reviewer", "password": "review123"})
    constructor = client.post("/auth/login", json={"username": "constructor", "password": "construct123"})
    bad = client.post("/auth/login", json={"username": "admin", "password": "bad"})

    assert_vue_shell_response(page)
    assert "admin / admin123" not in page.text
    assert "reviewer / review123" not in page.text
    assert 'value="admin"' not in page.text
    assert 'value="admin123"' not in page.text
    assert config.status_code == 200
    assert config.json()["data"]["demo_auth_enabled"] is True
    assert {item["username"] for item in config.json()["data"]["demo_accounts"]} == {"admin", "reviewer", "constructor"}
    assert {item["team_id"] for item in config.json()["data"]["demo_accounts"]} == {"demo-team"}
    assert admin.status_code == 200
    assert admin.json()["data"]["team_id"] == "demo-team"
    assert admin.json()["data"]["user"]["home"] == "/app"
    assert admin.json()["data"]["user"]["team_id"] == "demo-team"
    assert admin.json()["data"]["user"]["roles"] == ["admin"]
    assert reviewer.status_code == 200
    assert reviewer.json()["data"]["user"]["home"] == "/app"
    assert constructor.status_code == 200
    assert constructor.json()["data"]["user"]["roles"] == ["constructor"]
    assert constructor.json()["data"]["user"]["home"] == "/app?page=construction"
    assert bad.status_code == 401


def test_demo_auth_is_disabled_by_default_in_production(monkeypatch) -> None:
    production_settings = SimpleNamespace(
        app_env="production",
        demo_auth_enabled=None,
        admin_username="real-admin",
        admin_password="real-secret",
    )
    monkeypatch.setattr(auth, "settings", production_settings)

    demo_admin = client.post("/auth/login", json={"username": "admin", "password": "admin123"})
    real_admin = client.post("/auth/login", json={"username": "real-admin", "password": "real-secret"})
    production_config = client.get("/auth/config")

    assert demo_admin.status_code == 401
    assert real_admin.status_code == 200
    assert real_admin.json()["data"]["user"]["roles"] == ["admin"]
    assert production_config.status_code == 200
    assert production_config.json()["data"]["demo_auth_enabled"] is False
    assert production_config.json()["data"]["demo_accounts"] == []


def test_production_account_config_and_api_token_gate(monkeypatch, tmp_path) -> None:
    production_settings = SimpleNamespace(
        app_env="production",
        demo_auth_enabled=False,
        admin_username="root-admin",
        admin_password="RootPass12345",
        admin_team_id="north-team-01",
        auth_users_path=str(tmp_path / "users.json"),
        jwt_secret="jwt-secret-for-production-test-12345",
        jwt_expire_minutes=60,
    )
    monkeypatch.setattr(auth, "settings", production_settings)
    monkeypatch.setattr(account_store, "settings", production_settings)
    monkeypatch.setattr(security, "settings", production_settings)
    monkeypatch.setattr(main_module, "settings", production_settings)
    production_client = TestClient(main_module.create_app())

    admin_login = production_client.post(
        "/auth/login",
        json={"username": "root-admin", "password": "RootPass12345"},
    )
    assert admin_login.status_code == 200
    admin_token = admin_login.json()["data"]["access_token"]
    admin_headers = {"Authorization": f"bearer {admin_token}"}

    assert production_client.get("/local-test/summary").status_code == 401
    assert production_client.get("/auth/users").status_code == 401

    users = production_client.get("/auth/users", headers=admin_headers)
    assert users.status_code == 200
    assert {item["username"] for item in users.json()["data"]["items"]} == {"root-admin"}

    created = production_client.post(
        "/auth/users",
        headers=admin_headers,
        json={
            "username": "reviewer-a",
            "password": "ReviewPass12345",
            "name": "Reviewer A",
            "roles": ["reviewer"],
            "team_id": "north-team-01",
            "status": "active",
        },
    )
    assert created.status_code == 200
    reviewer_login = production_client.post(
        "/auth/login",
        headers={"x-forwarded-for": "10.0.0.5", "user-agent": "reviewer-agent"},
        json={"username": "reviewer-a", "password": "ReviewPass12345", "team_id": "other-team"},
    )
    assert reviewer_login.status_code == 200
    assert reviewer_login.json()["data"]["team_id"] == "north-team-01"
    reviewer_token = reviewer_login.json()["data"]["access_token"]

    summary = production_client.get(
        "/local-test/summary",
        headers={"Authorization": f"bearer {reviewer_token}", "X-Team-Id": "other-team"},
    )
    assert summary.status_code == 200
    assert summary.json()["data"]["summary"]["team_id"] == "north-team-01"

    users_after_login = production_client.get("/auth/users", headers=admin_headers)
    reviewer_user = next(item for item in users_after_login.json()["data"]["items"] if item["username"] == "reviewer-a")
    assert reviewer_user["last_login_ip"] == "10.0.0.5"
    assert len(reviewer_user["login_history"]) == 1
    assert reviewer_user["login_history"][0]["ip"] == "10.0.0.5"
    assert reviewer_user["login_history"][0]["device"] == "reviewer-agent"
    assert reviewer_user["login_history"][0]["at"]

    deleted = production_client.delete("/auth/users/reviewer-a", headers=admin_headers)
    assert deleted.status_code == 200
    assert deleted.json()["data"]["user"]["username"] == "reviewer-a"
    users_after_delete = production_client.get("/auth/users", headers=admin_headers)
    assert {item["username"] for item in users_after_delete.json()["data"]["items"]} == {"root-admin"}
    deleted_login = production_client.post(
        "/auth/login",
        json={"username": "reviewer-a", "password": "ReviewPass12345"},
    )
    assert deleted_login.status_code == 401
    delete_self = production_client.delete("/auth/users/root-admin", headers=admin_headers)
    assert delete_self.status_code == 400


def test_account_login_history_keeps_30_rows_and_marks_ip_common_user(monkeypatch, tmp_path) -> None:
    production_settings = SimpleNamespace(
        app_env="production",
        demo_auth_enabled=False,
        admin_username="root-admin",
        admin_password="RootPass12345",
        admin_team_id="north-team-01",
        auth_users_path=str(tmp_path / "users.json"),
        jwt_secret="jwt-secret-for-production-test-12345",
        jwt_expire_minutes=60,
    )
    monkeypatch.setattr(auth, "settings", production_settings)
    monkeypatch.setattr(account_store, "settings", production_settings)
    monkeypatch.setattr(security, "settings", production_settings)
    monkeypatch.setattr(main_module, "settings", production_settings)
    production_client = TestClient(main_module.create_app())

    admin_login = production_client.post(
        "/auth/login",
        json={"username": "root-admin", "password": "RootPass12345"},
    )
    admin_headers = {"Authorization": f"bearer {admin_login.json()['data']['access_token']}"}
    for username in ["reviewer-a", "reviewer-b"]:
        created = production_client.post(
            "/auth/users",
            headers=admin_headers,
            json={
                "username": username,
                "password": "ReviewPass12345",
                "name": username,
                "roles": ["reviewer"],
                "team_id": "north-team-01",
                "status": "active",
            },
        )
        assert created.status_code == 200

    for index in range(35):
        response = production_client.post(
            "/auth/login",
            headers={"x-forwarded-for": "10.0.0.8", "user-agent": f"reviewer-a-agent-{index}"},
            json={"username": "reviewer-a", "password": "ReviewPass12345"},
        )
        assert response.status_code == 200

    reviewer_b_login = production_client.post(
        "/auth/login",
        headers={"x-forwarded-for": "10.0.0.8", "user-agent": "reviewer-b-agent"},
        json={"username": "reviewer-b", "password": "ReviewPass12345"},
    )
    assert reviewer_b_login.status_code == 200

    users = production_client.get("/auth/users", headers=admin_headers).json()["data"]["items"]
    by_username = {item["username"]: item for item in users}
    reviewer_a_history = by_username["reviewer-a"]["login_history"]
    reviewer_b_history = by_username["reviewer-b"]["login_history"]

    assert len(reviewer_a_history) == 30
    assert reviewer_a_history[0]["ip"] == "10.0.0.8"
    assert reviewer_a_history[0]["device"] == "reviewer-a-agent-34"
    assert reviewer_b_history[0]["ip_common_user"] == "reviewer-a"
    assert reviewer_b_history[0]["ip_common_user_count"] == 30
    assert reviewer_b_history[0]["ip_login_count"] == 31


def test_api_docs_are_disabled_in_production(monkeypatch) -> None:
    monkeypatch.setattr(main_module, "settings", SimpleNamespace(app_env="production"))
    production_client = TestClient(main_module.create_app())

    assert production_client.get("/docs").status_code == 404
    assert production_client.get("/redoc").status_code == 404
    assert production_client.get("/openapi.json").status_code == 404


def test_project_routes_return_contract_shape() -> None:
    response = client.get("/projects")

    assert response.status_code == 200
    assert_success_shape(response.json())


def test_task_claim_placeholder_uses_contract_shape() -> None:
    response = client.post("/tasks/12/claim")

    assert response.status_code == 200
    payload = response.json()
    assert_success_shape(payload)
    assert payload["data"]["task_id"] == 12
    assert payload["data"]["status"] == "claimed"


def test_validation_error_uses_contract_shape() -> None:
    response = client.get("/tasks/not-an-int")

    assert response.status_code == 422
    payload = response.json()
    assert payload["data"] is None
    assert payload["error"]["code"] == "validation_error"
    assert isinstance(payload["request_id"], str)


def test_local_test_task_and_review_flow() -> None:
    client.post("/local-test/bootstrap")

    tasks_response = client.get("/local-test/tasks")
    assert tasks_response.status_code == 200
    tasks = tasks_response.json()["data"]["items"]
    status_response = client.get("/local-test/tasks/status")
    assert status_response.status_code == 200
    task_status = status_response.json()["data"]
    assert task_status["version"]
    assert task_status["total"] == len(tasks)
    assert "address_search_text" not in task_status
    task = next(item for item in tasks if item["can_claim"])
    assert "address" in task
    assert "address_search_text" in task
    assert isinstance(task["address_search_text"], str)

    claim_response = client.post(f"/local-test/tasks/{task['id']}/claim", json={"reviewer": "api-test"})
    assert claim_response.status_code == 200
    assert claim_response.json()["data"]["claimed_by"] == "api-test"

    groups_response = client.get(f"/local-test/tasks/{task['id']}/groups?limit=1")
    assert groups_response.status_code == 200
    group = groups_response.json()["data"]["items"][0]
    assert "photos" in group

    summary_response = client.get(f"/local-test/tasks/{task['id']}/groups?limit=1&summary=true")
    assert summary_response.status_code == 200
    summary_group = summary_response.json()["data"]["items"][0]
    assert "photos" not in summary_group
    assert "photo_count" in summary_group
    assert "reviewer" in summary_group

    review_response = client.patch(
        f"/local-test/groups/{group['id']}/review",
        json={"status": "approved", "reviewer": "api-test", "note": "api smoke"},
    )
    assert review_response.status_code == 200
    assert review_response.json()["data"]["status"] == "approved"

    photo_group = next(item for item in client.get(f"/local-test/tasks/{task['id']}/groups").json()["data"]["items"] if item["photos"])
    photo = photo_group["photos"][0]
    category_response = client.patch(
        f"/local-test/groups/{photo_group['id']}/photos/{photo['id']}/category",
        json={"category": "collector_barcode", "reviewer": "api-test"},
    )
    category_summary_response = client.patch(
        f"/local-test/groups/{photo_group['id']}/photos/{photo['id']}/category?include_group=true",
        json={"category": "module_meter", "reviewer": "api-test"},
    )
    delete_response = client.request(
        "DELETE",
        f"/local-test/groups/{photo_group['id']}/photos/{photo['id']}",
        json={"reviewer": "api-test"},
    )
    assert category_response.status_code == 200
    assert category_response.json()["data"]["category"] == "collector_barcode"
    assert category_summary_response.status_code == 200
    assert category_summary_response.json()["data"]["photo"]["category"] == "module_meter"
    assert category_summary_response.json()["data"]["group"]["id"] == photo_group["id"]
    assert "photos" not in category_summary_response.json()["data"]["group"]
    assert delete_response.status_code == 200
    assert delete_response.json()["data"]["deleted_photo"]["id"] == photo["id"]


def test_local_test_team_header_isolates_review_state() -> None:
    team_a = {"X-Team-Id": "team-a"}
    team_b = {"X-Team-Id": "team-b"}

    client.post("/local-test/bootstrap", headers=team_a)
    client.post("/local-test/bootstrap", headers=team_b)
    task_a = next(item for item in client.get("/local-test/tasks", headers=team_a).json()["data"]["items"] if item["can_claim"])
    task_b = next(item for item in client.get("/local-test/tasks", headers=team_b).json()["data"]["items"] if item["id"] == task_a["id"])

    claim_a = client.post(f"/local-test/tasks/{task_a['id']}/claim", headers=team_a, json={"reviewer": "alice"})
    after_a = client.get("/local-test/tasks", headers=team_a).json()["data"]["items"]
    after_b = client.get("/local-test/tasks", headers=team_b).json()["data"]["items"]
    teams = client.get("/local-test/teams", headers=team_a).json()["data"]["items"]

    assert claim_a.status_code == 200
    assert next(item for item in after_a if item["id"] == task_a["id"])["claimed_by"] == "alice"
    assert task_b["claimed_by"] is None
    assert next(item for item in after_b if item["id"] == task_a["id"])["claimed_by"] is None
    assert {"team-a", "team-b"}.issubset({item["team_id"] for item in teams})


def test_async_scan_template_import_job_completes() -> None:
    team = {"X-Team-Id": f"async-import-team-{uuid4().hex}"}
    catalog = build_api_workbook(
        [
            ["terminal", "meter_no", "address"],
            ["T-ASYNC", "ZZ0000001001", "Async road"],
        ]
    )
    scan = build_api_workbook(
        [
            ["barcode", "meter_match_key", "terminal", "collector", "module_asset_no", "photo_urls"],
            ["scan-1", "0000001001", "T-ASYNC", "collector", "asset-1", "https://example.invalid/1.jpg"],
            ["scan-2", "0000001001", "T-ASYNC", "collector", "asset-2", "https://example.invalid/2.jpg"],
        ]
    )

    catalog_response = client.post(
        "/local-test/catalog/total/import-xlsx",
        headers=team,
        files={"file": ("catalog.xlsx", catalog, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    job_response = client.post(
        "/local-test/scan/import-template-xlsx/jobs",
        headers=team,
        files={"file": ("scan.xlsx", scan, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert catalog_response.status_code == 200
    assert job_response.status_code == 200
    job_id = job_response.json()["data"]["job_id"]
    payload = {}
    for _ in range(50):
        poll = client.get(f"/local-test/scan/import-template-xlsx/jobs/{job_id}", headers=team)
        assert poll.status_code == 200
        payload = poll.json()["data"]
        if payload["status"] == "complete":
            break
        time.sleep(0.05)

    assert payload["status"] == "complete"
    assert payload["result"]["template_rows"] == 2
    assert payload["result"]["applied_records"] == 2
    assert payload["progress"]["phase"] == "complete"


def test_task_hall_page_is_available() -> None:
    assert_vue_shell_response(client.get("/task-hall"))
    assert_vue_shell_response(client.get("/task-hall?embedded=1"))

def test_app_shell_page_is_available() -> None:
    assert_vue_shell_response(client.get("/app"))

def test_claim_tasks_page_exposes_admin_release_all_control() -> None:
    assert_vue_shell_response(client.get("/claim-tasks"))
    assert_vue_shell_response(client.get("/claim-tasks?embedded=1"))

def test_admin_can_release_all_claimed_tasks() -> None:
    admin_login = client.post("/auth/login", json={"username": "admin", "password": "admin123"})
    reviewer_login = client.post("/auth/login", json={"username": "reviewer", "password": "review123"})
    admin_headers = {"Authorization": f"bearer {admin_login.json()['data']['access_token']}"}
    reviewer_headers = {"Authorization": f"bearer {reviewer_login.json()['data']['access_token']}"}

    client.post("/local-test/bootstrap", headers=admin_headers)
    tasks = client.get("/local-test/tasks", headers=admin_headers).json()["data"]["items"]
    claimable = [task for task in tasks if task["can_claim"]][:2]
    assert claimable
    for index, task in enumerate(claimable):
        claim = client.post(
            f"/local-test/tasks/{task['id']}/claim",
            headers=admin_headers,
            json={"reviewer": f"reviewer-{index}"},
        )
        assert claim.status_code == 200

    reviewer_denied = client.post(
        "/local-test/tasks/release-all",
        headers=reviewer_headers,
        json={"reviewer": "reviewer"},
    )
    released = client.post(
        "/local-test/tasks/release-all",
        headers=admin_headers,
        json={"reviewer": "admin"},
    )

    assert reviewer_denied.status_code == 403
    assert released.status_code == 200
    assert released.json()["data"]["released"] == len(claimable)
    after = client.get("/local-test/tasks", headers=admin_headers).json()["data"]["items"]
    claimed_ids = {task["id"] for task in claimable}
    assert not any(task.get("claimed_by") for task in after if task["id"] in claimed_ids)


def test_construction_task_open_claim_and_upload_batch() -> None:
    admin_login = client.post("/auth/login", json={"username": "admin", "password": "admin123"})
    reviewer_login = client.post("/auth/login", json={"username": "reviewer", "password": "review123"})
    constructor_login = client.post("/auth/login", json={"username": "constructor", "password": "construct123"})
    constructor_name = constructor_login.json()["data"]["user"]["name"]
    admin_headers = {"Authorization": f"bearer {admin_login.json()['data']['access_token']}"}
    reviewer_headers = {"Authorization": f"bearer {reviewer_login.json()['data']['access_token']}"}
    constructor_headers = {"Authorization": f"bearer {constructor_login.json()['data']['access_token']}"}

    client.post("/local-test/bootstrap", headers=admin_headers)
    client.post("/local-test/scan/clear", headers=admin_headers)
    task = client.get("/local-test/tasks", headers=admin_headers).json()["data"]["items"][0]
    hidden = client.get(
        "/local-test/construction/tasks?actor=constructor",
        headers=constructor_headers,
    ).json()["data"]["items"]
    denied_claim = client.post(
        f"/local-test/construction/tasks/{task['id']}/claim",
        headers=constructor_headers,
        json={"actor": "constructor"},
    )
    opened = client.patch(
        f"/local-test/construction/tasks/{task['id']}/open",
        headers=admin_headers,
        json={"actor": "admin"},
    )
    assigned = client.patch(
        f"/local-test/construction/tasks/{task['id']}/assign",
        headers=admin_headers,
        json={"actor": "admin", "constructor": "constructor"},
    )
    claimed = client.post(
        f"/local-test/construction/tasks/{task['id']}/claim",
        headers=constructor_headers,
        json={"actor": "constructor"},
    )
    groups = client.get(
        f"/local-test/construction/tasks/{task['id']}/groups?limit=1&summary=true",
        headers=constructor_headers,
    ).json()["data"]["items"]
    group = groups[0]
    missing_module_upload = client.post(
        f"/local-test/construction/groups/{group['id']}/upload-batch",
        headers=constructor_headers,
        data={
            "actor": "constructor",
            "client_batch_id": "batch-api-missing-module",
            "client_completed_at": "bad-client-time",
            "collector": "collector-api",
            "module_asset_no": "module-api",
            "photo_slots": ["before_box", "after_box"],
            "client_photo_ids": ["photo-a", "photo-b"],
        },
        files=[
            ("files", ("before.jpg", b"image-before", "image/jpeg")),
            ("files", ("after.jpg", b"image-after", "image/jpeg")),
        ],
    )
    uploaded = client.post(
        f"/local-test/construction/groups/{group['id']}/upload-batch",
        headers=constructor_headers,
        data={
            "actor": "constructor",
            "client_batch_id": "batch-api-1",
            "client_completed_at": "bad-client-time",
            "collector": "collector-api",
            "module_asset_no": "module-api",
            "photo_slots": ["before_box", "module_meter", "after_box"],
            "client_photo_ids": ["photo-a", "photo-b", "photo-c"],
        },
        files=[
            ("files", ("before.jpg", b"image-before-1", "image/jpeg")),
            ("files", ("meter.jpg", b"image-meter-1", "image/jpeg")),
            ("files", ("after.jpg", b"image-after-1", "image/jpeg")),
        ],
    )

    assert hidden == []
    assert denied_claim.status_code == 400
    assert opened.status_code == 200
    assert opened.json()["data"]["construction_enabled"] is True
    assert assigned.status_code == 200
    assert assigned.json()["data"]["construction_claimed_by"] == "constructor"
    assert claimed.status_code == 200
    assert claimed.json()["data"]["construction_claimed_by"] == "constructor"
    assert missing_module_upload.status_code == 400
    assert "模块与电能表" in missing_module_upload.json()["detail"]
    assert uploaded.status_code == 200
    payload = uploaded.json()["data"]
    assert payload["added"] == 3
    assert payload["skipped_duplicates"] == 0
    assert payload["group"]["status"] == "exception"
    assert payload["group"]["exception_note"] == "缺采集器照片"
    assert "missing_collector_photo" in payload["group"]["exception_reasons"]
    assert payload["group"]["photos"][0]["upload_source"] == "construction-mobile"
    assert payload["group"]["photos"][0]["construction_slot"] == "before_box"
    assert payload["group"]["photos"][0]["category"] == "unclassified"
    assert payload["group"]["photos"][0]["creator"] == constructor_name
    assert payload["group"]["photos"][0]["creator"] != "constructor"
    assert payload["group"]["photos"][0]["sha256"]
    assert payload["group"]["photos"][0]["storage_type"] == "local_upload"
    assert payload["group"]["photos"][0]["storage_key"].startswith("construction/")
    assert payload["uploaded_urls"][0].startswith("/static/uploads/construction/")

    after_first_upload = client.get(
        f"/local-test/construction/tasks/{task['id']}/groups?limit=1000&summary=true",
        headers=constructor_headers,
    ).json()["data"]["items"]
    assert group["id"] not in {item["id"] for item in after_first_upload}

    collector_upload = client.post(
        f"/local-test/construction/groups/{group['id']}/upload-batch",
        headers=constructor_headers,
        data={
            "actor": "constructor",
            "client_batch_id": "batch-api-collector-fix",
            "client_completed_at": "2026-06-08T09:10:00",
            "collector": "collector-api",
            "module_asset_no": "module-api",
            "photo_slots": ["collector_barcode"],
            "client_photo_ids": ["photo-collector"],
        },
        files=[
            ("files", ("collector.jpg", b"image-collector-1", "image/jpeg")),
        ],
    )
    assert collector_upload.status_code == 200
    collector_payload = collector_upload.json()["data"]
    assert collector_payload["group"]["status"] == "pending"
    assert collector_payload["group"]["exception_note"] == ""
    assert "missing_collector_photo" not in collector_payload["group"]["exception_reasons"]

    second_group = after_first_upload[0]
    complete_upload = client.post(
        f"/local-test/construction/groups/{second_group['id']}/upload-batch",
        headers=constructor_headers,
        data={
            "actor": "constructor",
            "client_batch_id": "batch-api-2",
            "client_completed_at": "2026-06-08T09:30:00",
            "collector": "collector-api-2",
            "module_asset_no": "module-api-2",
            "photo_slots": ["before_box", "after_box", "module_meter", "collector_barcode"],
            "client_photo_ids": ["photo-1", "photo-2", "photo-3", "photo-4"],
        },
        files=[
            ("files", ("before.jpg", b"image-1", "image/jpeg")),
            ("files", ("after.jpg", b"image-2", "image/jpeg")),
            ("files", ("meter.jpg", b"image-3", "image/jpeg")),
            ("files", ("collector.jpg", b"image-4", "image/jpeg")),
        ],
    )
    assert complete_upload.status_code == 200
    complete_payload = complete_upload.json()["data"]
    assert complete_payload["added"] == 4
    assert complete_payload["group"]["status"] == "pending"
    workload = client.get(f"/local-test/installers/{constructor_name}/daily-workload", headers=admin_headers)
    assert workload.status_code == 200
    client_day = next(item for item in workload.json()["data"]["items"] if item["date"] == "2026-06-08")
    assert client_day["group_count"] == 1
    assert client_day["completion_count"] == 1
    assert client_day["start_time"] == "09:30"
    assert client_day["end_time"] == "09:30"

    client.post(
        f"/local-test/tasks/{task['id']}/claim",
        headers=reviewer_headers,
        json={"reviewer": "reviewer"},
    )
    review_groups = client.get(
        f"/local-test/tasks/{task['id']}/groups?limit=1000&scan_only=false&summary=true",
        headers=reviewer_headers,
    ).json()["data"]["items"]
    assert second_group["id"] in {item["id"] for item in review_groups}
    review_detail = client.get(f"/local-test/groups/{second_group['id']}", headers=reviewer_headers).json()["data"]
    assert len(review_detail["photos"]) == 4
    assert all(photo["image_url"].startswith("/static/uploads/construction/") for photo in review_detail["photos"])
    assert all(photo["download_status"] == "downloaded" for photo in review_detail["photos"])
    assert all(photo["category"] == "unclassified" for photo in review_detail["photos"])

    repaired_detail = client.get(f"/local-test/groups/{group['id']}", headers=reviewer_headers).json()["data"]
    collector_photo = next(photo for photo in repaired_detail["photos"] if photo["construction_slot"] == "collector_barcode")
    deleted_collector = client.request(
        "DELETE",
        f"/local-test/groups/{group['id']}/photos/{collector_photo['id']}",
        headers=reviewer_headers,
        json={"reviewer": "reviewer"},
    )
    assert deleted_collector.status_code == 200

    deleted_group = deleted_collector.json()["data"]["group"]
    assert deleted_group["status"] == "exception"
    assert deleted_group["exception_note"] == "缺采集器照片"
    assert "missing_collector_photo" in deleted_group["exception_reasons"]
    exception_groups = client.get("/local-test/exception-groups", headers=reviewer_headers).json()["data"]["items"]
    assert group["id"] in {item["id"] for item in exception_groups}

    released = client.post(
        f"/local-test/construction/tasks/{task['id']}/release",
        headers=constructor_headers,
        json={"actor": "constructor"},
    )
    assert released.status_code == 200
    assert released.json()["data"]["construction_claimed_by"] in (None, "")
    after_release = client.get(
        "/local-test/construction/tasks?actor=constructor",
        headers=constructor_headers,
    ).json()["data"]["items"]
    assert task["id"] not in {item["id"] for item in after_release}


def test_construction_online_events_feed_fused_installer_workload() -> None:
    admin_login = client.post("/auth/login", json={"username": "admin", "password": "admin123"})
    admin_headers = {"Authorization": f"bearer {admin_login.json()['data']['access_token']}"}
    client.post(
        "/auth/users",
        headers=admin_headers,
        json={
            "username": "constructor",
            "password": "construct123",
            "name": "施工员",
            "roles": ["constructor"],
            "team_id": "demo-team",
            "status": "active",
        },
    )
    constructor_login = client.post("/auth/login", json={"username": "constructor", "password": "construct123"})
    constructor_name = constructor_login.json()["data"]["user"]["name"]
    constructor_headers = {"Authorization": f"bearer {constructor_login.json()['data']['access_token']}"}

    client.post("/local-test/bootstrap", headers=admin_headers)
    client.post("/local-test/scan/clear", headers=admin_headers)
    task = client.get("/local-test/tasks", headers=admin_headers).json()["data"]["items"][0]
    client.patch(
        f"/local-test/construction/tasks/{task['id']}/assign",
        headers=admin_headers,
        json={"actor": "admin", "constructor": "constructor"},
    )
    groups = client.get(
        f"/local-test/construction/tasks/{task['id']}/groups?limit=1000&summary=true",
        headers=constructor_headers,
    ).json()["data"]["items"]

    current = datetime(2026, 6, 8, 9, 0)
    spoofed = client.post(
        "/local-test/construction/heartbeat",
        headers=constructor_headers,
        json={"actor": "other-installer", "task_id": task["id"], "occurred_at": current.isoformat()},
    )
    assert spoofed.status_code == 403
    while current <= datetime(2026, 6, 8, 11, 0):
        heartbeat = client.post(
            "/local-test/construction/heartbeat",
            headers=constructor_headers,
            json={"actor": "constructor", "task_id": task["id"], "occurred_at": current.isoformat()},
        )
        assert heartbeat.status_code == 200
        current = current + local_simulation.timedelta(minutes=5)

    draft_done = client.post(
        "/local-test/construction/non-idle-events",
        headers=constructor_headers,
        json={
            "event_type": "group_draft_completed",
            "actor": "constructor",
            "task_id": task["id"],
            "group_id": groups[0]["id"],
            "client_batch_id": "deleted-draft",
            "occurred_at": "2026-06-08T09:20:00",
        },
    )
    draft_deleted = client.post(
        "/local-test/construction/non-idle-events",
        headers=constructor_headers,
        json={
            "event_type": "group_draft_deleted",
            "actor": "constructor",
            "task_id": task["id"],
            "group_id": groups[0]["id"],
            "client_batch_id": "deleted-draft",
            "occurred_at": "2026-06-08T09:25:00",
        },
    )
    assert draft_done.status_code == 200
    assert draft_deleted.status_code == 200

    for index, completed_at in enumerate(("2026-06-08T09:30:00", "2026-06-08T10:10:00")):
        upload = client.post(
            f"/local-test/construction/groups/{groups[index]['id']}/upload-batch",
            headers=constructor_headers,
            data={
                "actor": "constructor",
                "client_batch_id": f"online-kpi-batch-{uuid4().hex}-{index}",
                "client_completed_at": completed_at,
                "collector": f"collector-{index}",
                "module_asset_no": f"module-{index}",
                "photo_slots": ["before_box", "after_box", "module_meter", "collector_barcode"],
                "client_photo_ids": [f"before-{index}", f"after-{index}", f"meter-{index}", f"collector-{index}"],
            },
            files=[
                ("files", ("before.jpg", f"before-{index}".encode(), "image/jpeg")),
                ("files", ("after.jpg", f"after-{index}".encode(), "image/jpeg")),
                ("files", ("meter.jpg", f"meter-{index}".encode(), "image/jpeg")),
                ("files", ("collector.jpg", f"collector-{index}".encode(), "image/jpeg")),
            ],
        )
        assert upload.status_code == 200

    workload = client.get(f"/local-test/installers/{constructor_name}/daily-workload", headers=admin_headers)

    assert workload.status_code == 200
    item = next(row for row in workload.json()["data"]["items"] if row["date"] == "2026-06-08")
    assert item["attendance_window_minutes"] == 480
    assert item["countable_online_minutes"] == 120
    assert item["base_online_coefficient"] == 1.06
    assert item["idle_penalty_coefficient"] == 0
    assert item["final_online_coefficient"] == 1.06
    assert item["pending_non_idle_count"] == 0
    assert item["confirmed_non_idle_count"] == 2
    assert item["fused_work_duration_minutes"] >= item["work_duration_minutes"]
    assert item["fused_work_duration_minutes"] <= item["attendance_window_minutes"]
    assert item["fused_efficiency_duration_minutes"] <= item["attendance_window_minutes"]
    assert item["fused_efficiency_duration_minutes"] <= item["fused_work_duration_minutes"]


def test_construction_upload_rejects_placeholder_group_id_before_file_save() -> None:
    admin_login = client.post("/auth/login", json={"username": "admin", "password": "admin123"})
    constructor_login = client.post("/auth/login", json={"username": "constructor", "password": "construct123"})
    admin_headers = {"Authorization": f"bearer {admin_login.json()['data']['access_token']}"}
    constructor_headers = {"Authorization": f"bearer {constructor_login.json()['data']['access_token']}"}

    client.post("/local-test/bootstrap", headers=admin_headers)
    task = client.get("/local-test/tasks", headers=admin_headers).json()["data"]["items"][0]
    client.patch(
        f"/local-test/construction/tasks/{task['id']}/open",
        headers=admin_headers,
        json={"actor": "admin"},
    )
    client.patch(
        f"/local-test/construction/tasks/{task['id']}/assign",
        headers=admin_headers,
        json={"actor": "admin", "constructor": "constructor"},
    )
    upload_dir = Path("v2-api/app/static/uploads/construction")

    def saved_upload_files() -> set[str]:
        if not upload_dir.exists():
            return set()
        return {str(path.relative_to(upload_dir)) for path in upload_dir.rglob("*") if path.is_file()}

    before_files = saved_upload_files()
    uploaded = client.post(
        "/local-test/construction/groups/00000000/upload-batch",
        headers=constructor_headers,
        data={
            "actor": "constructor",
            "client_batch_id": "batch-placeholder-api",
            "client_completed_at": "2026-06-08T09:30:00",
            "collector": "collector-api",
            "module_asset_no": "module-api",
            "photo_slots": ["before_box", "module_meter", "after_box"],
            "client_photo_ids": ["photo-a", "photo-b", "photo-c"],
        },
        files=[
            ("files", ("before.jpg", b"placeholder-before", "image/jpeg")),
            ("files", ("meter.jpg", b"placeholder-meter", "image/jpeg")),
            ("files", ("after.jpg", b"placeholder-after", "image/jpeg")),
        ],
    )

    assert uploaded.status_code == 400
    assert "00000000" in uploaded.json()["detail"]
    assert saved_upload_files() == before_files

    group = client.get(
        f"/local-test/construction/tasks/{task['id']}/groups?limit=1&summary=true",
        headers=constructor_headers,
    ).json()["data"]["items"][0]
    client.patch(
        f"/local-test/groups/{group['id']}/metadata",
        headers=admin_headers,
        json={"actor": "admin", "updates": {"meter_no": "00000000"}},
    )

    before_files = saved_upload_files()
    placeholder_meter_upload = client.post(
        f"/local-test/construction/groups/{group['id']}/upload-batch",
        headers=constructor_headers,
        data={
            "actor": "constructor",
            "client_batch_id": "batch-placeholder-meter",
            "client_completed_at": "2026-06-08T09:40:00",
            "collector": "collector-api",
            "module_asset_no": "module-api",
            "photo_slots": ["before_box", "module_meter", "after_box"],
            "client_photo_ids": ["photo-a", "photo-b", "photo-c"],
        },
        files=[
            ("files", ("before.jpg", b"placeholder-meter-before", "image/jpeg")),
            ("files", ("meter.jpg", b"placeholder-meter-meter", "image/jpeg")),
            ("files", ("after.jpg", b"placeholder-meter-after", "image/jpeg")),
        ],
    )

    assert placeholder_meter_upload.status_code == 400
    assert "00000000" in placeholder_meter_upload.json()["detail"]
    assert saved_upload_files() == before_files


def test_exception_group_assignment_is_visible_to_constructor() -> None:
    admin_login = client.post("/auth/login", json={"username": "admin", "password": "admin123"})
    reviewer_login = client.post("/auth/login", json={"username": "reviewer", "password": "review123"})
    constructor_login = client.post("/auth/login", json={"username": "constructor", "password": "construct123"})
    admin_headers = {"Authorization": f"bearer {admin_login.json()['data']['access_token']}"}
    reviewer_headers = {"Authorization": f"bearer {reviewer_login.json()['data']['access_token']}"}
    constructor_headers = {"Authorization": f"bearer {constructor_login.json()['data']['access_token']}"}

    client.post("/local-test/bootstrap", headers=admin_headers)
    task = next(item for item in client.get("/local-test/tasks", headers=admin_headers).json()["data"]["items"] if item["can_claim"])
    claim = client.post(f"/local-test/tasks/{task['id']}/claim", headers=reviewer_headers, json={"reviewer": "reviewer"})
    assert claim.status_code == 200
    group = client.get(
        f"/local-test/tasks/{task['id']}/groups?limit=1&summary=true",
        headers=reviewer_headers,
    ).json()["data"]["items"][0]

    returned = client.patch(
        f"/local-test/groups/{group['id']}/return-exception",
        headers=reviewer_headers,
        json={"actor": "reviewer", "category": "照片缺失", "note": "现场补缺失照片"},
    )
    assert returned.status_code == 200
    order = returned.json()["data"]["order"]
    assigned = client.patch(
        f"/local-test/construction/exception-orders/{order['id']}/assign",
        headers=admin_headers,
        json={"actor": "admin", "constructor": "constructor", "note": "补采异常资料组"},
    )
    constructor_orders = client.get(
        "/local-test/construction/exception-orders?actor=constructor",
        headers=constructor_headers,
    ).json()["data"]["items"]
    constructor_tasks = client.get(
        "/local-test/construction/tasks?actor=constructor",
        headers=constructor_headers,
    ).json()["data"]["items"]

    assert assigned.status_code == 200
    assert assigned.json()["data"]["order"]["assigned_to"] == "constructor"
    assert order["id"] in {item["id"] for item in constructor_orders}
    assert str(group["id"]) in {str(item["group_id"]) for item in constructor_orders}
    assert task["id"] in {item["id"] for item in constructor_tasks}


def test_constructor_can_keep_up_to_five_assigned_terminals() -> None:
    admin_login = client.post("/auth/login", json={"username": "admin", "password": "admin123"})
    constructor_login = client.post("/auth/login", json={"username": "constructor", "password": "construct123"})
    admin_headers = {"Authorization": f"bearer {admin_login.json()['data']['access_token']}"}
    constructor_headers = {"Authorization": f"bearer {constructor_login.json()['data']['access_token']}"}

    client.post("/local-test/bootstrap", headers=admin_headers)
    tasks = client.get("/local-test/tasks", headers=admin_headers).json()["data"]["items"][:6]
    assert len(tasks) == 6
    for task in tasks:
        opened = client.patch(
            f"/local-test/construction/tasks/{task['id']}/open",
            headers=admin_headers,
            json={"actor": "admin"},
        )
        assert opened.status_code == 200

    allowed_tasks = tasks[:5]
    denied_task = tasks[5]
    assigned_responses = [
        client.patch(
            f"/local-test/construction/tasks/{task['id']}/assign",
            headers=admin_headers,
            json={"actor": "admin", "constructor": "constructor"},
        )
        for task in allowed_tasks
    ]
    claimed = client.post(
        f"/local-test/construction/tasks/{allowed_tasks[0]['id']}/claim",
        headers=constructor_headers,
        json={"actor": "constructor"},
    )
    visible = client.get(
        "/local-test/construction/tasks?actor=constructor",
        headers=constructor_headers,
    )
    denied = client.post(
        f"/local-test/construction/tasks/{denied_task['id']}/claim",
        headers=constructor_headers,
        json={"actor": "constructor"},
    )
    denied_assign = client.patch(
        f"/local-test/construction/tasks/{denied_task['id']}/assign",
        headers=admin_headers,
        json={"actor": "admin", "constructor": "constructor"},
    )

    assert all(response.status_code == 200 for response in assigned_responses)
    assert claimed.status_code == 200
    assert {task["id"] for task in visible.json()["data"]["items"]} == {task["id"] for task in allowed_tasks}
    assert denied.status_code == 400
    assert "assigned by an administrator" in denied.json()["detail"]
    assert denied_assign.status_code == 400
    assert "already has 5 active terminals" in denied_assign.json()["detail"]


def test_construction_tasks_include_meter_search_text_for_task_picker() -> None:
    admin_login = client.post("/auth/login", json={"username": "admin", "password": "admin123"})
    admin_headers = {"Authorization": f"bearer {admin_login.json()['data']['access_token']}"}

    client.post("/local-test/bootstrap", headers=admin_headers)
    task = client.get("/local-test/tasks", headers=admin_headers).json()["data"]["items"][0]
    groups = client.get(
        f"/local-test/tasks/{task['id']}/groups?limit=1000&scan_only=false&summary=true",
        headers=admin_headers,
    ).json()["data"]["items"]
    target_meter = next(item["meter_no"] for item in groups if item.get("meter_no"))

    assigned = client.patch(
        f"/local-test/construction/tasks/{task['id']}/assign",
        headers=admin_headers,
        json={"actor": "admin", "constructor": "constructor"},
    )
    assert assigned.status_code == 200

    construction_tasks = client.get(
        "/local-test/construction/tasks?actor=constructor",
        headers=admin_headers,
    )
    matching_task = next(item for item in construction_tasks.json()["data"]["items"] if item["id"] == task["id"])

    assert target_meter in matching_task["meter_search_text"]


def test_direct_workspace_routes_redirect_to_app_shell() -> None:
    for path in ["/project-board", "/claim-tasks", "/task-hall", "/construction", "/account-management", "/sync-config"]:
        assert_vue_shell_response(client.get(path, follow_redirects=False))
    response = client.get("/construction-cache", follow_redirects=False)
    assert response.status_code == 307
    assert response.headers["location"] == "/construction"
    response = client.get("/unmatched", follow_redirects=False)
    assert response.status_code == 307
    assert response.headers["location"] == "/task-hall"

def test_project_board_page_is_available() -> None:
    assert_vue_shell_response(client.get("/project-board"))
    assert_vue_shell_response(client.get("/project-board?embedded=1"))

def test_static_page_verifier_rejects_visible_mojibake(tmp_path, monkeypatch) -> None:
    verifier = load_static_page_verifier()
    page = tmp_path / "bad.html"
    page.write_text("<!doctype html><html><body>妞ゅ湱娲伴惇瀣緲 濡炪倕婀卞ú?/body></html>", encoding="utf-8")
    monkeypatch.setattr(verifier, "STATIC_ROOT", tmp_path)

    try:
        verifier.verify_page("bad.html", ["妞ゅ湱娲伴惇瀣緲"], node=None)
    except AssertionError as exc:
        assert "mojibake fragment" in str(exc)
    else:
        raise AssertionError("visible mojibake should be rejected")


def test_legacy_static_html_pages_are_not_served() -> None:
    for path in [
        "/static/app_shell.html",
        "/static/login.html",
        "/static/project_board.html",
        "/static/claim_tasks.html",
        "/static/task_hall.html",
        "/static/construction.html",
        "/static/construction_cache.html",
        "/static/unmatched.html",
        "/static/sync_config.html",
    ]:
        response = client.get(path)
        assert response.status_code == 404


def test_demo_review_image_assets_are_available() -> None:
    for index in range(1, 5):
        response = client.get(f"/static/demo-assets/review-photo-{index}.svg")

        assert response.status_code == 200
        assert "image/svg+xml" in response.headers["content-type"]


def test_claim_tasks_page_is_available() -> None:
    assert_vue_shell_response(client.get("/claim-tasks"))
    assert_vue_shell_response(client.get("/claim-tasks?embedded=1"))

def test_construction_page_is_available() -> None:
    assert_vue_shell_response(client.get("/construction"))
    assert_vue_shell_response(client.get("/construction?embedded=1"))


def test_global_search_page_is_available() -> None:
    assert_vue_shell_response(client.get("/global-search"))
    assert_vue_shell_response(client.get("/global-search?embedded=1"))


def test_unmatched_page_redirects_to_review_workbench() -> None:
    response = client.get("/unmatched?embedded=1", follow_redirects=False)

    assert response.status_code == 307
    assert response.headers["location"] == "/task-hall"

def test_group_target_route_is_searchable() -> None:
    client.post("/local-test/bootstrap")

    response = client.get("/local-test/group-targets?query=350&limit=5")

    assert response.status_code == 200
    payload = response.json()["data"]
    assert "items" in payload
    assert len(payload["items"]) <= 5


def test_admin_global_group_search_is_admin_only(monkeypatch, tmp_path) -> None:
    from app.api.routes import local_test

    production_settings = SimpleNamespace(
        app_env="production",
        demo_auth_enabled=False,
        admin_username="root-admin",
        admin_password="RootPass12345",
        admin_team_id="global-search-team",
        auth_users_path=str(tmp_path / "users.json"),
        jwt_secret="jwt-secret-for-global-search-test-12345",
        jwt_expire_minutes=60,
        state_backend="json",
    )
    monkeypatch.setattr(auth, "settings", production_settings)
    monkeypatch.setattr(account_store, "settings", production_settings)
    monkeypatch.setattr(security, "settings", production_settings)
    monkeypatch.setattr(local_test, "settings", production_settings)
    monkeypatch.setattr(main_module, "settings", production_settings)
    production_client = TestClient(main_module.create_app())

    admin_login = production_client.post(
        "/auth/login",
        json={"username": "root-admin", "password": "RootPass12345"},
    )
    assert admin_login.status_code == 200
    admin_headers = {"Authorization": f"bearer {admin_login.json()['data']['access_token']}"}

    created = production_client.post(
        "/auth/users",
        headers=admin_headers,
        json={
            "username": "reviewer-a",
            "password": "ReviewPass12345",
            "name": "Reviewer A",
            "roles": ["reviewer"],
            "team_id": "global-search-team",
            "status": "active",
        },
    )
    assert created.status_code == 200
    reviewer_login = production_client.post(
        "/auth/login",
        json={"username": "reviewer-a", "password": "ReviewPass12345"},
    )
    assert reviewer_login.status_code == 200
    reviewer_headers = {"Authorization": f"bearer {reviewer_login.json()['data']['access_token']}"}

    bootstrap = production_client.post("/local-test/bootstrap", headers=admin_headers)
    assert bootstrap.status_code == 200

    forbidden = production_client.get("/groups/search?query=350&limit=5", headers=reviewer_headers)
    assert forbidden.status_code == 403
    legacy_forbidden = production_client.get("/local-test/group-targets?query=350&limit=5", headers=reviewer_headers)
    assert legacy_forbidden.status_code == 403

    blank = production_client.get("/groups/search", headers=admin_headers)
    assert blank.status_code == 200
    assert blank.json()["data"]["total"] == 0
    assert blank.json()["data"]["items"] == []

    response = production_client.get("/groups/search?query=350&limit=5", headers=admin_headers)
    assert response.status_code == 200
    legacy_response = production_client.get("/local-test/group-targets?query=350&limit=5", headers=admin_headers)
    assert legacy_response.status_code == 200
    payload = response.json()["data"]
    assert payload["total"] >= 1
    assert len(payload["items"]) <= 5
    assert {"id", "task_id", "terminal", "meter_no", "status", "photo_count"}.issubset(payload["items"][0])


def test_admin_group_backoffice_edit_and_resets_are_audited(monkeypatch, tmp_path) -> None:
    from app.api.routes import local_test

    production_settings = SimpleNamespace(
        app_env="production",
        demo_auth_enabled=False,
        admin_username="root-admin",
        admin_password="RootPass12345",
        admin_team_id="group-admin-team",
        auth_users_path=str(tmp_path / "users.json"),
        jwt_secret="jwt-secret-for-group-admin-test-12345",
        jwt_expire_minutes=60,
        state_backend="json",
    )
    monkeypatch.setattr(auth, "settings", production_settings)
    monkeypatch.setattr(account_store, "settings", production_settings)
    monkeypatch.setattr(security, "settings", production_settings)
    monkeypatch.setattr(local_test, "settings", production_settings)
    monkeypatch.setattr(main_module, "settings", production_settings)
    production_client = TestClient(main_module.create_app())

    admin_login = production_client.post(
        "/auth/login",
        json={"username": "root-admin", "password": "RootPass12345"},
    )
    assert admin_login.status_code == 200
    admin_headers = {"Authorization": f"bearer {admin_login.json()['data']['access_token']}"}

    created = production_client.post(
        "/auth/users",
        headers=admin_headers,
        json={
            "username": "reviewer-a",
            "password": "ReviewPass12345",
            "name": "Reviewer A",
            "roles": ["reviewer"],
            "team_id": "group-admin-team",
            "status": "active",
        },
    )
    assert created.status_code == 200
    reviewer_login = production_client.post(
        "/auth/login",
        json={"username": "reviewer-a", "password": "ReviewPass12345"},
    )
    assert reviewer_login.status_code == 200
    reviewer_headers = {"Authorization": f"bearer {reviewer_login.json()['data']['access_token']}"}

    assert production_client.post("/local-test/bootstrap", headers=admin_headers).status_code == 200
    group = production_client.get("/groups/search?query=350&limit=1", headers=admin_headers).json()["data"]["items"][0]

    forbidden = production_client.patch(
        f"/groups/{group['id']}/metadata",
        headers=reviewer_headers,
        json={"updates": {"address": "reviewer should not edit"}},
    )
    assert forbidden.status_code == 403
    legacy_privileged_forbidden = production_client.patch(
        f"/local-test/groups/{group['id']}/metadata",
        headers=reviewer_headers,
        json={"actor": "forged-admin", "updates": {"status": "approved", "reviewer": "forged-admin"}},
    )
    assert legacy_privileged_forbidden.status_code == 403
    legacy_terminal_forbidden = production_client.patch(
        f"/local-test/groups/{group['id']}/terminal",
        headers=reviewer_headers,
        json={"actor": "forged-admin", "terminal": "FORGED-TERM"},
    )
    assert legacy_terminal_forbidden.status_code == 403

    edited = production_client.patch(
        f"/groups/{group['id']}/metadata",
        headers=admin_headers,
        json={
            "updates": {
                "meter_no": "ADMIN-METER-001",
                "terminal": "ADMIN-TERM-001",
                "address": "admin edited address",
                "status": "approved",
                "reviewer": "manual-reviewer",
                "review_note": "manual review note",
                "exception_note": "manual exception note",
                "collector": "admin collector",
                "module_asset_no": "admin module",
            }
        },
    )
    assert edited.status_code == 200
    edited_payload = edited.json()["data"]
    edited_group = edited_payload["group"]
    assert edited_group["meter_no"] == "ADMIN-METER-001"
    assert edited_group["terminal"] == "ADMIN-TERM-001"
    assert edited_group["address"] == "admin edited address"
    assert edited_group["status"] == "approved"
    assert edited_group["reviewer"] == "manual-reviewer"
    assert edited_group["collector"] == "admin collector"
    assert edited_group["module_asset_no"] == "admin module"
    assert set(edited_payload["changed_fields"]) >= {"meter_no", "terminal", "address", "status", "reviewer"}

    searched = production_client.get("/groups/search?query=ADMIN-METER-001&limit=1", headers=admin_headers)
    assert searched.status_code == 200
    searched_group = searched.json()["data"]["items"][0]
    assert searched_group["collector"] == "admin collector"
    assert searched_group["module_asset_no"] == "admin module"
    assert searched_group["exception_note"] == "manual exception note"

    reset_review = production_client.patch(
        f"/groups/{group['id']}/reset-unreviewed",
        headers=admin_headers,
        json={"reason": "admin smoke reset review"},
    )
    assert reset_review.status_code == 200
    reset_review_group = reset_review.json()["data"]["group"]
    assert reset_review_group["status"] == "pending"
    assert reset_review_group["reviewer"] == ""
    assert reset_review_group["review_note"] == ""
    assert reset_review_group["exception_note"] == ""

    reset_construction = production_client.patch(
        f"/groups/{group['id']}/reset-unconstructed",
        headers=admin_headers,
        json={"reason": "admin smoke reset construction"},
    )
    assert reset_construction.status_code == 200
    reset_construction_payload = reset_construction.json()["data"]
    assert reset_construction_payload["group"]["photo_count"] == 0
    assert reset_construction_payload["soft_deleted_photos"] >= 0

    audit = production_client.get("/local-test/audit-log?limit=20", headers=admin_headers)
    assert audit.status_code == 200
    actions = [item["action"] for item in audit.json()["data"]["items"]]
    assert "admin_group_metadata_update" in actions
    assert "admin_group_reset_unreviewed" in actions
    assert "group_reset_to_unconstructed" in actions


def test_unmatched_create_group_route_creates_terminal_task() -> None:
    client.post("/local-test/bootstrap")
    client.post(
        "/local-test/scan/import-url-rows",
        json={
            "rows": [
                {
                    "meter_no": "NO-MATCH-API",
                    "terminal": "",
                    "collector": "collector-api",
                    "module_asset_no": "module-api",
                    "photo_urls": "https://example.test/api-a.jpg,https://example.test/api-b.jpg",
                }
            ]
        },
    )
    unmatched = client.get("/local-test/unmatched?query=NO-MATCH-API").json()["data"]["items"][0]

    response = client.post(
        f"/local-test/unmatched/{unmatched['unmatched_id']}/create-group",
        json={"actor": "api-test", "terminal": "T-API", "updates": {"address": "api manual address"}},
    )
    tasks = client.get("/local-test/tasks").json()["data"]["items"]

    assert response.status_code == 200
    assert response.json()["data"]["group"]["terminal"] == "T-API"
    assert response.json()["data"]["group"]["photo_count"] == 2
    assert any(task["terminal"] == "T-API" and task["can_claim"] for task in tasks)


def test_unmatched_blank_route_creates_unmatched_record() -> None:
    client.post("/local-test/bootstrap")

    response = client.post("/local-test/unmatched/blank", json={"actor": "api-test"})
    record = response.json()["data"]["record"]
    listed = client.get(f"/local-test/unmatched?query={record['unmatched_id']}").json()["data"]

    assert response.status_code == 200
    assert record["record_type"] == "blank_group"
    assert listed["total"] == 1
    assert listed["items"][0]["unmatched_id"] == record["unmatched_id"]


def test_group_metadata_route_updates_form_fields() -> None:
    client.post("/local-test/bootstrap")
    created = client.post(
        "/local-test/groups",
        json={
            "actor": "api-test",
            "terminal": "T-FORM",
            "meter_no": "M-FORM",
            "address": "form address before",
        },
    ).json()["data"]["group"]
    client.post(
        f"/local-test/groups/{created['id']}/photos/import-urls",
        json={
            "actor": "api-test",
            "photo_urls": ["https://example.test/form-1.jpg"],
            "collector": "old collector",
            "module_asset_no": "old module",
            "creator": "old installer",
        },
    )

    response = client.patch(
        f"/local-test/groups/{created['id']}/metadata",
        json={
            "actor": "api-test",
            "updates": {
                "meter_no": "API-FORM",
                "address": "api form address",
                "collector": "api collector",
                "module_asset_no": "api module",
                "creator": "api installer",
            },
        },
    )
    updated = response.json()["data"]["group"]

    assert response.status_code == 200
    assert updated["meter_no"] == "API-FORM"
    assert updated["address"] == "api form address"
    assert updated["photos"][0]["collector"] == "api collector"
    assert updated["photos"][0]["asset_no"] == "api module"
    assert updated["photos"][0]["creator"] == "api installer"


def test_manual_group_and_photo_import_routes() -> None:
    client.post("/local-test/bootstrap")

    created = client.post(
        "/local-test/groups",
        json={
            "actor": "api-test",
            "terminal": "T-MANUAL",
            "meter_no": "M-MANUAL",
            "address": "manual address",
        },
    )
    group = created.json()["data"]["group"]

    imported = client.post(
        f"/local-test/groups/{group['id']}/photos/import-urls",
        json={
            "actor": "api-test",
            "photo_urls": ["https://example.test/manual-1.jpg", "https://example.test/manual-2.jpg"],
            "collector": "collector-manual",
            "module_asset_no": "module-manual",
            "creator": "installer-manual",
        },
    )

    assert created.status_code == 200
    assert group["terminal"] == "T-MANUAL"
    assert group["photo_count"] == 0
    assert imported.status_code == 200
    assert imported.json()["data"]["added"] == 2
    assert imported.json()["data"]["group"]["photo_count"] == 2
    imported_photo = imported.json()["data"]["group"]["photos"][0]
    assert imported_photo["storage_type"] == "external_url"
    assert imported_photo["storage_key"] == "https://example.test/manual-1.jpg"


def test_manual_group_photo_upload_route() -> None:
    client.post("/local-test/bootstrap")

    created = client.post(
        "/local-test/groups",
        json={
            "actor": "api-test",
            "terminal": "T-UPLOAD",
            "meter_no": "M-UPLOAD",
            "address": "manual upload address",
        },
    )
    group = created.json()["data"]["group"]

    uploaded = client.post(
        f"/local-test/groups/{group['id']}/photos/upload-images",
        data={"actor": "api-test", "collector": "collector-upload", "module_asset_no": "module-upload"},
        files=[
            ("files", ("upload-a.jpg", b"fake-image-a", "image/jpeg")),
            ("files", ("upload-b.png", b"fake-image-b", "image/png")),
        ],
    )

    assert uploaded.status_code == 200
    payload = uploaded.json()["data"]
    assert payload["added"] == 2
    assert payload["group"]["photo_count"] == 2
    assert payload["uploaded_urls"][0].startswith("/static/uploads/manual/")
    uploaded_photo = payload["group"]["photos"][0]
    assert uploaded_photo["storage_type"] == "local_upload"
    assert uploaded_photo["storage_key"].startswith("manual/")
    assert uploaded_photo["sha256"]


def test_catalog_routes_are_filterable() -> None:
    client.post("/local-test/bootstrap")

    response = client.get("/local-test/catalog/total?limit=5")
    filtered = client.get("/local-test/catalog/stage?query=350&limit=5")

    assert response.status_code == 200
    assert response.json()["data"]["total"] > 0
    assert len(response.json()["data"]["items"]) <= 5
    assert filtered.status_code == 200
    assert "items" in filtered.json()["data"]


def test_sync_config_page_is_available() -> None:
    assert_vue_shell_response(client.get("/sync-config"))

def test_clear_scan_data_route_resets_local_scan_state() -> None:
    client.post("/local-test/bootstrap")

    response = client.post("/local-test/scan/clear")

    assert response.status_code == 200
    payload = response.json()["data"]
    assert payload["summary"]["scan_rows"] == 0
    assert payload["summary"]["downloaded_photos"] == 0
    assert payload["summary"]["unclassified_photos"] == 0


def test_url_row_import_route_updates_local_tasks() -> None:
    client.post("/local-test/bootstrap")
    first_group = client.get("/local-test/groups?limit=1").json()["data"]["items"][0]
    client.post("/local-test/scan/clear")

    response = client.post(
        "/local-test/scan/import-url-rows",
        json={
            "rows": [
                {
                    "meter_no": first_group["meter_no"],
                    "terminal": first_group["terminal"],
                    "collector": "C-001",
                    "module_asset_no": "M-001",
                    "photo_urls": "https://example.test/1.jpg,https://example.test/2.jpg",
                }
            ]
        },
    )
    tasks = client.get("/local-test/tasks").json()["data"]["items"]

    assert response.status_code == 200
    assert response.json()["data"]["applied_records"] == 2
    assert any(task["can_claim"] for task in tasks)


def test_url_row_import_adds_incremental_photos_for_duplicate_meter_number() -> None:
    client.post("/local-test/bootstrap")
    first_group = client.get("/local-test/groups?limit=1").json()["data"]["items"][0]
    client.post("/local-test/scan/clear")

    rows = [
        {
            "meter_no": first_group["meter_no"],
            "terminal": first_group["terminal"],
            "collector": "C-001",
            "module_asset_no": "M-001",
            "photo_urls": "https://example.test/1.jpg",
        },
        {
            "meter_no": first_group["meter_no"],
            "terminal": first_group["terminal"],
            "collector": "C-002",
            "module_asset_no": "M-002",
            "photo_urls": "https://example.test/2.jpg",
        },
    ]
    response = client.post("/local-test/scan/import-url-rows", json={"rows": rows})
    group = client.get(f"/local-test/groups/{first_group['id']}").json()["data"]

    assert response.status_code == 200
    assert response.json()["data"]["applied_records"] == 2
    assert response.json()["data"]["skipped_duplicate_meters"] == 0
    assert response.json()["data"]["photos_new"] == 2
    assert group["photo_count"] == 2


def test_installer_daily_workload_includes_work_time_segments() -> None:
    client.post("/local-test/bootstrap")
    groups = client.get("/local-test/groups?limit=3").json()["data"]["items"]
    client.post("/local-test/scan/clear")

    rows = [
        {
            "meter_no": groups[0]["meter_no"],
            "terminal": groups[0]["terminal"],
            "creator": "kpi-installer",
            "created_at": "2026-06-22 08:10:00",
            "photo_urls": "https://example.test/kpi-1.jpg",
        },
        {
            "meter_no": groups[0]["meter_no"],
            "terminal": groups[0]["terminal"],
            "creator": "kpi-installer",
            "created_at": "2026-06-22 08:50:00",
            "photo_urls": "https://example.test/kpi-2.jpg",
        },
        {
            "meter_no": groups[1]["meter_no"],
            "terminal": groups[1]["terminal"],
            "creator": "kpi-installer",
            "created_at": "2026-06-22 09:20:00",
            "photo_urls": "https://example.test/kpi-3.jpg",
        },
        {
            "meter_no": groups[2]["meter_no"],
            "terminal": groups[2]["terminal"],
            "creator": "kpi-installer",
            "created_at": "2026-06-22 11:00:00",
            "photo_urls": "https://example.test/kpi-4.jpg",
        },
    ]

    response = client.post("/local-test/scan/import-url-rows", json={"rows": rows})
    workload = client.get("/local-test/installers/kpi-installer/daily-workload")

    assert response.status_code == 200
    assert workload.status_code == 200
    item = workload.json()["data"]["items"][0]
    assert item["date"] == "2026-06-22"
    assert item["start_time"] == "08:10"
    assert item["end_time"] == "11:00"
    assert item["work_duration_minutes"] == 50
    assert item["efficiency_duration_minutes"] == 50
    assert item["work_duration_label"]
    assert item["work_span_minutes"] == 170
    assert item["break_threshold_minutes"] == 45
    assert item["timepoint_count"] == 4
    assert item["completion_count"] == 3
    assert item["completion_per_effective_hour"] == 3.6
    assert item["weighted_completion"] >= 2.25
    segments = {segment["hour"]: segment["minutes"] for segment in item["hourly_segments"]}
    assert segments[8] == 35
    assert segments[9] == 15
    assert segments[10] == 0
    two_hour = {segment["start_hour"]: segment for segment in item["two_hour_segments"]}
    assert two_hour[8]["minutes"] == 50
    assert two_hour[8]["efficiency_minutes"] == 50
    assert two_hour[8]["completion_count"] == 2
    assert two_hour[8]["completion_per_effective_hour"] == 2.4
    assert len(two_hour[8]["addresses"]) == 2
    assert two_hour[10]["minutes"] == 0
    assert two_hour[10]["completion_count"] == 1
    assert two_hour[10]["addresses"][0]["difficulty_weight"] >= 0.75


def test_installer_kpi_clusters_same_building_number_public_equipment() -> None:
    completion_records = [
        {
            "group_id": "g-room",
            "meter_no": "110000000001",
            "terminal": "350000000001",
            "address": "上海市宝山区聚丰园路95弄18号201室",
            "completed_at": datetime(2026, 6, 22, 8, 10),
        },
        {
            "group_id": "g-public",
            "meter_no": "110000000002",
            "terminal": "350000000001",
            "address": "上海市宝山区聚丰园路95弄18号公用设备",
            "completed_at": datetime(2026, 6, 22, 8, 20),
        },
        {
            "group_id": "g-other-building",
            "meter_no": "110000000003",
            "terminal": "350000000001",
            "address": "上海市宝山区聚丰园路95弄19号公用设备",
            "completed_at": datetime(2026, 6, 22, 8, 30),
        },
    ]

    summary = local_simulation.build_work_time_summary(
        [record["completed_at"] for record in completion_records],
        completion_records,
    )

    segment = next(item for item in summary["two_hour_segments"] if item["start_hour"] == 8)
    addresses = {item["meter_no"]: item for item in segment["addresses"]}

    assert addresses["110000000001"]["address_cluster_key"].endswith("95弄18号")
    assert addresses["110000000002"]["address_cluster_key"] == addresses["110000000001"]["address_cluster_key"]
    assert addresses["110000000002"]["cluster_size"] == 2
    assert addresses["110000000003"]["address_cluster_key"].endswith("95弄19号")
    assert addresses["110000000003"]["address_cluster_key"] != addresses["110000000001"]["address_cluster_key"]


def test_excel_exports_return_real_workbooks() -> None:
    client.post("/local-test/bootstrap")
    task = client.get("/local-test/tasks").json()["data"]["items"][0]

    task_export = client.post("/exports/task-detail", json={"task_id": task["id"]})
    all_final_export = client.post("/exports/final-delivery", json={"project_id": 1})
    terminal_final_export = client.post("/exports/final-delivery", json={"task_id": task["id"]})
    exception_export = client.post("/exports/exception-meters", json={})

    assert task_export.status_code == 200
    assert task_export.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "attachment" in task_export.headers["content-disposition"]
    assert task_export.content.startswith(b"PK")
    assert all_final_export.status_code == 400
    assert terminal_final_export.status_code == 200
    assert terminal_final_export.content.startswith(b"PK")
    assert exception_export.status_code == 200
    assert exception_export.content.startswith(b"PK")
    from io import BytesIO

    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(exception_export.content), read_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    assert "\u5f02\u5e38\u539f\u56e0" in headers
    assert "\u73b0\u573a\u5904\u7406\u5efa\u8bae" in headers


def test_final_delivery_manifest_supports_frontend_zip_export() -> None:
    client.post("/local-test/bootstrap")
    task = client.get("/local-test/tasks").json()["data"]["items"][0]
    client.post(f"/local-test/tasks/{task['id']}/claim", json={"reviewer": "api-test"})
    group = client.get(f"/local-test/tasks/{task['id']}/groups?limit=1&scan_only=false").json()["data"]["items"][0]
    client.patch(
        f"/local-test/groups/{group['id']}/review",
        json={"status": "approved", "reviewer": "api-test", "note": "done"},
    )

    all_response = client.get("/local-test/export-manifest/final-delivery")
    response = client.get(f"/local-test/export-manifest/final-delivery?task_id={task['id']}")
    all_scope_response = client.get(f"/local-test/export-manifest/final-delivery?task_id={task['id']}&review_scope=all")

    assert all_response.status_code == 400
    assert response.status_code == 200
    payload = response.json()["data"]
    assert payload["photo_limit_per_group"] == 4
    assert payload["scope"]["task_id"] == task["id"]
    assert payload["scope"]["review_scope"] == "reviewed"
    assert payload["groups"]
    assert all(item["status"] in {"approved", "exception"} or item.get("has_archive_blocker") for item in payload["groups"])
    assert all_scope_response.status_code == 200
    assert len(all_scope_response.json()["data"]["groups"]) >= len(payload["groups"])
    first_group = payload["groups"][0]
    assert {"terminal", "address", "meter_no", "photos"}.issubset(first_group)
    if first_group["photos"]:
        first_photo = first_group["photos"][0]
        assert {"image_url", "category_label", "archive_filename"}.issubset(first_photo)


def test_group_detail_uses_local_data_without_legacy_sync(monkeypatch) -> None:
    client.post("/local-test/bootstrap")
    first_group = client.get("/local-test/groups?limit=1").json()["data"]["items"][0]

    def fail_if_called(group_id: str):
        raise AssertionError(f"legacy sync should not run for group detail: {group_id}")

    monkeypatch.setattr(sync_manager, "load_group_photo_urls", fail_if_called)

    response = client.get(f"/local-test/groups/{first_group['id']}")

    assert response.status_code == 200
    payload = response.json()["data"]
    assert payload["id"] == first_group["id"]
    assert "photos" in payload
