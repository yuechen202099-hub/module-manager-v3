from copy import deepcopy
from io import BytesIO
from pathlib import Path
from importlib.util import find_spec

import pytest
from fastapi.testclient import TestClient

from app.main import create_app
from app.services import local_simulation
from app.services.local_simulation import (
    DEFAULT_SCAN_FILE,
    DEFAULT_TOTAL_CATALOG,
    LocalTestPaths,
    apply_group_photo_urls,
    apply_synced_scan_records,
    add_photo_urls_to_group,
    blank_state,
    build_delivery_cache_for_group,
    build_final_delivery_export,
    build_final_delivery_manifest,
    bootstrap_local_simulation,
    classify_photo,
    claim_task,
    clear_scan_data,
    assign_construction_task,
    create_blank_unmatched_record,
    create_empty_group_for_terminal,
    create_group_from_unmatched_record,
    delete_group_photo,
    delete_unmatched_record,
    dedupe_unmatched_records,
    get_task_progress,
    get_delivery_cached_photo_path,
    get_group,
    get_state,
    import_scan_template_xlsx,
    import_total_catalog_xlsx,
    list_audit_events,
    list_exception_groups,
    list_task_groups,
    list_groups,
    list_unmatched_records,
    list_tasks,
    normalize_cell,
    release_task,
    rematch_unmatched_record,
    reset_group_to_unconstructed,
    return_group_to_exception_order,
    review_group,
    save_exception_note,
    search_group_targets,
    set_current_team,
    submit_construction_exception_order,
    sync_state_photos_to_oss,
    reset_current_team,
    update_group_metadata,
)


SAMPLE_FILES = [DEFAULT_TOTAL_CATALOG, DEFAULT_SCAN_FILE]


requires_sample_workbooks = pytest.mark.skipif(
    not all(path.exists() for path in SAMPLE_FILES) or find_spec("openpyxl") is None,
    reason="local sample workbooks or openpyxl are not available",
)


def fake_catalog_rows(source: str) -> list[dict]:
    return [
        {
            "source": source,
            "row_number": 2,
            "terminal": "T-001",
            "meter_no": "ZZ1001",
            "address": "A road",
            "meter_match_key": "1001",
        },
        {
            "source": source,
            "row_number": 3,
            "terminal": "T-002",
            "meter_no": "ZZ1002",
            "address": "B road",
            "meter_match_key": "1002",
        },
        {
            "source": source,
            "row_number": 4,
            "terminal": "T-003",
            "meter_no": "ZZ1003",
            "address": "C road",
            "meter_match_key": "1003",
        },
    ]


def fake_scan_rows() -> list[dict]:
    rows = []
    for index in range(4):
        rows.append(
            {
                "row_number": index + 2,
                "barcode": f"scan-{index}",
                "meter_match_key": "1001",
                "source_file": "local",
                "collector": "collector",
                "asset_no": f"asset-{index}",
                "asset_type": "module",
                "creator": "tester",
                "created_at": "2026-06-09",
                "has_image": True,
            }
        )
    rows.append(
        {
            "row_number": 6,
            "barcode": "scan-4",
            "meter_match_key": "1002",
            "source_file": "local",
            "collector": "collector",
            "asset_no": "asset-4",
            "asset_type": "module",
            "creator": "tester",
            "created_at": "2026-06-09",
            "has_image": True,
        }
    )
    return rows


def archive_all_group_photos(group: dict, reviewer: str = "alice") -> None:
    categories = ["before_box", "collector_barcode", "module_meter", "after_box"]
    for index, photo in enumerate(list(group["photos"])):
        classify_photo(group["id"], photo["id"], categories[index % len(categories)], reviewer=reviewer)


def build_catalog_workbook_bytes(rows: list[tuple[str, str, str]]) -> bytes:
    pytest.importorskip("openpyxl")
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["终端", "表号", "安装地址"])
    for row in rows:
        sheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


@pytest.fixture()
def synthetic_state(monkeypatch: pytest.MonkeyPatch) -> dict:
    def read_catalog(path: Path, source: str) -> list[dict]:
        return fake_catalog_rows(source)

    monkeypatch.setattr(local_simulation, "read_catalog_rows", read_catalog)
    monkeypatch.setattr(local_simulation, "read_scan_rows", lambda path: fake_scan_rows())
    return bootstrap_local_simulation(LocalTestPaths(Path("total.xlsx"), Path("stage.xlsx"), Path("scan.xlsx")))


def test_new_team_starts_empty_and_imports_total_catalog() -> None:
    token = set_current_team("empty-total-import-team")
    try:
        state = get_state()
        assert state["loaded"] is False
        assert state["summary"]["groups"] == 0
        assert state["tasks"] == []

        workbook = build_catalog_workbook_bytes(
            [
                ("T-001", "ZZ1001", "A road"),
                ("T-001", "ZZ1001", "A road duplicate"),
                ("T-002", "ZZ1002", "B road"),
            ]
        )
        result = import_total_catalog_xlsx(workbook)
        tasks = list_tasks()

        assert result["catalog_rows"] == 3
        assert result["imported_rows"] == 2
        assert result["skipped_duplicate_meters"] == 1
        assert result["summary"]["groups"] == 2
        assert result["summary"]["stage_catalog_rows"] == 0
        assert {task["terminal"] for task in tasks} == {"T-001", "T-002"}
        assert all(task["can_claim"] is False for task in tasks)
    finally:
        reset_current_team(token)


def test_local_upload_photos_are_synced_to_oss(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    token = set_current_team("oss-local-sync-team")
    try:
        state = get_state()
        state.clear()
        state.update(blank_state("oss-local-sync-team"))
        upload_dir = tmp_path / "manual"
        upload_dir.mkdir()
        (upload_dir / "photo.jpg").write_bytes(b"local-photo-bytes")
        state["groups"] = [
            {
                "id": "g-oss-001",
                "task_id": 1,
                "meter_match_key": "1001",
                "meter_no": "ZZ1001",
                "terminal": "T-001",
                "address": "A road",
                "status": "pending",
                "photo_count": 1,
                "photos": [
                    {
                        "id": "p-local-001",
                        "image_url": "/static/uploads/manual/photo.jpg",
                        "storage_type": "local_upload",
                        "storage_key": "manual/photo.jpg",
                        "source_file": "manual",
                    }
                ],
            }
        ]

        def fake_save_image_bytes(**kwargs):
            assert kwargs["content"] == b"local-photo-bytes"
            assert kwargs["scope"] == "imported"
            return {
                "url": "oss://bucket/imported/photo.jpg",
                "sha256": "sha256-local",
                "storage_type": "oss",
                "storage_key": "imported/photo.jpg",
                "storage_bucket": "bucket",
                "storage_source": "imported-oss-upload",
                "content_type": "image/jpeg",
            }

        monkeypatch.setattr(local_simulation, "active_storage_backend", lambda: "oss")
        monkeypatch.setattr(local_simulation, "static_upload_root", lambda: tmp_path)
        monkeypatch.setattr(local_simulation, "save_image_bytes", fake_save_image_bytes)
        monkeypatch.setattr(local_simulation, "save_all_team_states", lambda: None)

        report = sync_state_photos_to_oss(team_id="oss-local-sync-team", max_workers=1)
        photo = state["groups"][0]["photos"][0]

        assert report["uploaded"] == 1
        assert report["failed"] == 0
        assert photo["image_url"] == "oss://bucket/imported/photo.jpg"
        assert photo["storage_type"] == "oss"
        assert photo["storage_key"] == "imported/photo.jpg"
        assert photo["pre_oss_image_url"] == "/static/uploads/manual/photo.jpg"
    finally:
        reset_current_team(token)


def test_team_states_are_isolated(monkeypatch: pytest.MonkeyPatch) -> None:
    def read_catalog(path: Path, source: str) -> list[dict]:
        return fake_catalog_rows(source)

    monkeypatch.setattr(local_simulation, "read_catalog_rows", read_catalog)
    monkeypatch.setattr(local_simulation, "read_scan_rows", lambda path: fake_scan_rows())

    token_a = set_current_team("team-a")
    try:
        team_a = bootstrap_local_simulation(LocalTestPaths(Path("total.xlsx"), Path("stage.xlsx"), Path("scan.xlsx")))
        claim_task(1, reviewer="alice")
        assert team_a["tasks"][0]["claimed_by"] == "alice"
    finally:
        reset_current_team(token_a)

    token_b = set_current_team("team-b")
    try:
        team_b = bootstrap_local_simulation(LocalTestPaths(Path("total.xlsx"), Path("stage.xlsx"), Path("scan.xlsx")))
        assert team_b["summary"]["team_id"] == "team-b"
        assert team_b["tasks"][0]["claimed_by"] is None
        claim_task(1, reviewer="bob")
        assert team_b["tasks"][0]["claimed_by"] == "bob"
    finally:
        reset_current_team(token_b)

    token_a = set_current_team("team-a")
    try:
        assert list_tasks()[0]["claimed_by"] == "alice"
    finally:
        reset_current_team(token_a)


@requires_sample_workbooks
def test_bootstrap_local_simulation_uses_sample_workbooks() -> None:
    state = bootstrap_local_simulation(LocalTestPaths())
    summary = state["summary"]

    assert summary["total_catalog_rows"] > 20_000
    assert summary["stage_catalog_rows"] == 0
    assert summary["scan_rows"] > 0
    assert summary["groups"] == summary["total_catalog_rows"]
    assert summary["matched_groups"] > 0


@requires_sample_workbooks
def test_local_groups_are_displayed_with_total_catalog_meter_number() -> None:
    bootstrap_local_simulation(LocalTestPaths())
    result = list_groups(limit=10)

    assert result["total"] > 0
    first = result["items"][0]
    assert first["meter_no"]
    assert first["address"]
    assert first["meter_no"] != first["meter_match_key"]


@requires_sample_workbooks
def test_group_detail_can_be_loaded_from_generated_id() -> None:
    bootstrap_local_simulation(LocalTestPaths())
    result = list_groups(limit=1)
    group = get_group(result["items"][0]["id"])

    assert group is not None
    assert group["id"].startswith("g-")


def test_task_can_be_claimed_and_released(synthetic_state: dict) -> None:
    claimed = claim_task(1, reviewer="alice")
    assert claimed["status"] == "in_review"
    assert claimed["claimed_by"] == "alice"
    assert claimed["claimed_at"]

    released = release_task(1, reviewer="alice")
    assert released["status"] == "released"
    assert released["claimed_by"] is None
    assert released["released_at"]


def test_tasks_are_split_by_terminal_and_require_scan_info(synthetic_state: dict) -> None:
    tasks = list_tasks()

    assert [task["terminal"] for task in tasks] == ["T-001", "T-002", "T-003"]
    assert tasks[0]["scan_rows"] == 4
    assert tasks[0]["can_claim"] is True
    assert tasks[0]["complete_groups"] == 1
    assert tasks[0]["completeness_rate"] == 1.0
    assert tasks[0]["renovation_count"] == 1
    assert tasks[0]["uploaded_count"] == 1
    assert tasks[0]["upload_rate"] == 1.0
    assert tasks[0]["review_rate"] == 0.0
    assert tasks[1]["partial_groups"] == 1
    assert tasks[1]["completeness_rate"] == 1.0
    assert tasks[1]["exception_groups"] == 1
    assert tasks[1]["unreviewed_count"] == 0
    assert tasks[1]["review_rate"] == 0.0
    assert tasks[2]["scan_rows"] == 0
    assert tasks[2]["can_claim"] is False
    assert tasks[2]["completeness_rate"] == 0.0
    assert tasks[2]["incomplete_groups"] == 0
    assert tasks[2]["unconstructed_groups"] == 1

    with pytest.raises(ValueError):
        claim_task(tasks[2]["id"], reviewer="alice")


def test_summary_reports_installer_group_share(synthetic_state: dict) -> None:
    summary = synthetic_state["summary"]
    distribution = {item["installer"]: item for item in summary["installer_distribution"]}

    assert distribution["tester"]["group_count"] == 2
    assert distribution["tester"]["share"] == 1.0
    assert "未填写" not in distribution


def test_normalize_cell_repairs_latin1_mojibake() -> None:
    mojibake = "\u00e5\u00ae\u009d\u00e5\u00b1\u00b1\u00e5\u008c\u00ba\u00e9\u0094\u00a6\u00e7\u00a7\u008b\u00e8\u00b7\u00af1152\u00e5\u008f\u00b7"

    assert normalize_cell(mojibake) == "\u5b9d\u5c71\u533a\u9526\u79cb\u8def1152\u53f7"
    assert normalize_cell("\u4e0a\u6d77\u5e02\u5b9d\u5c71\u533a") == "\u4e0a\u6d77\u5e02\u5b9d\u5c71\u533a"


def test_review_group_updates_status_and_summary(synthetic_state: dict) -> None:
    state = synthetic_state
    first_id = state["groups"][0]["id"]

    claim_task(1, reviewer="alice")
    reviewed = review_group(first_id, status="approved", reviewer="alice", note="sample passed")
    tasks = list_tasks()

    assert reviewed["status"] == "approved"
    assert reviewed["review_note"] == "sample passed"
    assert state["summary"]["approved_groups"] == 1
    assert state["summary"]["reviewed_groups"] == 1
    assert state["summary"]["exception_groups"] == 1
    assert state["summary"]["review_progress"] == 0.3333
    assert tasks[0]["completed_groups"] == 1
    assert tasks[0]["progress"] == 1.0


def test_review_group_rejects_unknown_status(synthetic_state: dict) -> None:
    state = synthetic_state
    first_id = state["groups"][0]["id"]

    with pytest.raises(ValueError):
        review_group(first_id, status="done", reviewer="alice")


def test_review_group_blocks_non_claiming_reviewer(synthetic_state: dict) -> None:
    first_id = synthetic_state["groups"][0]["id"]
    claim_task(1, reviewer="alice")

    with pytest.raises(ValueError):
        review_group(first_id, status="approved", reviewer="bob")


def test_exception_note_marks_group_and_progress(synthetic_state: dict) -> None:
    second_id = synthetic_state["groups"][1]["id"]

    claim_task(2, reviewer="alice")
    reviewed = save_exception_note(second_id, reviewer="alice", note="missing required photos")
    progress = get_task_progress(2)

    assert reviewed["status"] == "exception"
    assert reviewed["exception_note"] == "missing required photos"
    assert progress["exception_groups"] == 1
    assert progress["reviewed_groups"] == 0
    assert progress["pending_groups"] == 0


def test_pending_archive_blocker_counts_as_problem_but_not_archived(synthetic_state: dict) -> None:
    group = synthetic_state["groups"][1]
    group["status"] = "pending"
    local_simulation.refresh_summary()
    progress = get_task_progress(group["task_id"])

    assert group["has_archive_blocker"] is True
    assert synthetic_state["summary"]["exception_groups"] == 1
    assert synthetic_state["summary"]["reviewed_groups"] == 0
    assert progress["exception_groups"] == 1
    assert progress["reviewed_groups"] == 0
    assert progress["pending_groups"] == 0


def test_problem_group_is_not_counted_again_as_missing_photo(synthetic_state: dict) -> None:
    group = synthetic_state["groups"][1]

    assert group["photo_count"] == 1
    local_simulation.refresh_summary()

    assert group["has_archive_blocker"] is True
    assert synthetic_state["summary"]["exception_groups"] == 1
    assert synthetic_state["summary"]["incomplete_groups"] == 0
    assert synthetic_state["summary"]["scanned_groups"] == 2


def test_task_groups_can_be_filtered_by_status(synthetic_state: dict) -> None:
    result = list_task_groups(2, status="incomplete")

    assert result["total"] == 1
    assert result["items"][0]["photo_count"] == 1


def test_task_groups_can_be_limited_to_scanned_groups(synthetic_state: dict) -> None:
    scanned = list_task_groups(3, scan_only=True)
    all_groups = list_task_groups(3, scan_only=False)

    assert scanned["total"] == 0
    assert all_groups["total"] == 1


def test_task_groups_can_return_lightweight_summaries(synthetic_state: dict) -> None:
    result = list_task_groups(1, limit=1, summary_only=True)

    assert result["total"] >= 1
    assert "photos" not in result["items"][0]
    assert "photo_count" in result["items"][0]
    assert "reviewer" in result["items"][0]


def test_task_groups_are_ordered_for_review_queue(synthetic_state: dict) -> None:
    template = deepcopy(synthetic_state["groups"][0])

    def make_group(group_id: str, meter_no: str, status: str, photo_count: int, archived: bool = False) -> dict:
        group = deepcopy(template)
        group.update(
            {
                "id": group_id,
                "task_id": 1,
                "terminal": template["terminal"],
                "meter_no": meter_no,
                "status": status,
                "photo_count": photo_count,
                "has_archive_blocker": status == "exception",
                "exception_reasons": ["测试异常"] if status == "exception" else [],
            }
        )
        group["photos"] = deepcopy(template["photos"][:photo_count])
        for index, photo in enumerate(group["photos"], start=1):
            photo["id"] = f"{group_id}-photo-{index}"
            photo["archive_status"] = "archived" if archived else "pending"
        return group

    synthetic_state["groups"] = [
        group for group in synthetic_state["groups"] if group["task_id"] != 1
    ] + [
        make_group("queue-unconstructed", "0004", "pending", 0),
        make_group("queue-exception", "0003", "exception", 4),
        make_group("queue-done", "0002", "approved", 4, archived=True),
        make_group("queue-reviewable", "0001", "pending", 4),
    ]

    result = list_task_groups(1, scan_only=False)

    assert [item["id"] for item in result["items"]] == [
        "queue-reviewable",
        "queue-exception",
        "queue-unconstructed",
        "queue-done",
    ]


def test_unconstructed_groups_stay_out_of_exception_queue_but_remain_in_task_list(synthetic_state: dict) -> None:
    no_scan_group = synthetic_state["groups"][2]

    assert no_scan_group["photo_count"] == 0
    assert no_scan_group.get("has_archive_blocker") is False
    assert no_scan_group.get("exception_reasons") == []

    exceptions = list_exception_groups()
    task_groups = list_task_groups(3, scan_only=False)

    assert no_scan_group["id"] not in {group["id"] for group in exceptions["items"]}
    assert no_scan_group["id"] in {group["id"] for group in task_groups["items"]}
    assert task_groups["items"][0]["construction_status"] == "unconstructed"
    assert synthetic_state["summary"]["unconstructed_groups"] == 1
    assert synthetic_state["summary"]["incomplete_groups"] == 0
    assert synthetic_state["summary"]["exception_groups"] == 1


def test_downloaded_photo_can_be_classified(synthetic_state: dict) -> None:
    first_group = synthetic_state["groups"][0]
    photo = first_group["photos"][0]

    with pytest.raises(ValueError, match="must be claimed"):
        classify_photo(first_group["id"], photo["id"], "collector_barcode", reviewer="alice")

    claim_task(1, reviewer="alice")
    classified = classify_photo(first_group["id"], photo["id"], "collector_barcode", reviewer="alice")

    assert classified["category"] == "collector_barcode"
    assert classified["category_label"] == "\u91c7\u96c6\u5668\u6761\u5f62\u7801"
    assert classified["classified_by"] == "alice"
    assert classified["archive_status"] == "archived"
    assert classified["archive_filename"] == "\u91c7\u96c6\u5668\u6761\u5f62\u7801.jpg"
    assert classified["archived_at"]
    assert synthetic_state["summary"]["unclassified_photos"] == 4


def test_previewable_url_photo_can_be_classified_without_download_status(synthetic_state: dict) -> None:
    first_group = synthetic_state["groups"][0]
    photo = first_group["photos"][0]
    photo["image_url"] = "https://example.test/photo.jpg"
    photo["download_status"] = "oss_reused"

    claim_task(1, reviewer="alice")
    classified = classify_photo(first_group["id"], photo["id"], "collector_barcode", reviewer="alice")

    assert classified["category"] == "collector_barcode"
    assert classified["download_status"] == "downloaded"


def test_delivery_cache_builds_for_approved_group(
    synthetic_state: dict,
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    group = synthetic_state["groups"][0]
    for photo in group["photos"]:
        photo["image_url"] = f"https://example.test/{photo['id']}.jpg"

    monkeypatch.setattr(local_simulation.settings, "delivery_cache_path", str(tmp_path))
    monkeypatch.setattr(local_simulation, "schedule_delivery_cache_build", lambda *args, **kwargs: None)
    monkeypatch.setattr(local_simulation, "save_all_team_states", lambda: None)
    monkeypatch.setattr(
        local_simulation,
        "download_delivery_photo_content",
        lambda photo: (f"cached-{photo['id']}".encode("utf-8"), ".jpg", "image/jpeg"),
    )

    claim_task(group["task_id"], reviewer="alice")
    archive_all_group_photos(group, reviewer="alice")
    result = build_delivery_cache_for_group(group["id"], force=True)
    manifest = build_final_delivery_manifest(task_id=group["task_id"])
    first_photo = manifest["groups"][0]["photos"][0]
    cached_path = get_delivery_cached_photo_path(group["id"], first_photo["id"])

    assert group["status"] == "approved"
    assert result["status"] == "ready"
    assert first_photo["delivery_cache_url"].startswith(f"/local-test/delivery-cache/{group['id']}/")
    assert cached_path.read_bytes().startswith(b"cached-")


def test_photo_can_be_classified_as_unmatched_data_group(synthetic_state: dict) -> None:
    first_group = synthetic_state["groups"][0]
    photo = first_group["photos"][0]

    claim_task(1, reviewer="alice")
    classified = classify_photo(first_group["id"], photo["id"], "unmatched_group", reviewer="alice")

    assert classified["category"] == "unmatched_group"
    assert classified["category_label"] == "未匹配数据组"
    assert classified["archive_filename"] == "未匹配数据组.jpg"


def test_delete_group_photo_requires_claim_and_resets_review_state(synthetic_state: dict) -> None:
    group = synthetic_state["groups"][0]
    original_photo_count = group["photo_count"]
    photo = group["photos"][0]

    with pytest.raises(ValueError, match="must be claimed"):
        delete_group_photo(group["id"], photo["id"], reviewer="alice")

    claim_task(group["task_id"], reviewer="alice")
    deleted = delete_group_photo(group["id"], photo["id"], reviewer="alice")
    audits = list_audit_events()

    assert deleted["deleted_photo"]["id"] == photo["id"]
    assert deleted["group"]["photo_count"] == original_photo_count - 1
    assert all(item["id"] != photo["id"] for item in deleted["group"]["photos"])
    assert deleted["group"]["status"] == "incomplete"
    assert deleted["group"]["reviewer"] is None
    assert audits["items"][0]["action"] == "delete_group_photo"


def test_clear_scan_data_resets_downloaded_photos_and_claimable_tasks(synthetic_state: dict) -> None:
    state = clear_scan_data()
    tasks = list_tasks()

    assert state["summary"]["scan_rows"] == 0
    assert state["summary"]["downloaded_photos"] == 0
    assert state["summary"]["unclassified_photos"] == 0
    assert state["scan_unmatched"] == []
    assert all(group["photos"] == [] for group in state["groups"])
    assert all(group["photo_count"] == 0 for group in state["groups"])
    assert all(task["can_claim"] is False for task in tasks)


def test_apply_synced_scan_records_matches_catalog_and_refreshes_tasks(synthetic_state: dict) -> None:
    clear_scan_data()
    before = list_tasks()
    assert before[0]["can_claim"] is False

    result = apply_synced_scan_records(
        [
            {
                "file_id": "remote-file-1",
                "source_file": "remote-source",
                "installer": "installer",
                "barcode": "ABCDEFGHIJK000001001X",
                "meter_match_key": "1001",
                "terminal": "T-001",
                "collector": "collector",
                "meter_no": "ZZ1001",
                "module_asset_no": "asset-1",
                "address": "A road",
                "asset_type": "module",
                "creator": "tester",
                "created_at": "2026-06-09",
                "image_count": 2,
                "image_urls": ["https://download.example/photo-1.jpg", "https://download.example/photo-2.jpg"],
            }
        ]
    )
    after = list_tasks()
    group = synthetic_state["groups"][0]

    assert result["applied_records"] == 2
    assert result["unmatched_records"] == 0
    assert group["photo_count"] == 2
    assert group["photos"][0]["image_url"] == "https://download.example/photo-1.jpg"
    assert group["photos"][1]["image_url"] == "https://download.example/photo-2.jpg"
    assert after[0]["can_claim"] is True
    assert after[0]["scan_rows"] == 2

    duplicate = apply_synced_scan_records(
        [
            {
                "file_id": "remote-file-1",
                "source_file": "remote-source",
                "barcode": "ABCDEFGHIJK000001001X",
                "meter_match_key": "1001",
                "image_urls": ["https://download.example/photo-1.jpg", "https://download.example/photo-2.jpg"],
            }
        ]
    )

    assert duplicate["applied_records"] == 0
    assert duplicate["skipped_duplicate_meters"] == 0
    assert duplicate["photos_duplicate"] == 2
    assert group["photo_count"] == 2


def test_apply_synced_scan_records_supplements_duplicate_meter_in_same_batch(synthetic_state: dict) -> None:
    clear_scan_data()

    result = apply_synced_scan_records(
        [
            {
                "file_id": "remote-file-a",
                "source_file": "remote-source",
                "barcode": "ABCDEFGHIJK000001001X",
                "meter_match_key": "1001",
                "meter_no": "ZZ1001",
                "image_urls": ["https://download.example/a.jpg"],
            },
            {
                "file_id": "remote-file-b",
                "source_file": "remote-source",
                "barcode": "ABCDEFGHIJK000001001X",
                "meter_match_key": "1001",
                "meter_no": "ZZ1001",
                "image_urls": ["https://download.example/b.jpg"],
            },
        ]
    )

    group = synthetic_state["groups"][0]
    assert result["applied_records"] == 2
    assert result["skipped_duplicate_meters"] == 0
    assert result["photos_new"] == 2
    assert group["photo_count"] == 2
    assert group["photos"][0]["image_url"] == "https://download.example/a.jpg"
    assert group["photos"][1]["image_url"] == "https://download.example/b.jpg"


def test_reset_group_to_unconstructed_soft_clears_photos(synthetic_state: dict) -> None:
    clear_scan_data()
    apply_synced_scan_records(
        [
            {
                "file_id": "remote-reset",
                "source_file": "remote-source",
                "barcode": "ABCDEFGHIJK000001001X",
                "meter_match_key": "1001",
                "meter_no": "ZZ1001",
                "collector": "collector-a",
                "module_asset_no": "asset-a",
                "image_urls": ["https://download.example/reset-a.jpg", "https://download.example/reset-b.jpg"],
            }
        ]
    )
    group = synthetic_state["groups"][0]
    claim_task(group["task_id"], "alice")

    result = reset_group_to_unconstructed(group["id"], actor="alice", reason="现场返工")

    assert result["soft_deleted_photos"] == 2
    assert group["photos"] == []
    assert len(group["deleted_photos"]) == 2
    assert group["photo_count"] == 0
    assert group["status"] == "pending"
    assert group.get("construction_collector") is None
    assert group.get("construction_module_asset_no") is None
    assert group["deleted_photos"][0]["is_active"] is False


def test_return_group_to_exception_order_creates_assigned_work_order(synthetic_state: dict) -> None:
    clear_scan_data()
    apply_synced_scan_records(
        [
            {
                "file_id": "remote-exception",
                "source_file": "remote-source",
                "barcode": "ABCDEFGHIJK000001001X",
                "meter_match_key": "1001",
                "meter_no": "ZZ1001",
                "collector": "collector-a",
                "module_asset_no": "asset-a",
                "image_urls": ["https://download.example/exception-a.jpg"],
            }
        ]
    )
    group = synthetic_state["groups"][0]
    claim_task(group["task_id"], "alice")
    assign_construction_task(group["task_id"], actor="admin", constructor="constructor")

    result = return_group_to_exception_order(
        group["id"],
        actor="alice",
        category="module_error",
        note="模块号错误",
    )
    assert result["order"]["assigned_to"] == "constructor"
    assert result["order"]["status"] == "assigned"

    submitted = submit_construction_exception_order(
        result["order"]["id"],
        actor="constructor",
        updates={"collector": "collector-b", "module_asset_no": "asset-b"},
        note="现场已修正",
    )

    assert group["status"] == "pending"
    assert submitted["order"]["status"] == "submitted"
    assert submitted["group"]["construction_collector"] == "collector-b"
    assert submitted["group"]["construction_module_asset_no"] == "asset-b"


def test_unmatched_records_are_searchable_and_audited(synthetic_state: dict) -> None:
    result = apply_synced_scan_records(
        [
            {
                "file_id": "remote-unmatched-1",
                "source_file": "remote-source",
                "barcode": "NO-MATCH-001",
                "meter_match_key": "no-match",
                "terminal": "T-404",
                "collector": "collector-x",
                "module_asset_no": "module-x",
                "image_urls": ["https://download.example/unmatched.jpg"],
            }
        ]
    )

    records = list_unmatched_records(query="NO-MATCH")
    deleted = delete_unmatched_record(records["items"][0]["unmatched_id"], actor="alice", reason="bad source")
    audits = list_audit_events()

    assert result["unmatched_records"] == 1
    assert records["total"] == 1
    assert deleted["barcode"] == "NO-MATCH-001"
    assert audits["items"][0]["action"] == "delete_unmatched"
    assert audits["items"][0]["actor"] == "alice"


def test_unmatched_dedupe_removes_duplicate_meter_records(synthetic_state: dict) -> None:
    synthetic_state["scan_unmatched"] = []
    synthetic_state["scan_unmatched"].extend(
        [
            {
                "unmatched_id": "dup-plain",
                "barcode": "3130001201100201234503",
                "meter_no": "110020123450",
                "meter_match_key": "002012345",
                "terminal": "T-404",
                "address": "A road 101",
                "photo_urls": ["https://download.example/a.jpg"],
            },
            {
                "unmatched_id": "dup-assigned",
                "barcode": "3130001201100201234503",
                "meter_no": "110020123450",
                "meter_match_key": "002012345",
                "terminal": "T-404",
                "address": "A road 101",
                "assigned_to": "constructor-a",
                "photo_urls": ["https://download.example/a.jpg", "https://download.example/b.jpg"],
            },
        ]
    )

    result = dedupe_unmatched_records(actor="admin")
    records = list_unmatched_records(query="110020123450")
    audits = list_audit_events()

    assert result["removed"] == 1
    assert result["duplicate_ids"] == ["dup-plain"]
    assert records["total"] == 1
    assert records["items"][0]["unmatched_id"] == "dup-assigned"
    assert audits["items"][0]["action"] == "dedupe_unmatched"


def test_blank_group_creation_enters_unmatched_queue(synthetic_state: dict) -> None:
    record = create_blank_unmatched_record(actor="alice")
    records = list_unmatched_records(query=record["unmatched_id"])
    audits = list_audit_events()

    assert record["record_type"] == "blank_group"
    assert records["total"] == 1
    assert records["items"][0]["unmatched_id"] == record["unmatched_id"]
    assert audits["items"][0]["action"] == "create_blank_unmatched"


def test_group_targets_can_be_fuzzy_searched_for_manual_association(synthetic_state: dict) -> None:
    by_address = search_group_targets(query="T-001 A road")
    by_photo_field = search_group_targets(query="collector asset-1")
    by_terminal = search_group_targets(terminal="T-002")

    assert by_address["total"] == 1
    assert by_address["items"][0]["meter_no"] == "ZZ1001"
    assert by_photo_field["total"] == 1
    assert by_photo_field["items"][0]["terminal"] == "T-001"
    assert by_terminal["total"] == 1
    assert by_terminal["items"][0]["meter_no"] == "ZZ1002"


def test_unmatched_record_can_create_group_for_terminal(synthetic_state: dict) -> None:
    apply_synced_scan_records(
        [
            {
                "file_id": "manual-create-1",
                "source_file": "manual-source",
                "barcode": "MANUAL-001",
                "meter_match_key": "manual-key",
                "terminal": "",
                "collector": "collector-manual",
                "module_asset_no": "module-manual",
                "photo_urls": "https://example.test/manual-a.jpg,https://example.test/manual-b.jpg",
            }
        ]
    )
    record = list_unmatched_records(query="MANUAL-001")["items"][0]

    result = create_group_from_unmatched_record(
        record["unmatched_id"],
        actor="alice",
        terminal="T-NEW",
        updates={"address": "manual address", "meter_no": "ZZ-MANUAL"},
    )
    created = result["group"]
    tasks = list_tasks()
    audits = list_audit_events()

    assert created["terminal"] == "T-NEW"
    assert created["meter_no"] == "ZZ-MANUAL"
    assert created["address"] == "manual address"
    assert created["photo_count"] == 2
    assert any(task["terminal"] == "T-NEW" and task["can_claim"] for task in tasks)
    assert list_unmatched_records(query="MANUAL-001")["total"] == 0
    assert audits["items"][0]["action"] == "create_group_from_unmatched"


def test_unmatched_record_attaches_to_existing_terminal_group(synthetic_state: dict) -> None:
    target = synthetic_state["groups"][0]
    original_group_count = len(synthetic_state["groups"])
    synthetic_state["scan_unmatched"].append(
        {
            "unmatched_id": "manual-attach-1",
            "barcode": target["meter_no"],
            "meter_no": target["meter_no"],
            "meter_match_key": target["meter_match_key"],
            "terminal": "",
            "collector": "collector-attach",
            "module_asset_no": "module-attach",
            "image_urls": ["https://example.test/attach-a.jpg"],
        }
    )

    result = create_group_from_unmatched_record(
        "manual-attach-1",
        actor="alice",
        terminal=target["terminal"],
        updates={"address": "corrected address"},
    )
    audits = list_audit_events()

    assert result["attached"] is True
    assert result["group"]["id"] == target["id"]
    assert result["added_photos"] == 1
    assert len(synthetic_state["groups"]) == original_group_count
    assert result["group"]["address"] == "corrected address"
    assert list_unmatched_records(query="manual-attach-1")["total"] == 0
    assert audits["items"][0]["action"] == "attach_unmatched_to_existing_group"


def test_replacement_rematch_adds_delivery_export_remark(synthetic_state: dict) -> None:
    openpyxl = pytest.importorskip("openpyxl")
    target = synthetic_state["groups"][0]
    claim_task(target["task_id"], "alice")
    archive_all_group_photos(target)
    review_group(target["id"], status="approved", reviewer="alice", note="ready")
    apply_synced_scan_records(
        [
            {
                "file_id": "replacement-export-1",
                "source_file": "replacement-source",
                "barcode": "NEW-METER-001",
                "meter_no": "NEW-METER-001",
                "meter_match_key": "NEW-METER-001",
                "terminal": target["terminal"],
                "collector": "collector-replacement",
                "module_asset_no": "module-replacement",
                "photo_urls": "https://example.test/replacement-a.jpg",
            }
        ]
    )
    record = list_unmatched_records(query="NEW-METER-001")["items"][0]

    result = rematch_unmatched_record(
        record["unmatched_id"],
        actor="alice",
        meter_no="NEW-METER-001",
        old_meter_no=target["meter_no"],
        terminal=target["terminal"],
    )

    workbook = openpyxl.load_workbook(
        BytesIO(build_final_delivery_export(terminal=target["terminal"], review_scope="all")),
        read_only=True,
    )
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    remark_index = rows[0].index("备注")
    target_rows = [row for row in rows[1:] if row[1] == target["meter_no"]]

    assert result["matched"] is True
    assert target_rows
    assert any(row[remark_index] == f"换表：旧表号 {target['meter_no']}" for row in target_rows)


def test_admin_can_create_empty_group_and_import_missing_photos(synthetic_state: dict) -> None:
    result = create_empty_group_for_terminal(
        "T-MANUAL",
        actor="admin",
        meter_no="ZZ-MANUAL",
        address="manual address",
    )
    created = result["group"]

    imported = add_photo_urls_to_group(
        created["id"],
        actor="admin",
        photo_urls=["https://example.test/manual-a.jpg", "https://example.test/manual-b.jpg"],
        collector="collector-manual",
        module_asset_no="module-manual",
        creator="installer-manual",
    )
    tasks = list_tasks()
    audits = list_audit_events()

    assert created["terminal"] == "T-MANUAL"
    assert created["photo_count"] == 2
    assert created["status"] == "incomplete"
    assert imported["added"] == 2
    assert imported["group"]["photos"][0]["creator"] == "installer-manual"
    assert any(task["terminal"] == "T-MANUAL" and task["can_claim"] for task in tasks)
    assert audits["items"][0]["action"] == "add_group_photos"


def test_group_metadata_form_updates_group_and_photo_fields(synthetic_state: dict) -> None:
    group = synthetic_state["groups"][0]

    result = update_group_metadata(
        group["id"],
        actor="alice",
        updates={
            "meter_no": "ZZ-FORM",
            "address": "form address",
            "collector": "collector-form",
            "module_asset_no": "module-form",
            "creator": "installer-form",
        },
    )
    audits = list_audit_events()

    assert result["group"]["meter_no"] == "ZZ-FORM"
    assert result["group"]["address"] == "form address"
    assert all(photo["collector"] == "collector-form" for photo in result["group"]["photos"])
    assert all(photo["asset_no"] == "module-form" for photo in result["group"]["photos"])
    assert all(photo["creator"] == "installer-form" for photo in result["group"]["photos"])
    assert audits["items"][0]["action"] == "update_group_metadata"


def test_incomplete_group_is_marked_exception_after_last_photo_archived(synthetic_state: dict) -> None:
    group = synthetic_state["groups"][1]
    photo = group["photos"][0]

    claim_task(2, reviewer="alice")
    classify_photo(group["id"], photo["id"], "collector_barcode", reviewer="alice")

    assert group["status"] == "exception"
    assert group["has_archive_blocker"] is True
    assert "照片不足" in group["exception_note"]


def test_archive_blocks_duplicate_module_and_missing_required_scan_fields(synthetic_state: dict) -> None:
    clear_scan_data()
    apply_synced_scan_records(
        [
            {
                "meter_match_key": "1001",
                "barcode": f"dup-left-{index}",
                "collector": "collector-a",
                "module_asset_no": "DUP-MODULE",
                "image_urls": [f"https://example.test/left-{index}.jpg"],
            }
            for index in range(4)
        ]
        + [
            {
                "meter_match_key": "1002",
                "barcode": f"dup-right-{index}",
                "collector": "",
                "module_asset_no": "DUP-MODULE",
                "image_urls": [f"https://example.test/right-{index}.jpg"],
            }
            for index in range(4)
        ]
    )
    duplicate_group = synthetic_state["groups"][0]
    missing_group = synthetic_state["groups"][1]

    claim_task(1, reviewer="alice")
    claim_task(2, reviewer="bob")
    archive_all_group_photos(duplicate_group, reviewer="alice")
    archive_all_group_photos(missing_group, reviewer="bob")

    assert duplicate_group["status"] == "exception"
    assert "模块号重复" in duplicate_group["exception_note"]
    assert missing_group["status"] == "exception"
    assert "缺少采集器信息" in missing_group["exception_note"]

    clear_scan_data()
    apply_synced_scan_records(
        [
            {
                "meter_match_key": "1001",
                "barcode": f"missing-module-{index}",
                "collector": "collector-a",
                "module_asset_no": "",
                "image_urls": [f"https://example.test/missing-module-{index}.jpg"],
            }
            for index in range(4)
        ]
    )
    module_group = synthetic_state["groups"][0]
    claim_task(1, reviewer="alice")
    archive_all_group_photos(module_group, reviewer="alice")

    assert module_group["status"] == "exception"
    assert "缺少模块资产编号" in module_group["exception_note"]


@pytest.mark.skipif(find_spec("openpyxl") is None, reason="openpyxl is not available")
def test_import_scan_template_xlsx_reads_business_fields_and_hyperlink(synthetic_state: dict, monkeypatch: pytest.MonkeyPatch) -> None:
    from io import BytesIO

    from openpyxl import Workbook

    clear_scan_data()
    monkeypatch.setattr(
        local_simulation,
        "resolve_detail_image_urls",
        lambda url: [
            "https://example.test/barcodeImgDetail?downloadImg=cloud://photo-a.jpg",
            "https://example.test/barcodeImgDetail?downloadImg=cloud://photo-b.jpg",
        ],
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "合并"
    sheet.append(["编号", "扫码内容", "数量", "采集器", "地址", "模块资产编号", "资产类型", "创建者", "创建时间", "图片 (电脑查看)", "来自文件"])
    sheet.append(
        [
            1,
            "ABCDEFGHIJK1001X",
            1,
            "COLLECTOR-01",
            "scan address",
            "MODULE-01",
            "单相双模",
            "安装员A",
            "2026-06-09 23:08:30",
            "查看图片",
            "installer-folder",
        ]
    )
    sheet["J2"].hyperlink = "https://example.test/barcodeImgDetail?itemIdentifer=abc"
    stream = BytesIO()
    workbook.save(stream)

    result = import_scan_template_xlsx(stream.getvalue())
    group = synthetic_state["groups"][0]
    photo = group["photos"][0]

    assert result["template_rows"] == 1
    assert result["applied_records"] == 2
    assert photo["barcode"] == "ABCDEFGHIJK1001X"
    assert photo["collector"] == "COLLECTOR-01"
    assert photo["asset_no"] == "MODULE-01"
    assert photo["creator"] == "安装员A"
    assert photo["image_url"] == "https://example.test/barcodeImgDetail?downloadImg=cloud://photo-a.jpg"
    assert group["photos"][1]["image_url"] == "https://example.test/barcodeImgDetail?downloadImg=cloud://photo-b.jpg"


def test_synced_photo_urls_can_be_loaded_per_group(synthetic_state: dict) -> None:
    clear_scan_data()
    result = apply_synced_scan_records(
        [
            {
                "file_id": "remote-file-1",
                "source_file": "remote-source",
                "barcode": "ABCDEFGHIJK000001001X",
                "meter_match_key": "1001",
                "image_file_ids": ["cloud://photo-1.jpg", "cloud://photo-2.jpg"],
                "image_urls": [],
            }
        ]
    )
    group = synthetic_state["groups"][0]

    assert result["applied_records"] == 2
    assert group["photos"][0]["image_file_id"] == "cloud://photo-1.jpg"
    assert group["photos"][0]["image_url"] == ""

    loaded = apply_group_photo_urls(
        group["id"],
        {
            "cloud://photo-1.jpg": "https://download.example/photo-1.jpg",
            "cloud://photo-2.jpg": "https://download.example/photo-2.jpg",
        },
    )

    assert loaded["loaded_photo_urls"] == 2
    assert group["photos"][0]["image_url"] == "https://download.example/photo-1.jpg"
    assert group["photos"][1]["image_url"] == "https://download.example/photo-2.jpg"


def test_local_test_routes_cover_review_flow(synthetic_state: dict) -> None:
    client = TestClient(create_app())

    claim_response = client.post("/local-test/tasks/1/claim", json={"reviewer": "alice"})
    groups_response = client.get("/local-test/tasks/1/groups?limit=1")
    group_id = groups_response.json()["data"]["items"][0]["id"]
    review_response = client.patch(
        f"/local-test/groups/{group_id}/review",
        json={"status": "approved", "reviewer": "alice", "note": "ok"},
    )
    progress_response = client.get("/local-test/tasks/1/progress")

    assert claim_response.status_code == 200
    assert groups_response.status_code == 200
    assert groups_response.json()["data"]["total"] == 1
    assert review_response.status_code == 200
    assert review_response.json()["data"]["status"] == "approved"
    assert progress_response.json()["data"]["reviewed_groups"] == 1
    assert progress_response.json()["data"]["completeness_rate"] == 1.0


@requires_sample_workbooks
def test_sample_paths_are_absolute_windows_files() -> None:
    for path in SAMPLE_FILES:
        assert isinstance(path, Path)
        assert path.suffix == ".xlsx"
