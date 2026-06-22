from __future__ import annotations

import html
import hashlib
import json
import mimetypes
import os
import re
import threading
import urllib.request
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from contextvars import ContextVar
from dataclasses import dataclass
from datetime import UTC, datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path
from typing import Any, Callable
from urllib.parse import parse_qsl, urlencode, urljoin, urlparse, urlunparse

from app.core.config import settings
from app.services.matching import build_long_scan_match_key, build_total_catalog_match_key
from app.services.photo_storage import (
    active_storage_backend,
    parse_oss_image_url,
    resolve_photo_preview_url,
    save_image_bytes,
    sign_oss_server_url,
    static_upload_root,
    validate_image_content,
)


DEFAULT_TOTAL_CATALOG = Path("C:/Users/Administrator/Desktop/\u603b\u4f53\u6570\u636e.xlsx")
DEFAULT_STAGE_CATALOG = Path("C:/Users/Administrator/Desktop/\u7b2c\u4e00\u6279\u6570\u636e.xlsx")
LOCAL_WORK_TZ = timezone(timedelta(hours=8))
WORK_SESSION_BREAK_MINUTES = 60
DEFAULT_SCAN_FILE = Path("C:/Users/Administrator/Desktop/\u6279\u91cf\u626b\u7801_20260608125555.xlsx")
REVIEWABLE_STATUSES = {"pending", "incomplete", "approved", "exception", "unmatched"}
OPEN_STATUSES = {"pending", "incomplete", "unmatched"}
DONE_STATUSES = {"approved", "exception"}
MAX_ACTIVE_CONSTRUCTION_TASKS_PER_CONSTRUCTOR = 5
PHOTO_CATEGORIES = {
    "unclassified": "\u672a\u5206\u7c7b",
    "before_box": "\u8868\u7bb1\u6574\u4f53\u6539\u9020\u524d",
    "collector_barcode": "\u91c7\u96c6\u5668\u6761\u5f62\u7801",
    "module_meter": "\u6a21\u5757\u4e0e\u7535\u80fd\u8868",
    "after_box": "\u8868\u7bb1\u6574\u4f53\u6539\u9020\u540e",
    "unmatched_group": "\u672a\u5339\u914d\u6570\u636e\u7ec4",
    "other": "\u5176\u4ed6",
}
CONSTRUCTION_SLOT_CATEGORIES = {
    "before_box": "before_box",
    "collector_barcode": "collector_barcode",
    "module_meter": "module_meter",
    "after_box": "after_box",
    "other": "other",
}
VOLATILE_URL_QUERY_KEYS = {
    "access_token",
    "expires",
    "expire",
    "expires_in",
    "signature",
    "sign",
    "token",
    "x-oss-signature",
    "x-oss-expires",
    "x-oss-date",
    "x-oss-credential",
    "x-oss-security-token",
}
CONSTRUCTION_EXCEPTION_CATEGORIES = {
    "meter_error": "表号错误",
    "collector_error": "采集器号错误",
    "module_error": "模块号错误",
    "photo_error": "照片错误",
    "missing_photo": "照片缺失",
    "site_blocked": "现场无法施工",
    "device_mismatch": "设备不符",
    "other": "其他",
}
IMAGE_SRC_RE = re.compile(r"<img\b[^>]*(?:src|data-src)=['\"]([^'\"]+)['\"]", re.IGNORECASE)
DEFAULT_TEAM_ID = "default-team"


@dataclass(frozen=True)
class LocalTestPaths:
    total_catalog: Path = DEFAULT_TOTAL_CATALOG
    stage_catalog: Path = DEFAULT_STAGE_CATALOG
    scan_file: Path = DEFAULT_SCAN_FILE


_current_team_id: ContextVar[str] = ContextVar("local_simulation_team_id", default=DEFAULT_TEAM_ID)


def blank_state(team_id: str = DEFAULT_TEAM_ID) -> dict[str, Any]:
    return {
        "team_id": team_id,
        "loaded": False,
        "paths": {},
        "summary": {},
        "projects": [],
        "tasks": [],
        "groups": [],
        "total_catalog": [],
        "stage_catalog": [],
        "scan_unmatched": [],
        "stage_unmatched": [],
        "construction_exception_orders": [],
        "review_events": [],
        "photo_events": [],
        "audit_events": [],
    }


def empty_summary() -> dict[str, Any]:
    return {
        "team_id": DEFAULT_TEAM_ID,
        "total_catalog_rows": 0,
        "stage_catalog_rows": 0,
        "scan_rows": 0,
        "groups": 0,
        "matched_groups": 0,
        "incomplete_groups": 0,
        "unconstructed_groups": 0,
        "approved_groups": 0,
        "exception_groups": 0,
        "reviewed_groups": 0,
        "unreviewed_groups": 0,
        "stage_unmatched": 0,
        "scan_unmatched": 0,
        "photo_rows_linked": 0,
        "scanned_groups": 0,
        "installer_distribution": [],
        "downloaded_photos": 0,
        "unclassified_photos": 0,
        "review_progress": 0.0,
    }


_team_states: dict[str, dict[str, Any]] = {}
_state: dict[str, Any] = blank_state(DEFAULT_TEAM_ID)
_state["summary"] = empty_summary()
_team_states[DEFAULT_TEAM_ID] = _state
_persistence_lock = threading.RLock()
_delivery_cache_executor = ThreadPoolExecutor(max_workers=2)
_delivery_cache_lock = threading.RLock()
_delivery_cache_inflight: set[tuple[str, str]] = set()


def persisted_state_path() -> Path | None:
    configured = os.getenv("LOCAL_SIMULATION_STATE_PATH", "").strip()
    if configured:
        return Path(configured)
    return None


def save_all_team_states() -> None:
    path = persisted_state_path()
    if path is None:
        return
    payload = {
        "version": 1,
        "saved_at": datetime.now(UTC).isoformat(),
        "teams": _team_states,
    }
    with _persistence_lock:
        path.parent.mkdir(parents=True, exist_ok=True)
        tmp_path = path.with_name(f"{path.name}.tmp")
        tmp_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
        tmp_path.replace(path)


def load_all_team_states() -> bool:
    path = persisted_state_path()
    if path is None:
        return False
    if not path.exists():
        return False
    with _persistence_lock:
        payload = json.loads(path.read_text(encoding="utf-8"))
        teams = payload.get("teams") if isinstance(payload, dict) else None
        if not isinstance(teams, dict):
            return False
        next_states: dict[str, dict[str, Any]] = {}
        for team_id, state in teams.items():
            if not isinstance(state, dict):
                continue
            normalized_team_id = normalize_team_id(str(team_id))
            state["team_id"] = normalized_team_id
            state.setdefault("summary", empty_summary())
            next_states[normalized_team_id] = state
        if not next_states:
            return False
        _team_states.clear()
        _team_states.update(next_states)
        if DEFAULT_TEAM_ID not in _team_states:
            _team_states[DEFAULT_TEAM_ID] = blank_state(DEFAULT_TEAM_ID)
            _team_states[DEFAULT_TEAM_ID]["summary"] = empty_summary()
        global _state
        _state = _team_states[DEFAULT_TEAM_ID]
        return True


def normalize_team_id(team_id: str | None) -> str:
    value = re.sub(r"[^0-9A-Za-z_-]+", "-", str(team_id or "").strip()).strip("-").lower()
    return value or DEFAULT_TEAM_ID


def set_current_team(team_id: str | None):
    return _current_team_id.set(normalize_team_id(team_id))


def reset_current_team(token) -> None:
    _current_team_id.reset(token)


def current_team_id() -> str:
    return normalize_team_id(_current_team_id.get())


load_all_team_states()


def state_for_team(team_id: str | None = None) -> dict[str, Any]:
    team = normalize_team_id(team_id or current_team_id())
    if team not in _team_states:
        _team_states[team] = blank_state(team)
        _team_states[team]["summary"] = empty_summary()
    _team_states[team].setdefault("summary", empty_summary())
    _team_states[team]["summary"]["team_id"] = team
    return _team_states[team]


def list_team_states() -> list[dict[str, Any]]:
    return [
        {
            "team_id": team_id,
            "loaded": state.get("loaded", False),
            "groups": len(state.get("groups", [])),
            "tasks": len(state.get("tasks", [])),
            "summary": state.get("summary", {}),
        }
        for team_id, state in sorted(_team_states.items())
    ]


def get_state() -> dict[str, Any]:
    return state_for_team()


def clear_scan_data() -> dict[str, Any]:
    state = get_state()
    for group in state["groups"]:
        group["photos"] = []
        group["photo_count"] = 0
        if group["status"] != "unmatched":
            group["status"] = "incomplete"
        group["reviewer"] = None
        group["review_note"] = ""
        group["exception_note"] = ""
        group["reviewed_at"] = None
    state["scan_unmatched"] = []
    state["photo_events"] = []
    state["summary"]["scan_rows"] = 0
    refresh_summary()
    return state


ProgressCallback = Callable[[dict[str, Any]], None]
OSS_SYNC_MAX_BYTES = 30 * 1024 * 1024


def apply_synced_scan_records(
    records: list[dict[str, Any]],
    progress_callback: ProgressCallback | None = None,
) -> dict[str, Any]:
    state = get_state()
    groups_by_key = {group["meter_match_key"]: group for group in state["groups"]}
    known_photo_keys = {
        make_photo_unique_key(photo)
        for group in state["groups"]
        for photo in group.get("photos", []) + group.get("deleted_photos", [])
    }
    applied = 0
    skipped_duplicates = 0
    skipped_duplicate_meters = 0
    unmatched = []
    processed_photos = 0
    groups_matched: set[str] = set()
    groups_unmatched = 0
    groups_existing: set[str] = set()
    for index, record in enumerate(records, start=1):
        match_key = str(record.get("meter_match_key") or "")
        group = groups_by_key.get(match_key)
        if group is None:
            unmatched.append(ensure_unmatched_record(record))
            groups_unmatched += 1
            continue
        groups_matched.add(str(group.get("id") or match_key))
        if group.get("photo_count", 0) > 0:
            groups_existing.add(str(group.get("id") or match_key))
        rows = scan_record_to_photo_rows(record, index)
        group_changed = False
        for row in rows:
            processed_photos += 1
            unique_key = make_scan_unique_key(row)
            if unique_key in known_photo_keys:
                skipped_duplicates += 1
                if progress_callback and processed_photos % 5 == 0:
                    progress_callback(
                        {
                            "phase": "linking_photos",
                            "processed_records": index,
                            "total_records": len(records),
                            "processed_photos": processed_photos,
                            "applied_records": applied,
                            "skipped_duplicates": skipped_duplicates,
                            "skipped_duplicate_meters": skipped_duplicate_meters,
                            "photos_seen": processed_photos,
                            "photos_new": applied,
                            "photos_duplicate": skipped_duplicates,
                        }
                    )
                continue
            photo = build_photo_record(group["photo_count"] + 1, row)
            group["photos"].append(photo)
            group["photo_count"] = len(group["photos"])
            known_photo_keys.add(unique_key)
            applied += 1
            group_changed = True
            if progress_callback and processed_photos % 5 == 0:
                progress_callback(
                    {
                        "phase": "linking_photos",
                        "processed_records": index,
                        "total_records": len(records),
                            "processed_photos": processed_photos,
                            "applied_records": applied,
                            "skipped_duplicates": skipped_duplicates,
                            "skipped_duplicate_meters": skipped_duplicate_meters,
                            "photos_seen": processed_photos,
                            "photos_new": applied,
                            "photos_duplicate": skipped_duplicates,
                        }
                    )
        if group_changed:
            group["status"] = "incomplete" if group["photo_count"] < 4 else "pending"
            group["reviewer"] = None
            group["review_note"] = ""
            group["exception_note"] = ""
            group["reviewed_at"] = None
    state["scan_unmatched"] = merge_unmatched_records(state.get("scan_unmatched", []), unmatched)
    state["summary"]["scan_rows"] = sum(group["photo_count"] for group in state["groups"]) + len(unmatched)
    refresh_summary()
    if progress_callback:
        progress_callback(
            {
                "phase": "complete",
                "processed_records": len(records),
                "total_records": len(records),
                "processed_photos": processed_photos,
                "applied_records": applied,
                "skipped_duplicates": skipped_duplicates,
                "skipped_duplicate_meters": skipped_duplicate_meters,
                "photos_seen": processed_photos,
                "photos_new": applied,
                "photos_duplicate": skipped_duplicates,
            }
        )
    return {
        "rows_total": len(records),
        "received_records": len(records),
        "groups_matched": len(groups_matched),
        "groups_unmatched": groups_unmatched,
        "groups_existing": len(groups_existing),
        "photos_seen": processed_photos,
        "photos_new": applied,
        "photos_duplicate": skipped_duplicates,
        "photos_reused_oss": 0,
        "photos_uploaded_oss": 0,
        "photos_failed": 0,
        "applied_records": applied,
        "skipped_duplicates": skipped_duplicates,
        "skipped_duplicate_meters": skipped_duplicate_meters,
        "unmatched_records": len(unmatched),
        "summary": state["summary"],
    }


def import_url_scan_rows(rows: list[dict[str, Any]], progress_callback: ProgressCallback | None = None) -> dict[str, Any]:
    records = [normalize_url_import_row(row, index) for index, row in enumerate(rows, start=1)]
    return apply_synced_scan_records(records, progress_callback=progress_callback)


def import_scan_template_xlsx(content: bytes, progress_callback: ProgressCallback | None = None) -> dict[str, Any]:
    rows = read_scan_template_xlsx_rows(content)
    if progress_callback:
        progress_callback({"phase": "parsed", "template_rows": len(rows), "processed_records": 0, "total_records": len(rows)})
    detail_stats = expand_detail_pages_for_rows(rows, progress_callback=progress_callback)
    result = import_url_scan_rows(rows, progress_callback=progress_callback)
    result["template_rows"] = len(rows)
    result.update(detail_stats)
    return result


def import_total_catalog_xlsx(content: bytes) -> dict[str, Any]:
    state = get_state()
    incoming_rows = read_catalog_xlsx_rows(content, source="total")
    existing_keys = {str(row.get("meter_match_key") or "") for row in state.get("total_catalog", [])}
    seen_keys: set[str] = set()
    accepted_rows: list[dict[str, Any]] = []
    duplicate_rows: list[dict[str, Any]] = []
    for row in incoming_rows:
        key = str(row.get("meter_match_key") or "")
        if not key or key in existing_keys or key in seen_keys:
            duplicate_rows.append(row)
            continue
        seen_keys.add(key)
        accepted_rows.append(row)

    total_rows = [*state.get("total_catalog", []), *accepted_rows]
    rebuild_state_from_total_catalog(total_rows, preserve_existing=True)
    state = get_state()
    return {
        "catalog_rows": len(incoming_rows),
        "imported_rows": len(accepted_rows),
        "skipped_duplicate_meters": len(duplicate_rows),
        "summary": state["summary"],
    }


def read_scan_template_xlsx_rows(content: bytes) -> list[dict[str, Any]]:
    load_workbook = get_workbook_loader()
    workbook = load_workbook(BytesIO(content), read_only=False, data_only=True)
    sheet = workbook.worksheets[0]
    header_cells = next(sheet.iter_rows(min_row=1, max_row=1, values_only=False))
    headers = [normalize_cell(cell.value) for cell in header_cells]
    rows: list[dict[str, Any]] = []
    blank_run = 0
    for row_number, cells in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
        values: dict[str, Any] = {"row_number": row_number}
        for index, cell in enumerate(cells):
            if index >= len(headers) or not headers[index]:
                continue
            header = headers[index]
            value = normalize_cell(cell.value)
            values[header] = value
            if cell.hyperlink and cell.hyperlink.target:
                values[f"{header}_url"] = cell.hyperlink.target
                if "\u56fe\u7247" in header:
                    values["photo_urls"] = cell.hyperlink.target
        if not any(str(value).strip() for key, value in values.items() if key != "row_number"):
            blank_run += 1
            if blank_run >= 20:
                break
            continue
        blank_run = 0
        rows.append(values)
    return rows


def expand_detail_pages_for_rows(
    rows: list[dict[str, Any]],
    max_workers: int = 16,
    progress_callback: ProgressCallback | None = None,
) -> dict[str, int]:
    detail_urls: set[str] = set()
    for row in rows:
        for url in split_urls(pick_photo_url_field(row)):
            if is_photo_detail_page(url):
                detail_urls.add(url)
    if not detail_urls:
        return {"detail_pages": 0, "resolved_detail_pages": 0, "resolved_image_urls": 0}

    resolved_by_url: dict[str, list[str]] = {}
    completed_detail_pages = 0
    running_resolved_image_urls = 0
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(resolve_detail_image_urls, url): url for url in detail_urls}
        for future in as_completed(futures):
            url = futures[future]
            try:
                resolved_by_url[url] = future.result()
            except OSError:
                resolved_by_url[url] = []
            completed_detail_pages += 1
            running_resolved_image_urls += len(resolved_by_url[url])
            if progress_callback and (completed_detail_pages % 5 == 0 or completed_detail_pages == len(detail_urls)):
                progress_callback(
                    {
                        "phase": "resolving_photos",
                        "processed_detail_pages": completed_detail_pages,
                        "total_detail_pages": len(detail_urls),
                        "resolved_image_urls": running_resolved_image_urls,
                    }
                )

    resolved_image_urls = 0
    resolved_detail_pages = 0
    for row in rows:
        expanded: list[str] = []
        seen: set[str] = set()
        for url in split_urls(pick_photo_url_field(row)):
            candidates = resolved_by_url.get(url) if is_photo_detail_page(url) else None
            if candidates:
                resolved_detail_pages += 1
                resolved_image_urls += len(candidates)
            for candidate in candidates or [url]:
                if candidate not in seen:
                    expanded.append(candidate)
                    seen.add(candidate)
        if expanded:
            row["photo_urls"] = "\n".join(expanded)
    return {
        "detail_pages": len(detail_urls),
        "resolved_detail_pages": resolved_detail_pages,
        "resolved_image_urls": resolved_image_urls,
    }


def pick_photo_url_field(row: dict[str, Any]) -> str:
    return pick_first(
        row,
        "photo_urls",
        "image_urls",
        "\u7167\u7247URL",
        "\u56fe\u7247URL",
        "\u56fe\u7247 (\u7535\u8111\u67e5\u770b)_url",
        "\u56fe\u7247(\u7535\u8111\u67e5\u770b)_url",
        "url",
        "URL",
    )


def normalize_url_import_row(row: dict[str, Any], index: int) -> dict[str, Any]:
    barcode = pick_first(row, "barcode", "scan_code", "扫码内容", "条形码", "长条码")
    meter_no = pick_first(row, "meter_no", "表号", "短表号")
    if not meter_no and barcode:
        meter_no = barcode
    meter_match_key = pick_first(row, "meter_match_key", "匹配键")
    if not meter_match_key:
        if barcode:
            try:
                meter_match_key = build_long_scan_match_key(barcode)
            except ValueError:
                meter_match_key = ""
        elif meter_no:
            meter_match_key = build_total_catalog_match_key(meter_no)
    image_urls = expand_photo_urls(split_urls(pick_photo_url_field(row)))
    return {
        "file_id": f"import-row-{index}",
        "source_file": pick_first(row, "source_file", "来源文件", "来自文件", "批次") or "url-import",
        "installer": pick_first(row, "installer", "安装人员", "创建者"),
        "barcode": barcode or meter_no or f"row-{index}",
        "meter_match_key": meter_match_key,
        "terminal": pick_first(row, "terminal", "终端"),
        "collector": pick_first(row, "collector", "采集器"),
        "meter_no": meter_no,
        "module_asset_no": pick_first(row, "module_asset_no", "module", "模块", "模块资产编号"),
        "address": pick_first(row, "address", "地址"),
        "asset_type": pick_first(row, "asset_type", "资产类型"),
        "creator": pick_first(row, "creator", "创建者"),
        "created_at": pick_first(row, "created_at", "创建时间"),
        "image_count": len(image_urls),
        "image_urls": image_urls,
        "image_file_ids": [],
    }


def pick_first(row: dict[str, Any], *keys: str) -> str:
    normalized = {str(key).strip().lower(): value for key, value in row.items()}
    for key in keys:
        value = row.get(key)
        if value is None:
            value = normalized.get(key.lower())
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


def split_urls(value: str) -> list[str]:
    if not value:
        return []
    normalized = value.replace("\r", "\n").replace("；", ";").replace("，", ",")
    parts = []
    for block in normalized.split("\n"):
        for segment in block.replace(";", ",").split(","):
            item = segment.strip()
            if item:
                parts.append(item)
    return parts


def expand_photo_urls(urls: list[str]) -> list[str]:
    expanded: list[str] = []
    seen: set[str] = set()
    for url in urls:
        resolved = resolve_detail_image_urls(url) if is_photo_detail_page(url) else []
        candidates = resolved or [url]
        for candidate in candidates:
            if candidate not in seen:
                expanded.append(candidate)
                seen.add(candidate)
    return expanded


def is_photo_detail_page(url: str) -> bool:
    return "barcodeImgDetail" in url and "itemIdentifer=" in url


def resolve_detail_image_urls(detail_url: str, timeout: int = 10) -> list[str]:
    try:
        request = urllib.request.Request(detail_url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(request, timeout=timeout) as response:
            body = response.read(2_000_000)
    except OSError:
        return []
    text = body.decode("utf-8", errors="replace")
    urls: list[str] = []
    seen: set[str] = set()
    for raw_src in IMAGE_SRC_RE.findall(text):
        src = html.unescape(raw_src.strip())
        if not src or src.startswith("data:"):
            continue
        image_url = urljoin(detail_url, src)
        if not looks_like_review_photo(image_url) or image_url in seen:
            continue
        urls.append(image_url)
        seen.add(image_url)
    return urls


def looks_like_review_photo(url: str) -> bool:
    lower = url.lower()
    return "downloadimg=" in lower or "scan_photos" in lower or lower.split("?", 1)[0].endswith((".jpg", ".jpeg", ".png", ".webp", ".bmp"))


def infer_photo_storage(image_url: str) -> dict[str, str]:
    url = str(image_url or "").strip()
    if not url:
        return {"storage_type": "", "storage_key": ""}
    parsed = urlparse(url)
    if url.startswith("/static/uploads/"):
        return {"storage_type": "local_upload", "storage_key": url.removeprefix("/static/uploads/")}
    if parsed.scheme == "oss":
        return {"storage_type": "oss", "storage_key": parsed.path.lstrip("/")}
    if parsed.scheme in {"http", "https"}:
        return {"storage_type": "external_url", "storage_key": url}
    if url.startswith("/"):
        return {"storage_type": "local_upload", "storage_key": url.lstrip("/")}
    return {"storage_type": "external_url", "storage_key": url}


def normalized_photo_source_url(image_url: str) -> str:
    raw = str(image_url or "").strip()
    if not raw:
        return ""
    parsed = urlparse(raw)
    query_items = parse_qsl(parsed.query, keep_blank_values=True)
    for key, value in query_items:
        if key.lower() == "downloadimg" and value:
            return value.strip()
    if parsed.scheme in {"http", "https"}:
        kept = [
            (key, value)
            for key, value in query_items
            if key.lower() not in VOLATILE_URL_QUERY_KEYS
        ]
        return urlunparse(
            (
                parsed.scheme.lower(),
                parsed.netloc.lower(),
                parsed.path,
                "",
                urlencode(kept, doseq=True),
                "",
            )
        )
    return raw


def hash_text(value: str) -> str:
    return hashlib.sha256(str(value or "").encode("utf-8")).hexdigest()


def source_fingerprint_from_values(
    *,
    image_file_id: str = "",
    image_url: str = "",
    barcode: str = "",
    source_file: str = "",
    photo_index: str | int = "",
    row_number: str = "",
) -> str:
    file_id = str(image_file_id or "").strip()
    if file_id:
        return f"file:{hash_text(file_id)}"
    normalized_url = normalized_photo_source_url(image_url)
    if normalized_url:
        return f"url:{hash_text(normalized_url)}"
    fallback = "|".join(
        [
            str(barcode or "").strip(),
            str(source_file or "").strip(),
            str(photo_index or "").strip(),
            str(row_number or "").strip(),
        ]
    )
    return f"row:{hash_text(fallback)}"


def ensure_photo_identity_fields(photo: dict[str, Any]) -> dict[str, Any]:
    source_url = str(photo.get("source_url") or photo.get("pre_oss_image_url") or photo.get("image_url") or "").strip()
    image_file_id = str(photo.get("source_file_id") or photo.get("image_file_id") or "").strip()
    source_url_normalized = normalized_photo_source_url(source_url)
    photo.setdefault("source_url", source_url)
    photo.setdefault("source_url_hash", hash_text(source_url_normalized) if source_url_normalized else "")
    photo.setdefault("source_file_id", image_file_id)
    photo.setdefault(
        "source_fingerprint",
        source_fingerprint_from_values(
            image_file_id=image_file_id,
            image_url=source_url,
            barcode=str(photo.get("barcode") or ""),
            source_file=str(photo.get("source_file") or ""),
            photo_index=str(photo.get("row_number") or ""),
            row_number=str(photo.get("row_number") or ""),
        ),
    )
    photo.setdefault("import_batch_id", "")
    photo.setdefault("is_active", True)
    photo.setdefault("deleted_at", None)
    photo.setdefault("deleted_by", "")
    photo.setdefault("delete_reason", "")
    return photo


def ensure_photo_storage_fields(photo: dict[str, Any]) -> dict[str, Any]:
    storage = infer_photo_storage(str(photo.get("image_url") or ""))
    photo.setdefault("storage_type", storage["storage_type"])
    photo.setdefault("storage_key", storage["storage_key"])
    photo.setdefault("storage_bucket", "")
    photo.setdefault("storage_source", photo.get("source_file") or storage["storage_type"])
    photo.setdefault("sha256", "")
    ensure_photo_identity_fields(photo)
    return photo


def photo_needs_oss_sync(photo: dict[str, Any]) -> bool:
    ensure_photo_storage_fields(photo)
    image_url = str(photo.get("image_url") or "").strip()
    storage_type = str(photo.get("storage_type") or "").strip()
    storage_key = str(photo.get("storage_key") or "").strip()
    if storage_type == "oss" or image_url.startswith("oss://"):
        return False
    return bool(image_url or storage_key)


def photo_source_filename(photo: dict[str, Any]) -> str:
    for value in (
        photo.get("original_filename"),
        photo.get("archive_filename"),
        photo.get("storage_key"),
        photo.get("image_url"),
        photo.get("image_file_id"),
        photo.get("id"),
    ):
        text = str(value or "").split("?", 1)[0].strip().rstrip("/")
        if not text:
            continue
        name = Path(text).name
        if Path(name).suffix.lower() in {".jpg", ".jpeg", ".png", ".webp", ".bmp"}:
            return name
    return f"{photo.get('id') or 'photo'}.jpg"


def local_photo_candidates(photo: dict[str, Any]) -> list[Path]:
    ensure_photo_storage_fields(photo)
    values = [
        str(photo.get("storage_key") or ""),
        str(photo.get("image_url") or ""),
        str(photo.get("object_key") or ""),
    ]
    paths: list[Path] = []
    root = static_upload_root()
    for value in values:
        text = value.strip()
        if not text:
            continue
        if text.startswith("/static/uploads/"):
            text = text.removeprefix("/static/uploads/")
        text = text.lstrip("/").removeprefix("static/uploads/").removeprefix("uploads/")
        if Path(text).suffix.lower() not in {".jpg", ".jpeg", ".png", ".webp", ".bmp"}:
            continue
        paths.append(root / text)
    return paths


def load_photo_content_for_oss(photo: dict[str, Any]) -> tuple[bytes, str, str, str]:
    ensure_photo_storage_fields(photo)
    storage_type = str(photo.get("storage_type") or "").strip()
    image_url = str(photo.get("image_url") or "").strip()
    parsed = urlparse(image_url)
    if storage_type == "local_upload" or image_url.startswith("/static/uploads/") or (
        storage_type and parsed.scheme not in {"http", "https", "oss"}
    ):
        for path in local_photo_candidates(photo):
            if path.exists() and path.is_file():
                content = path.read_bytes()
                validate_image_content(content, mimetypes.guess_type(path.name)[0] or "", str(path))
                return (
                    content,
                    mimetypes.guess_type(path.name)[0] or "application/octet-stream",
                    "local_upload",
                    str(path),
                )
        checked = ", ".join(str(path) for path in local_photo_candidates(photo)) or "(none)"
        raise FileNotFoundError(f"Local photo file not found; checked: {checked}")
    if parsed.scheme in {"http", "https"}:
        request = urllib.request.Request(
            image_url,
            headers={
                "User-Agent": "module-manager-v2-oss-sync/1.0",
                "Accept": "image/*,*/*;q=0.8",
            },
        )
        with urllib.request.urlopen(request, timeout=25) as response:
            content_type = response.headers.get("Content-Type", "") or "application/octet-stream"
            expected_length = response.headers.get("Content-Length", "")
            content = response.read(OSS_SYNC_MAX_BYTES + 1)
        if len(content) > OSS_SYNC_MAX_BYTES:
            raise ValueError("Photo exceeds OSS sync max size")
        if not content:
            raise ValueError("Downloaded empty photo")
        if expected_length and expected_length.isdigit() and len(content) != int(expected_length):
            raise ValueError("Downloaded photo is incomplete")
        validate_image_content(content, content_type, image_url)
        return content, content_type, "external_url", image_url
    raise ValueError("Photo has no supported source for OSS sync")


def delivery_cache_root() -> Path:
    configured = str(settings.delivery_cache_path or "").strip()
    if configured:
        return Path(configured)
    state_path = persisted_state_path()
    if state_path is not None:
        return state_path.parent / "delivery_cache"
    return Path(__file__).resolve().parents[2] / "data" / "delivery_cache"


def safe_cache_part(value: str, fallback: str = "item") -> str:
    text = re.sub(r"[^0-9A-Za-z_-]+", "-", str(value or "").strip()).strip("-")
    return (text or fallback)[:96]


def delivery_photo_cache_version(photo: dict[str, Any]) -> str:
    ensure_photo_storage_fields(photo)
    payload = {
        "id": photo.get("id", ""),
        "image_url": photo.get("image_url", ""),
        "storage_type": photo.get("storage_type", ""),
        "storage_key": photo.get("storage_key", ""),
        "sha256": photo.get("sha256", ""),
        "category": photo.get("category", ""),
        "archive_filename": photo.get("archive_filename", ""),
        "is_active": photo.get("is_active", True),
    }
    return hashlib.sha256(json.dumps(payload, ensure_ascii=False, sort_keys=True).encode("utf-8")).hexdigest()


def delivery_group_cache_version(group: dict[str, Any]) -> str:
    versions = [
        delivery_photo_cache_version(photo)
        for photo in group.get("photos", [])
        if photo.get("is_active", True)
    ]
    return hashlib.sha256("|".join(versions).encode("utf-8")).hexdigest()


def delivery_cache_relative_path(group: dict[str, Any], photo: dict[str, Any], version: str, suffix: str) -> Path:
    team = safe_cache_part(current_team_id(), DEFAULT_TEAM_ID)
    group_id = safe_cache_part(str(group.get("id") or ""), "group")
    photo_id = safe_cache_part(str(photo.get("id") or ""), "photo")
    suffix = suffix.lower() if suffix.lower() in {".jpg", ".jpeg", ".png", ".webp", ".bmp"} else ".jpg"
    return Path(team) / group_id / f"{photo_id}-{version[:16]}{suffix}"


def delivery_cache_file_for_photo(group: dict[str, Any], photo: dict[str, Any]) -> Path | None:
    rel = str(photo.get("delivery_cache_path") or "").strip()
    if not rel:
        return None
    root = delivery_cache_root().resolve()
    candidate = (root / rel).resolve()
    try:
        candidate.relative_to(root)
    except ValueError:
        return None
    return candidate


def delivery_cache_url_for_photo(group: dict[str, Any], photo: dict[str, Any]) -> str:
    path = delivery_cache_file_for_photo(group, photo)
    version = str(photo.get("delivery_cache_version") or "")
    if not path or not version or not path.exists():
        return ""
    if version != delivery_photo_cache_version(photo):
        return ""
    return f"/local-test/delivery-cache/{group.get('id')}/{photo.get('id')}?v={version[:16]}"


def group_delivery_cache_ready(group: dict[str, Any]) -> bool:
    photos = [photo for photo in group.get("photos", []) if photo.get("is_active", True)]
    return bool(photos) and all(delivery_cache_url_for_photo(group, photo) for photo in photos)


def get_delivery_cached_photo_path(group_id: str, photo_id: str) -> Path:
    group = get_group(group_id)
    if group is None:
        raise KeyError(group_id)
    photo = next((item for item in group.get("photos", []) if item.get("id") == photo_id), None)
    if photo is None:
        raise KeyError(photo_id)
    path = delivery_cache_file_for_photo(group, photo)
    if not path or not path.exists() or photo.get("delivery_cache_version") != delivery_photo_cache_version(photo):
        raise FileNotFoundError(photo_id)
    return path


def photo_cache_download_url(photo: dict[str, Any]) -> str:
    ensure_photo_storage_fields(photo)
    storage_type = str(photo.get("storage_type") or "").strip()
    image_url = str(photo.get("image_url") or "").strip()
    storage_key = str(photo.get("storage_key") or "").strip()
    if storage_type == "oss" or image_url.startswith("oss://"):
        _, key = parse_oss_image_url(image_url)
        return sign_oss_server_url(storage_key or key, settings.oss_preview_process)
    return resolve_photo_preview_url(photo)


def download_delivery_photo_content(photo: dict[str, Any]) -> tuple[bytes, str, str]:
    for path in local_photo_candidates(photo):
        if path.exists() and path.is_file():
            content = path.read_bytes()
            if content:
                return content, Path(path).suffix.lower() or ".jpg", mimetypes.guess_type(path.name)[0] or "image/jpeg"
    url = photo_cache_download_url(photo)
    if not url or url.startswith("/"):
        raise ValueError("Photo has no downloadable delivery cache source")
    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": "module-manager-v2-delivery-cache/1.0",
            "Accept": "image/*,*/*;q=0.8",
        },
    )
    with urllib.request.urlopen(request, timeout=30) as response:
        content_type = response.headers.get("Content-Type", "") or "image/jpeg"
        expected_length = response.headers.get("Content-Length", "")
        content = response.read(OSS_SYNC_MAX_BYTES + 1)
    if len(content) > OSS_SYNC_MAX_BYTES:
        raise ValueError("Photo exceeds delivery cache max size")
    if not content:
        raise ValueError("Downloaded empty delivery cache photo")
    if expected_length and expected_length.isdigit() and len(content) != int(expected_length):
        raise ValueError("Downloaded delivery cache photo is incomplete")
    validate_image_content(content, content_type, url)
    suffix = mimetypes.guess_extension(content_type.split(";", 1)[0].strip()) or Path(urlparse(url).path).suffix or ".jpg"
    if suffix == ".jpe":
        suffix = ".jpg"
    return content, suffix, content_type


def photo_can_build_delivery_cache(photo: dict[str, Any]) -> bool:
    ensure_photo_storage_fields(photo)
    image_url = str(photo.get("image_url") or "").strip()
    storage_type = str(photo.get("storage_type") or "").strip()
    if storage_type == "oss" or image_url.startswith("oss://"):
        return True
    if storage_type == "local_upload" or image_url.startswith("/static/uploads/"):
        return any(path.exists() and path.is_file() for path in local_photo_candidates(photo))
    return False


def mark_delivery_cache_stale(group: dict[str, Any], reason: str = "") -> None:
    if not group:
        return
    if group.get("delivery_cache_status") in {"ready", "building", "partial_failed", "failed", "stale"}:
        group["delivery_cache_status"] = "stale"
        group["delivery_cache_error"] = reason
    for photo in group.get("photos", []):
        if photo.get("delivery_cache_path"):
            photo["delivery_cache_status"] = "stale"


def build_delivery_cache_for_group(group_id: str, force: bool = False) -> dict[str, Any]:
    group = get_group(group_id)
    if group is None:
        raise KeyError(group_id)
    if not is_reviewed_group(group):
        mark_delivery_cache_stale(group, "group is not approved")
        return {"status": "skipped", "reason": "group is not approved", "group_id": group_id}
    photos = [photo for photo in group.get("photos", []) if photo.get("is_active", True)]
    if not photos:
        mark_delivery_cache_stale(group, "group has no active photos")
        return {"status": "skipped", "reason": "group has no active photos", "group_id": group_id}

    root = delivery_cache_root()
    group["delivery_cache_status"] = "building"
    group["delivery_cache_error"] = ""
    group["delivery_cache_version"] = delivery_group_cache_version(group)
    built = 0
    reused = 0
    failed: list[dict[str, str]] = []

    for photo in photos:
        version = delivery_photo_cache_version(photo)
        existing = delivery_cache_file_for_photo(group, photo)
        if not force and existing and existing.exists() and photo.get("delivery_cache_version") == version:
            photo["delivery_cache_status"] = "ready"
            reused += 1
            continue
        try:
            content, suffix, content_type = download_delivery_photo_content(photo)
            rel = delivery_cache_relative_path(group, photo, version, suffix)
            target = root / rel
            target.parent.mkdir(parents=True, exist_ok=True)
            tmp = target.with_suffix(f"{target.suffix}.tmp")
            tmp.write_bytes(content)
            tmp.replace(target)
            photo["delivery_cache_path"] = str(rel).replace("\\", "/")
            photo["delivery_cache_version"] = version
            photo["delivery_cache_status"] = "ready"
            photo["delivery_cache_content_type"] = content_type
            photo["delivery_cache_built_at"] = now_iso()
            built += 1
        except Exception as exc:
            photo["delivery_cache_status"] = "failed"
            photo["delivery_cache_error"] = str(exc)
            failed.append({"photo_id": str(photo.get("id") or ""), "error": str(exc)})

    group["delivery_cache_status"] = "ready" if not failed else ("partial_failed" if built or reused else "failed")
    group["delivery_cache_built_at"] = now_iso()
    group["delivery_cache_error"] = "; ".join(item["error"] for item in failed[:3])
    append_audit_event(
        "delivery_cache_build",
        "system",
        {"group_id": group_id, "built": built, "reused": reused, "failed": len(failed)},
    )
    save_all_team_states()
    return {"status": group["delivery_cache_status"], "built": built, "reused": reused, "failed": failed, "group_id": group_id}


def schedule_delivery_cache_build(group_id: str, team_id: str | None = None, force: bool = False) -> None:
    team = normalize_team_id(team_id or current_team_id())
    group = get_group(group_id)
    if not group or not is_reviewed_group(group):
        return
    if not any(photo_can_build_delivery_cache(photo) for photo in group.get("photos", [])):
        return
    key = (team, group_id)
    with _delivery_cache_lock:
        if key in _delivery_cache_inflight:
            return
        _delivery_cache_inflight.add(key)

    def worker() -> None:
        token = set_current_team(team)
        try:
            build_delivery_cache_for_group(group_id, force=force)
        except Exception:
            group = get_group(group_id)
            if group:
                group["delivery_cache_status"] = "failed"
                group["delivery_cache_error"] = "delivery cache worker failed"
                save_all_team_states()
        finally:
            reset_current_team(token)
            with _delivery_cache_lock:
                _delivery_cache_inflight.discard(key)

    _delivery_cache_executor.submit(worker)


def copy_oss_reference(target: dict[str, Any], source: dict[str, Any]) -> None:
    ensure_photo_storage_fields(source)
    ensure_photo_identity_fields(target)
    target.setdefault("pre_oss_image_url", target.get("image_url", ""))
    target.setdefault("pre_oss_storage_type", target.get("storage_type", ""))
    target.setdefault("pre_oss_storage_key", target.get("storage_key", ""))
    target.setdefault("pre_oss_storage_bucket", target.get("storage_bucket", ""))
    target["image_url"] = source.get("image_url", "")
    target["storage_type"] = "oss"
    target["storage_bucket"] = source.get("storage_bucket", "")
    target["storage_key"] = source.get("storage_key", "")
    target["object_key"] = source.get("object_key") or source.get("storage_key", "")
    target["storage_source"] = source.get("storage_source", "reused-oss")
    target["sha256"] = source.get("sha256", target.get("sha256", ""))
    target["content_type"] = source.get("content_type", target.get("content_type", ""))
    target["byte_size"] = source.get("byte_size", target.get("byte_size", 0))
    target["oss_synced_at"] = now_iso()
    target["oss_source_type"] = "reused_oss"
    target["oss_source_ref"] = source.get("source_fingerprint", "")
    target["download_status"] = "oss_reused"


def sync_state_photos_to_oss(
    *,
    team_id: str | None = None,
    progress_callback: ProgressCallback | None = None,
    max_workers: int = 4,
) -> dict[str, Any]:
    if active_storage_backend() != "oss":
        return {
            "enabled": False,
            "status": "skipped",
            "reason": "storage backend is not oss",
            "selected_photos": 0,
            "uploaded": 0,
            "failed": 0,
        }
    state = state_for_team(team_id)
    candidates: list[tuple[dict[str, Any], dict[str, Any]]] = []
    oss_by_fingerprint: dict[str, dict[str, Any]] = {}
    oss_by_sha: dict[str, dict[str, Any]] = {}
    for group in state.get("groups", []):
        for photo in group.get("photos", []):
            ensure_photo_storage_fields(photo)
            fingerprint = str(photo.get("source_fingerprint") or "").strip()
            sha256 = str(photo.get("sha256") or "").strip()
            if str(photo.get("storage_type") or "") == "oss" or str(photo.get("image_url") or "").startswith("oss://"):
                if fingerprint:
                    oss_by_fingerprint[fingerprint] = photo
                if sha256:
                    oss_by_sha[sha256] = photo
                continue
            if photo_needs_oss_sync(photo):
                candidates.append((group, photo))
    total = len(candidates)
    report = {
        "enabled": True,
        "status": "complete",
        "selected_photos": total,
        "uploaded": 0,
        "reused_existing_oss": 0,
        "skipped_existing_oss": 0,
        "failed": 0,
        "bytes": 0,
        "failures": [],
    }
    if progress_callback:
        progress_callback(
            {
                "phase": "oss_uploading",
                "processed_photos": 0,
                "total_photos": total,
                "uploaded_photos": 0,
                "failed_photos": 0,
            }
        )
    if not candidates:
        return report

    remaining: list[tuple[dict[str, Any], dict[str, Any]]] = []
    for group, photo in candidates:
        fingerprint = str(photo.get("source_fingerprint") or "").strip()
        sha256 = str(photo.get("sha256") or "").strip()
        reusable = (oss_by_fingerprint.get(fingerprint) if fingerprint else None) or (oss_by_sha.get(sha256) if sha256 else None)
        if reusable:
            copy_oss_reference(photo, reusable)
            group["last_photo_imported_at"] = now_iso()
            report["reused_existing_oss"] += 1
            continue
        remaining.append((group, photo))
    candidates = remaining
    if not candidates:
        refresh_summary()
        save_all_team_states()
        return report

    lock = threading.Lock()

    def upload_one(item: tuple[dict[str, Any], dict[str, Any]]) -> dict[str, Any]:
        group, photo = item
        content, content_type, source_type, source_ref = load_photo_content_for_oss(photo)
        filename = photo_source_filename(photo)
        stored = save_image_bytes(
            scope="imported",
            filename=filename,
            content=content,
            content_type=content_type,
            team_id=str(state.get("team_id") or current_team_id()),
            group_id=str(group.get("id") or ""),
            key_hint=f"{group.get('id') or 'group'}-{photo.get('id') or 'photo'}",
        )
        return {
            "group": group,
            "photo": photo,
            "stored": stored,
            "source_type": source_type,
            "source_ref": source_ref,
            "byte_size": len(content),
            "content_type": content_type,
        }

    processed = 0
    with ThreadPoolExecutor(max_workers=max(1, max_workers)) as executor:
        futures = {executor.submit(upload_one, item): item for item in candidates}
        for future in as_completed(futures):
            group, photo = futures[future]
            processed += 1
            try:
                item = future.result()
                stored = item["stored"]
                with lock:
                    if str(photo.get("storage_type") or "") != "oss":
                        photo.setdefault("pre_oss_image_url", photo.get("image_url", ""))
                        photo.setdefault("pre_oss_storage_type", photo.get("storage_type", ""))
                        photo.setdefault("pre_oss_storage_key", photo.get("storage_key", ""))
                        photo.setdefault("pre_oss_storage_bucket", photo.get("storage_bucket", ""))
                    photo["image_url"] = stored["url"]
                    photo["storage_type"] = "oss"
                    photo["storage_bucket"] = stored.get("storage_bucket", "")
                    photo["storage_key"] = stored["storage_key"]
                    photo["object_key"] = stored["storage_key"]
                    photo["storage_source"] = stored["storage_source"]
                    photo["sha256"] = stored["sha256"]
                    photo["content_type"] = stored.get("content_type") or item["content_type"]
                    photo["byte_size"] = item["byte_size"]
                    photo["oss_synced_at"] = now_iso()
                    photo["oss_source_type"] = item["source_type"]
                    photo["oss_source_ref"] = item["source_ref"]
                    photo["download_status"] = "oss_migrated"
                    group["last_photo_imported_at"] = now_iso()
                    report["uploaded"] += 1
                    report["bytes"] += item["byte_size"]
                    fingerprint = str(photo.get("source_fingerprint") or "").strip()
                    sha256 = str(photo.get("sha256") or "").strip()
                    if fingerprint:
                        oss_by_fingerprint[fingerprint] = photo
                    if sha256:
                        oss_by_sha[sha256] = photo
            except Exception as exc:  # noqa: BLE001 - keep import jobs running and expose per-photo failures
                report["failed"] += 1
                photo["oss_sync_status"] = "failed"
                photo["oss_sync_error"] = str(exc)
                report["failures"].append(
                    {
                        "group_id": group.get("id"),
                        "photo_id": photo.get("id"),
                        "image_url": photo.get("image_url"),
                        "storage_key": photo.get("storage_key"),
                        "error": str(exc),
                    }
                )
            if progress_callback and (processed % 5 == 0 or processed == total):
                progress_callback(
                    {
                        "phase": "oss_uploading",
                        "processed_photos": processed,
                        "total_photos": total,
                        "uploaded_photos": report["uploaded"],
                        "failed_photos": report["failed"],
                    }
                )
            if processed % 20 == 0:
                save_all_team_states()
    refresh_summary()
    save_all_team_states()
    if report["failed"]:
        report["status"] = "partial_failed"
    return report


def scan_record_to_photo_rows(record: dict[str, Any], index: int) -> list[dict[str, Any]]:
    image_urls = record.get("image_urls") or []
    image_file_ids = record.get("image_file_ids") or []
    photo_slots = max(len(image_urls), len(image_file_ids), 1)
    rows = []
    for photo_index in range(1, photo_slots + 1):
        image_url = str(image_urls[photo_index - 1]) if photo_index <= len(image_urls) else ""
        image_file_id = str(image_file_ids[photo_index - 1]) if photo_index <= len(image_file_ids) else ""
        source_url = normalized_photo_source_url(image_url)
        source_fingerprint = source_fingerprint_from_values(
            image_file_id=image_file_id,
            image_url=image_url,
            barcode=str(record.get("barcode") or ""),
            source_file=str(record.get("source_file") or ""),
            photo_index=photo_index,
            row_number=str(record.get("file_id") or index),
        )
        rows.append(
            {
                "row_number": f"ez-{record.get('file_id') or index}-{index}-{photo_index}",
                "barcode": str(record.get("barcode") or ""),
                "meter_match_key": str(record.get("meter_match_key") or ""),
                "source_file": str(record.get("source_file") or ""),
                "collector": str(record.get("collector") or ""),
                "asset_no": str(record.get("module_asset_no") or ""),
                "asset_type": str(record.get("asset_type") or ""),
                "creator": str(record.get("creator") or record.get("installer") or ""),
                "created_at": str(record.get("created_at") or ""),
                "has_image": bool(image_url or image_file_id),
                "image_file_id": image_file_id,
                "image_url": str(image_url),
                "source_url": str(image_url),
                "source_url_hash": hash_text(source_url) if source_url else "",
                "source_file_id": image_file_id,
                "source_fingerprint": source_fingerprint,
                "import_batch_id": str(record.get("import_batch_id") or ""),
            }
        )
    return rows


def make_scan_unique_key(row: dict[str, Any]) -> str:
    fingerprint = str(row.get("source_fingerprint") or "").strip()
    if fingerprint:
        return fingerprint
    return "|".join(
        [
            str(row.get("barcode") or ""),
            str(row.get("source_file") or ""),
            str(row.get("image_file_id") or ""),
            str(row.get("image_url") or ""),
        ]
    )


def make_photo_unique_key(photo: dict[str, Any]) -> str:
    ensure_photo_identity_fields(photo)
    fingerprint = str(photo.get("source_fingerprint") or "").strip()
    if fingerprint:
        return fingerprint
    return "|".join(
        [
            str(photo.get("barcode") or ""),
            str(photo.get("source_file") or ""),
            str(photo.get("image_file_id") or ""),
            str(photo.get("image_url") or ""),
        ]
    )


def make_construction_photo_unique_key(photo: dict[str, Any]) -> tuple[str, str, str]:
    return (
        str(photo.get("client_batch_id") or ""),
        str(photo.get("client_photo_id") or ""),
        str(photo.get("sha256") or ""),
    )


def make_import_meter_duplicate_key(record: dict[str, Any]) -> str:
    meter_no = str(record.get("meter_no") or "").strip()
    if meter_no:
        return build_total_catalog_match_key(meter_no) or meter_no
    match_key = str(record.get("meter_match_key") or "").strip()
    if match_key:
        return match_key
    barcode = str(record.get("barcode") or "").strip()
    if barcode:
        try:
            return build_long_scan_match_key(barcode)
        except ValueError:
            return barcode
    return ""


def make_import_meter_duplicate_key_from_group(group: dict[str, Any]) -> str:
    meter_no = str(group.get("meter_no") or "").strip()
    if meter_no:
        return build_total_catalog_match_key(meter_no) or meter_no
    return str(group.get("meter_match_key") or "").strip()


def make_unmatched_id(record: dict[str, Any]) -> str:
    payload = {
        "barcode": record.get("barcode") or record.get("meter_no") or "",
        "source_file": record.get("source_file") or "",
        "terminal": record.get("terminal") or "",
        "meter_match_key": record.get("meter_match_key") or "",
        "image_urls": record.get("image_urls") or [],
        "photo_urls": record.get("photo_urls") or "",
    }
    raw = json.dumps(payload, ensure_ascii=False, sort_keys=True)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()[:16]


def ensure_unmatched_record(record: dict[str, Any]) -> dict[str, Any]:
    item = dict(record)
    item["unmatched_id"] = str(item.get("unmatched_id") or make_unmatched_id(item))
    item["record_type"] = str(item.get("record_type") or "scan")
    item["created_at"] = str(item.get("created_at") or now_iso())
    item["text_index"] = build_record_text_index(item)
    return item


def merge_unmatched_records(existing: list[dict[str, Any]], incoming: list[dict[str, Any]]) -> list[dict[str, Any]]:
    merged: dict[str, dict[str, Any]] = {}
    for record in existing + incoming:
        item = ensure_unmatched_record(record)
        merged[item["unmatched_id"]] = {**merged.get(item["unmatched_id"], {}), **item}
    return list(merged.values())


def build_record_text_index(record: dict[str, Any]) -> str:
    values: list[str] = []
    for value in record.values():
        if isinstance(value, list):
            values.extend(str(item) for item in value)
        elif not isinstance(value, dict):
            values.append(str(value))
    return " ".join(values).lower()


def list_unmatched_records(query: str = "", limit: int = 100, offset: int = 0) -> dict[str, Any]:
    state = get_state()
    records = [ensure_unmatched_record(item) for item in state.get("scan_unmatched", [])]
    q = query.strip().lower()
    if q:
        terms = [item for item in re.split(r"\s+", q) if item]
        records = [item for item in records if all(term in item.get("text_index", "") for term in terms)]
    records = sorted(records, key=lambda item: (str(item.get("terminal") or ""), str(item.get("barcode") or "")))
    return {"total": len(records), "items": records[offset : offset + limit]}


def get_unmatched_record(unmatched_id: str) -> dict[str, Any] | None:
    state = get_state()
    for record in state.get("scan_unmatched", []):
        item = ensure_unmatched_record(record)
        if item["unmatched_id"] == unmatched_id:
            return item
    return None


def delete_unmatched_record(unmatched_id: str, actor: str, reason: str = "") -> dict[str, Any]:
    state = get_state()
    kept = []
    deleted = None
    for record in state.get("scan_unmatched", []):
        item = ensure_unmatched_record(record)
        if item["unmatched_id"] == unmatched_id:
            deleted = item
        else:
            kept.append(item)
    if deleted is None:
        raise KeyError(unmatched_id)
    state["scan_unmatched"] = kept
    append_audit_event("delete_unmatched", actor, {"unmatched_id": unmatched_id, "reason": reason, "record": deleted})
    refresh_summary()
    return deleted


def create_blank_unmatched_record(actor: str) -> dict[str, Any]:
    state = get_state()
    record = ensure_unmatched_record(
        {
            "record_type": "blank_group",
            "unmatched_id": f"manual-blank-{now_iso()}",
            "barcode": "",
            "meter_no": "",
            "meter_match_key": "",
            "terminal": "",
            "address": "",
            "collector": "",
            "module_asset_no": "",
            "asset_no": "",
            "creator": "",
            "photo_urls": "",
            "image_urls": [],
            "source_file": "manual-blank-group",
        }
    )
    state["scan_unmatched"] = merge_unmatched_records(state.get("scan_unmatched", []), [record])
    append_audit_event("create_blank_unmatched", actor, {"unmatched_id": record["unmatched_id"]})
    refresh_summary()
    return record


UNMATCHED_EDIT_FIELDS = {
    "barcode",
    "meter_no",
    "meter_match_key",
    "terminal",
    "address",
    "collector",
    "module_asset_no",
    "asset_no",
    "creator",
    "note",
    "assignment_note",
    "replacement_old_meter_no",
    "replacement_by",
    "replacement_at",
}


def update_unmatched_record(
    unmatched_id: str,
    actor: str,
    updates: dict[str, Any] | None = None,
) -> dict[str, Any]:
    state = get_state()
    updates = updates or {}
    for index, record in enumerate(state.get("scan_unmatched", [])):
        item = ensure_unmatched_record(record)
        if item["unmatched_id"] != unmatched_id:
            continue
        for key in UNMATCHED_EDIT_FIELDS:
            if key in updates:
                item[key] = str(updates.get(key) or "").strip()
        if item.get("module_asset_no") and not item.get("asset_no"):
            item["asset_no"] = item["module_asset_no"]
        item["updated_by"] = actor
        item["updated_at"] = now_iso()
        item["text_index"] = build_record_text_index(item)
        state["scan_unmatched"][index] = item
        append_audit_event("update_unmatched", actor, {"unmatched_id": unmatched_id, "updates": updates})
        refresh_summary()
        return item
    raise KeyError(unmatched_id)


def assign_unmatched_record(
    unmatched_id: str,
    actor: str,
    constructor: str,
    note: str = "",
    due_date: str = "",
) -> dict[str, Any]:
    constructor = constructor.strip()
    if not constructor:
        raise ValueError("Constructor is required")
    record = update_unmatched_record(
        unmatched_id,
        actor,
        {
            "assignment_note": note,
        },
    )
    terminal = str(record.get("terminal") or "").strip()
    if terminal:
        task = ensure_task_for_terminal(terminal)
        assign_construction_task(int(task["id"]), actor=actor, constructor=constructor, note=note, due_date=due_date)
    record.update(
        {
            "assigned_to": constructor,
            "assigned_by": actor,
            "assigned_at": now_iso(),
            "assignment_note": note.strip(),
            "due_date": due_date.strip(),
            "field_task_type": "unmatched",
        }
    )
    record["text_index"] = build_record_text_index(record)
    state = get_state()
    state["scan_unmatched"] = [
        record if ensure_unmatched_record(item)["unmatched_id"] == unmatched_id else item
        for item in state.get("scan_unmatched", [])
    ]
    append_audit_event(
        "assign_unmatched",
        actor,
        {"unmatched_id": unmatched_id, "constructor": constructor, "note": note, "due_date": due_date},
    )
    return record


def unassign_unmatched_record(unmatched_id: str, actor: str, reason: str = "") -> dict[str, Any]:
    state = get_state()
    for index, record in enumerate(state.get("scan_unmatched", [])):
        item = ensure_unmatched_record(record)
        if item["unmatched_id"] != unmatched_id:
            continue
        previous = item.get("assigned_to") or ""
        item["assigned_to"] = ""
        item["unassigned_by"] = actor
        item["unassigned_at"] = now_iso()
        item["unassign_reason"] = reason.strip()
        item["text_index"] = build_record_text_index(item)
        state["scan_unmatched"][index] = item
        append_audit_event(
            "unassign_unmatched",
            actor,
            {"unmatched_id": unmatched_id, "previous_constructor": previous, "reason": reason},
        )
        return item
    raise KeyError(unmatched_id)


def mark_unmatched_outside_project(unmatched_id: str, actor: str, note: str = "") -> dict[str, Any]:
    state = get_state()
    for index, record in enumerate(state.get("scan_unmatched", [])):
        item = ensure_unmatched_record(record)
        if item["unmatched_id"] != unmatched_id:
            continue
        item["project_outside"] = True
        item["project_outside_by"] = actor
        item["project_outside_at"] = now_iso()
        item["project_outside_note"] = note.strip()
        item["field_task_type"] = "outside_project"
        item["text_index"] = build_record_text_index(item)
        state["scan_unmatched"][index] = item
        append_audit_event("mark_unmatched_outside_project", actor, {"unmatched_id": unmatched_id, "note": note})
        refresh_summary()
        return item
    raise KeyError(unmatched_id)


def find_group_by_meter_reference(meter_reference: str, terminal: str = "") -> dict[str, Any] | None:
    reference = str(meter_reference or "").strip()
    if not reference:
        return None
    candidates = {
        reference,
        build_total_catalog_match_key(reference) or reference,
    }
    terminal = terminal.strip()
    for group in get_state()["groups"]:
        if terminal and str(group.get("terminal") or "") != terminal:
            continue
        values = {
            str(group.get("id") or ""),
            str(group.get("meter_no") or ""),
            str(group.get("meter_match_key") or ""),
            build_total_catalog_match_key(str(group.get("meter_no") or "")) or "",
        }
        if candidates & values:
            return group
    return None


def rematch_unmatched_record(
    unmatched_id: str,
    actor: str,
    meter_no: str = "",
    old_meter_no: str = "",
    terminal: str = "",
    updates: dict[str, Any] | None = None,
) -> dict[str, Any]:
    merged_updates = dict(updates or {})
    if meter_no:
        merged_updates["meter_no"] = meter_no
        merged_updates["barcode"] = meter_no
        merged_updates["meter_match_key"] = build_total_catalog_match_key(meter_no) or meter_no
    if terminal:
        merged_updates["terminal"] = terminal
    if old_meter_no:
        merged_updates["replacement_old_meter_no"] = old_meter_no
        merged_updates["replacement_by"] = actor
        merged_updates["replacement_at"] = now_iso()
    record = update_unmatched_record(unmatched_id, actor, merged_updates)
    target = find_group_by_meter_reference(old_meter_no or meter_no or record.get("meter_no") or record.get("barcode"), terminal or record.get("terminal", ""))
    if target is None and not terminal:
        target = find_group_by_meter_reference(old_meter_no or meter_no or record.get("meter_no") or record.get("barcode"), "")
    if target is None:
        append_audit_event(
            "rematch_unmatched_no_target",
            actor,
            {"unmatched_id": unmatched_id, "meter_no": meter_no, "old_meter_no": old_meter_no, "terminal": terminal},
        )
        return {"record": record, "matched": False}
    associate_updates = dict(record)
    if old_meter_no:
        associate_updates["replacement_old_meter_no"] = old_meter_no
        associate_updates["replacement_target_group_id"] = target.get("id", "")
    associated = associate_unmatched_record(
        unmatched_id,
        actor,
        target_group_id=str(target.get("id") or ""),
        updates=associate_updates,
    )
    associated["matched"] = True
    associated["replacement_old_meter_no"] = old_meter_no
    return associated


def associate_unmatched_record(
    unmatched_id: str,
    actor: str,
    target_group_id: str = "",
    target_meter_no: str = "",
    updates: dict[str, Any] | None = None,
) -> dict[str, Any]:
    state = get_state()
    record = get_unmatched_record(unmatched_id)
    if record is None:
        raise KeyError(unmatched_id)
    group = None
    if target_group_id:
        group = get_group(target_group_id)
    if group is None and target_meter_no:
        group = next((item for item in state["groups"] if item["meter_no"] == target_meter_no), None)
    if group is None:
        raise ValueError("Target data group was not found")

    payload = {**record, **(updates or {})}
    payload["meter_match_key"] = group["meter_match_key"]
    payload["meter_no"] = group["meter_no"]
    payload["terminal"] = group["terminal"]
    result = apply_synced_scan_records([payload])
    delete_unmatched_record(unmatched_id, actor, "associated to data group")
    append_audit_event(
        "associate_unmatched",
        actor,
        {"unmatched_id": unmatched_id, "group_id": group["id"], "meter_no": group["meter_no"], "applied": result},
    )
    return {"group": group, "import_result": result}


def ensure_construction_task_fields(task: dict[str, Any]) -> dict[str, Any]:
    task.setdefault("construction_enabled", False)
    task.setdefault("construction_claimed_by", None)
    task.setdefault("construction_claimed_at", None)
    task.setdefault("construction_released_at", None)
    task.setdefault("construction_opened_by", None)
    task.setdefault("construction_opened_at", None)
    task.setdefault("construction_closed_at", None)
    task["assigned_constructor"] = task.get("construction_claimed_by")
    task["assigned_at"] = task.get("construction_claimed_at")
    if task.get("construction_enabled"):
        task["construction_status"] = "assigned" if task.get("construction_claimed_by") else "open"
    else:
        task["construction_status"] = "closed"
    return task


def construction_defaults() -> dict[str, Any]:
    return ensure_construction_task_fields({}).copy()


CONSTRUCTION_TASK_FIELDS = tuple(construction_defaults().keys())


def carry_construction_task_fields(tasks: list[dict[str, Any]], existing_tasks: list[dict[str, Any]]) -> list[dict[str, Any]]:
    existing_by_terminal = {str(task.get("terminal") or ""): task for task in existing_tasks}
    for task in tasks:
        existing = existing_by_terminal.get(str(task.get("terminal") or ""))
        if existing:
            for field in CONSTRUCTION_TASK_FIELDS:
                if field in existing:
                    task[field] = existing[field]
        ensure_construction_task_fields(task)
    return tasks


def ensure_task_for_terminal(terminal: str) -> dict[str, Any]:
    state = get_state()
    task = next((item for item in state["tasks"] if str(item.get("terminal") or "") == terminal), None)
    if task:
        return ensure_construction_task_fields(task)
    task_id = max((int(item.get("id", 0)) for item in state["tasks"]), default=0) + 1
    task = {
        "id": task_id,
        "project_id": 1,
        "terminal": terminal,
        "name": f"终端 {terminal}",
        "status": "published",
        "claimed_by": None,
        "claimed_at": None,
        "released_at": None,
        "total_groups": 0,
        "completed_groups": 0,
        "exception_groups": 0,
        "incomplete_groups": 0,
        "pending_groups": 0,
        "scan_rows": 0,
        "groups_with_scan": 0,
        "renovation_count": 0,
        "uploaded_count": 0,
        "reviewed_count": 0,
        "unreviewed_count": 0,
        "upload_rate": 0.0,
        "review_rate": 0.0,
        "complete_groups": 0,
        "partial_groups": 0,
        "has_scan_info": False,
        "can_claim": False,
        "claim_block_reason": "该终端暂无扫码信息，不能领取",
        "progress": 0.0,
        "completeness_rate": 0.0,
        **construction_defaults(),
    }
    state["tasks"].append(task)
    return task


def next_group_id() -> str:
    max_id = 0
    for group in get_state()["groups"]:
        raw = str(group.get("id", ""))
        if raw.startswith("g-") and raw[2:].isdigit():
            max_id = max(max_id, int(raw[2:]))
    return f"g-{max_id + 1:05d}"


def normalize_unmatched_photo_payload(record: dict[str, Any], updates: dict[str, Any]) -> dict[str, Any]:
    payload = {**record, **updates}
    if not payload.get("image_urls"):
        payload["image_urls"] = expand_photo_urls(split_urls(str(payload.get("photo_urls") or "")))
    payload.setdefault("image_file_ids", [])
    return payload


def find_existing_group_for_unmatched(terminal: str, meter_match_key: str, meter_no: str) -> dict[str, Any] | None:
    terminal = terminal.strip()
    meter_match_key = meter_match_key.strip()
    meter_no = meter_no.strip()
    if not terminal or not (meter_match_key or meter_no):
        return None
    for group in get_state()["groups"]:
        if str(group.get("terminal") or "") != terminal:
            continue
        if meter_match_key and str(group.get("meter_match_key") or "") == meter_match_key:
            return group
        if meter_no and str(group.get("meter_no") or "") == meter_no:
            return group
    return None


def append_unmatched_photos_to_group(group: dict[str, Any], photo_rows: list[dict[str, Any]]) -> int:
    existing = {make_photo_unique_key(photo) for photo in group.get("photos", [])}
    added = 0
    for row in photo_rows:
        photo = build_photo_record(len(group["photos"]) + 1, row)
        key = make_photo_unique_key(photo)
        if key in existing:
            continue
        group["photos"].append(photo)
        existing.add(key)
        added += 1
    group["photo_count"] = len(group["photos"])
    if group["status"] in DONE_STATUSES or group["photo_count"] < 4:
        group["status"] = "incomplete"
    elif group["status"] == "incomplete":
        group["status"] = "pending"
    group["reviewer"] = None
    group["reviewed_at"] = None
    return added


def create_group_from_unmatched_record(
    unmatched_id: str,
    actor: str,
    terminal: str,
    updates: dict[str, Any] | None = None,
) -> dict[str, Any]:
    terminal = terminal.strip()
    if not terminal:
        raise ValueError("Target terminal is required")
    state = get_state()
    record = get_unmatched_record(unmatched_id)
    if record is None:
        raise KeyError(unmatched_id)

    payload = normalize_unmatched_photo_payload(record, updates or {})
    meter_no = str(payload.get("meter_no") or payload.get("barcode") or payload.get("meter_match_key") or unmatched_id)
    meter_match_key = str(payload.get("meter_match_key") or build_total_catalog_match_key(meter_no) or meter_no)
    task = ensure_task_for_terminal(terminal)
    photo_rows = scan_record_to_photo_rows({**payload, "meter_match_key": meter_match_key}, len(state["groups"]) + 1)
    existing_group = find_existing_group_for_unmatched(terminal, meter_match_key, meter_no)
    if existing_group is not None:
        added = append_unmatched_photos_to_group(existing_group, photo_rows)
        existing_group["terminal"] = terminal
        existing_group["stage_terminal"] = terminal
        existing_group["task_id"] = task["id"]
        if payload.get("address"):
            existing_group["address"] = str(payload.get("address") or "")
        delete_unmatched_record(unmatched_id, actor, "attached to existing terminal data group")
        append_audit_event(
            "attach_unmatched_to_existing_group",
            actor,
            {
                "unmatched_id": unmatched_id,
                "terminal": terminal,
                "group_id": existing_group["id"],
                "meter_no": existing_group["meter_no"],
                "added_photos": added,
            },
        )
        refresh_summary()
        return {"group": existing_group, "task": task, "attached": True, "added_photos": added}

    photos = [build_photo_record(index, row) for index, row in enumerate(photo_rows, start=1)]
    group = {
        "id": next_group_id(),
        "task_id": task["id"],
        "meter_match_key": meter_match_key,
        "meter_no": meter_no,
        "terminal": terminal,
        "address": str(payload.get("address") or ""),
        "stage_meter_no": str(payload.get("stage_meter_no") or ""),
        "stage_terminal": terminal,
        "status": "pending" if len(photos) >= 4 else "incomplete",
        "reviewer": None,
        "review_note": "",
        "exception_note": "",
        "reviewed_at": None,
        "photo_count": len(photos),
        "photos": photos[:8],
        "manual_created": True,
        "source_unmatched_id": unmatched_id,
    }
    state["groups"].append(group)
    delete_unmatched_record(unmatched_id, actor, "created data group for terminal")
    append_audit_event(
        "create_group_from_unmatched",
        actor,
        {"unmatched_id": unmatched_id, "terminal": terminal, "group_id": group["id"], "meter_no": group["meter_no"]},
    )
    refresh_summary()
    return {"group": group, "task": task, "attached": False, "added_photos": len(photos)}


def create_empty_group_for_terminal(
    terminal: str,
    actor: str,
    meter_no: str = "",
    address: str = "",
    meter_match_key: str = "",
) -> dict[str, Any]:
    terminal = terminal.strip()
    state = get_state()
    task = ensure_task_for_terminal(terminal or "未关联终端")
    meter_no = meter_no.strip() or f"manual-{next_group_id()}"
    meter_match_key = meter_match_key.strip() or build_total_catalog_match_key(meter_no) or meter_no
    group = {
        "id": next_group_id(),
        "task_id": task["id"],
        "meter_match_key": meter_match_key,
        "meter_no": meter_no,
        "terminal": terminal,
        "address": address.strip(),
        "stage_meter_no": "",
        "stage_terminal": terminal,
        "status": "incomplete",
        "reviewer": None,
        "review_note": "",
        "exception_note": "",
        "reviewed_at": None,
        "photo_count": 0,
        "photos": [],
        "manual_created": True,
    }
    state["groups"].append(group)
    append_audit_event("create_empty_group", actor, {"terminal": terminal, "group_id": group["id"], "meter_no": meter_no})
    refresh_summary()
    return {"group": group, "task": task}


def update_group_terminal(group_id: str, terminal: str, actor: str) -> dict[str, Any]:
    group = get_group(group_id)
    if group is None:
        raise KeyError(group_id)
    terminal = terminal.strip()
    if not terminal:
        raise ValueError("Target terminal is required")
    previous_terminal = str(group.get("terminal") or "")
    task = ensure_task_for_terminal(terminal)
    group["terminal"] = terminal
    group["stage_terminal"] = terminal
    group["task_id"] = task["id"]
    append_audit_event(
        "update_group_terminal",
        actor,
        {"group_id": group_id, "previous_terminal": previous_terminal, "terminal": terminal, "task_id": task["id"]},
    )
    refresh_summary()
    return {"group": group, "task": task}


def update_group_metadata(group_id: str, actor: str, updates: dict[str, Any]) -> dict[str, Any]:
    group = get_group(group_id)
    if group is None:
        raise KeyError(group_id)
    allowed_group_fields = {"meter_no", "address", "meter_match_key"}
    previous = {field: group.get(field) for field in allowed_group_fields}
    previous_reasons = list(group.get("exception_reasons") or [])
    previous_auto_note = "; ".join(str(item) for item in previous_reasons if str(item).strip())
    for field in allowed_group_fields:
        if field in updates:
            group[field] = str(updates.get(field) or "").strip()
    photo_field_map = {
        "collector": "collector",
        "module_asset_no": "asset_no",
        "creator": "creator",
    }
    for incoming, photo_field in photo_field_map.items():
        if incoming not in updates:
            continue
        value = str(updates.get(incoming) or "").strip()
        for photo in group.get("photos", []):
            photo[photo_field] = value
    reasons = validate_group_archive(group)
    set_group_exception_flags(group, reasons)
    if previous_auto_note and str(group.get("exception_note") or "").strip() == previous_auto_note:
        group["exception_note"] = ""
    if group.get("status") == "exception" and not reasons and not group.get("exception_note"):
        if is_group_fully_archived(group):
            group["status"] = "approved"
            group["reviewer"] = actor
            group["review_note"] = "\u8d44\u6599\u7ec4\u4fee\u6b63\u5b8c\u6210"
            group["reviewed_at"] = now_iso()
        else:
            group["status"] = "incomplete" if group.get("photo_count", 0) < 4 else "pending"
            group["review_note"] = ""
            group["reviewed_at"] = None
    append_audit_event(
        "update_group_metadata",
        actor,
        {"group_id": group_id, "previous": previous, "updates": {key: updates.get(key) for key in sorted(updates)}},
    )
    refresh_summary()
    return {"group": group}


def add_photo_urls_to_group(
    group_id: str,
    actor: str,
    photo_urls: list[str],
    collector: str = "",
    module_asset_no: str = "",
    creator: str = "",
    photo_metadata: dict[str, dict[str, Any]] | None = None,
) -> dict[str, Any]:
    group = get_group(group_id)
    if group is None:
        raise KeyError(group_id)
    expanded_urls = expand_photo_urls(photo_urls)
    existing = {make_photo_unique_key(photo) for photo in group.get("photos", []) + group.get("deleted_photos", [])}
    added = 0
    skipped_duplicates = 0
    for url in expanded_urls:
        metadata = (photo_metadata or {}).get(url, {})
        source_url = normalized_photo_source_url(url)
        row = {
            "row_number": f"manual-{group_id}-{len(group['photos']) + 1}",
            "barcode": group.get("meter_no", ""),
            "meter_match_key": group.get("meter_match_key", ""),
            "source_file": "manual-photo-import",
            "collector": collector,
            "asset_no": module_asset_no,
            "asset_type": "",
            "creator": creator,
            "created_at": now_iso(),
            "has_image": True,
            "image_file_id": "",
            "image_url": url,
            "source_url": url,
            "source_url_hash": hash_text(source_url) if source_url else "",
            "source_fingerprint": source_fingerprint_from_values(
                image_url=url,
                barcode=group.get("meter_no", ""),
                source_file="manual-photo-import",
                photo_index=len(group["photos"]) + 1,
                row_number=group_id,
            ),
            "storage_type": metadata.get("storage_type", ""),
            "storage_key": metadata.get("storage_key", ""),
            "storage_bucket": metadata.get("storage_bucket", ""),
            "storage_source": metadata.get("storage_source", ""),
            "sha256": metadata.get("sha256", ""),
        }
        photo = build_photo_record(len(group["photos"]) + 1, row)
        key = make_photo_unique_key(photo)
        if key in existing:
            skipped_duplicates += 1
            continue
        group["photos"].append(photo)
        existing.add(key)
        added += 1
    group["photo_count"] = len(group["photos"])
    if group["status"] in DONE_STATUSES or group["photo_count"] < 4:
        group["status"] = "incomplete"
    elif group["status"] == "incomplete":
        group["status"] = "pending"
    group["reviewer"] = None
    group["reviewed_at"] = None
    mark_delivery_cache_stale(group, "manual photos changed")
    append_audit_event("add_group_photos", actor, {"group_id": group_id, "added": added, "skipped_duplicates": skipped_duplicates})
    refresh_summary()
    return {"group": group, "added": added, "skipped_duplicates": skipped_duplicates}


def append_audit_event(action: str, actor: str, payload: dict[str, Any]) -> dict[str, Any]:
    state = state_for_team()
    event = {
        "id": f"audit-{len(state.get('audit_events', [])) + 1:06d}",
        "action": action,
        "actor": actor,
        "payload": payload,
        "created_at": now_iso(),
    }
    state.setdefault("audit_events", []).append(event)
    return event


def list_audit_events(limit: int = 100, offset: int = 0) -> dict[str, Any]:
    events = list(reversed(get_state().get("audit_events", [])))
    return {"total": len(events), "items": events[offset : offset + limit]}


def apply_group_photo_urls(group_id: str, urls: dict[str, str]) -> dict[str, Any]:
    group = get_group(group_id)
    if group is None:
        raise KeyError(group_id)
    applied = 0
    for photo in group.get("photos", []):
        file_id = str(photo.get("image_file_id") or "")
        if file_id and not photo.get("image_url") and file_id in urls:
            photo["image_url"] = urls[file_id]
            photo["source_url"] = urls[file_id]
            photo["source_url_hash"] = hash_text(normalized_photo_source_url(urls[file_id]))
            ensure_photo_identity_fields(photo)
            photo["download_status"] = "downloaded"
            photo["downloaded_at"] = now_iso()
            applied += 1
    refresh_summary()
    return {"group_id": group_id, "loaded_photo_urls": applied, "group": group}


def bootstrap_local_simulation(paths: LocalTestPaths | None = None) -> dict[str, Any]:
    paths = paths or LocalTestPaths()
    total_rows = read_catalog_rows(paths.total_catalog, source="total")
    scan_rows = read_scan_rows(paths.scan_file)
    return rebuild_state_from_total_catalog(total_rows, scan_rows=scan_rows, paths=paths)


def rebuild_state_from_total_catalog(
    total_rows: list[dict[str, Any]],
    scan_rows: list[dict[str, Any]] | None = None,
    paths: LocalTestPaths | None = None,
    preserve_existing: bool = False,
) -> dict[str, Any]:
    state = state_for_team()
    team_id = current_team_id()
    scan_rows = scan_rows or []
    total_by_key: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in total_rows:
        total_by_key[row["meter_match_key"]].append(row)

    scans_by_key: dict[str, list[dict[str, Any]]] = defaultdict(list)
    scan_unmatched = []
    for row in scan_rows:
        if row["meter_match_key"] in total_by_key:
            scans_by_key[row["meter_match_key"]].append(row)
        else:
            scan_unmatched.append(ensure_unmatched_record(row))

    existing_groups_by_key = {
        str(group.get("meter_match_key") or ""): group
        for group in state.get("groups", [])
        if preserve_existing and group.get("meter_match_key")
    }
    terminal_task_ids = build_terminal_task_ids(total_rows)
    groups = []
    for index, total in enumerate(total_rows, start=1):
        existing_group = existing_groups_by_key.get(str(total["meter_match_key"]))
        if existing_group:
            photos = existing_group.get("photos", [])
            status = existing_group.get("status", "pending")
            reviewer = existing_group.get("reviewer")
            review_note = existing_group.get("review_note", "")
            exception_note = existing_group.get("exception_note", "")
            reviewed_at = existing_group.get("reviewed_at")
        else:
            photos = [
                build_photo_record(photo_index, row)
                for photo_index, row in enumerate(scans_by_key.get(total["meter_match_key"], []), start=1)
            ]
            status = "incomplete" if len(photos) < 4 else "pending"
            reviewer = None
            review_note = ""
            exception_note = ""
            reviewed_at = None
        terminal = total["terminal"]
        groups.append(
            {
                "id": f"g-{index:05d}",
                "task_id": terminal_task_ids.get(terminal, 0),
                "meter_match_key": total["meter_match_key"],
                "meter_no": total["meter_no"],
                "terminal": terminal,
                "address": total["address"],
                "stage_meter_no": "",
                "stage_terminal": "",
                "status": status,
                "reviewer": reviewer,
                "review_note": review_note,
                "exception_note": exception_note,
                "reviewed_at": reviewed_at,
                "photo_count": len(photos),
                "photos": photos[:8],
            }
        )

    summary = build_summary(total_rows, [], scan_rows, groups, [], scan_unmatched)
    tasks = build_terminal_tasks(groups)
    if preserve_existing:
        tasks = carry_construction_task_fields(tasks, state.get("tasks", []))

    summary["team_id"] = team_id
    state.update(
        {
            "team_id": team_id,
            "loaded": True,
            "paths": {
                "total_catalog": str(paths.total_catalog) if paths else "",
                "stage_catalog": "",
                "scan_file": str(paths.scan_file) if paths else "",
            },
            "summary": summary,
            "projects": [
                {
                    "id": 1,
                    "name": "V2.1 Local Simulation",
                    "team_id": team_id,
                    "status": "active",
                    "summary": summary,
                }
            ],
            "tasks": tasks,
            "groups": groups,
            "total_catalog": total_rows,
            "stage_catalog": [],
            "scan_unmatched": scan_unmatched,
            "stage_unmatched": [],
            "review_events": state.get("review_events", []) if preserve_existing else [],
            "photo_events": state.get("photo_events", []) if preserve_existing else [],
            "audit_events": state.get("audit_events", []) if preserve_existing else [],
        }
    )
    refresh_group_exceptions()
    return state


def build_terminal_task_ids(stage_rows: list[dict[str, Any]]) -> dict[str, int]:
    terminals = sorted({row["terminal"] for row in stage_rows if row["terminal"]})
    return {terminal: index for index, terminal in enumerate(terminals, start=1)}


def build_photo_record(photo_index: int, row: dict[str, Any]) -> dict[str, Any]:
    has_image = bool(row.get("has_image"))
    storage = infer_photo_storage(str(row.get("image_url", "")))
    photo = {
        "id": f"p-{row['row_number']}-{photo_index}",
        "row_number": row["row_number"],
        "barcode": row["barcode"],
        "meter_match_key": row["meter_match_key"],
        "source_file": row.get("source_file", ""),
        "collector": row.get("collector", ""),
        "asset_no": row.get("asset_no", ""),
        "asset_type": row.get("asset_type", ""),
        "creator": row.get("creator", ""),
        "created_at": row.get("created_at", ""),
        "has_image": has_image,
        "download_status": "downloaded" if has_image else "missing",
        "downloaded_at": now_iso() if has_image else None,
        "image_file_id": row.get("image_file_id", ""),
        "image_url": row.get("image_url", ""),
        "source_url": row.get("source_url", row.get("image_url", "")),
        "source_url_hash": row.get("source_url_hash", ""),
        "source_file_id": row.get("source_file_id", row.get("image_file_id", "")),
        "source_fingerprint": row.get("source_fingerprint", ""),
        "import_batch_id": row.get("import_batch_id", ""),
        "is_active": True,
        "deleted_at": None,
        "deleted_by": "",
        "delete_reason": "",
        "storage_type": row.get("storage_type") or storage["storage_type"],
        "storage_key": row.get("storage_key") or storage["storage_key"],
        "storage_bucket": row.get("storage_bucket", ""),
        "storage_source": row.get("storage_source") or row.get("source_file", "") or storage["storage_type"],
        "sha256": row.get("sha256", ""),
        "category": "unclassified",
        "category_label": PHOTO_CATEGORIES["unclassified"],
        "archive_status": "pending",
        "archive_filename": "",
        "archived_at": None,
    }
    return ensure_photo_storage_fields(photo)


def build_terminal_tasks(groups: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_terminal: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for group in groups:
        by_terminal[group["terminal"] or "UNKNOWN"].append(group)

    tasks = []
    for task_id, terminal in enumerate(sorted(by_terminal), start=1):
        terminal_groups = by_terminal[terminal]
        metrics = calculate_task_metrics(terminal_groups)
        scan_rows = sum(group["photo_count"] for group in terminal_groups)
        has_scan_info = scan_rows > 0
        address = first_task_address(terminal_groups)
        address_search_text = task_address_search_text(terminal_groups)
        task = {
            "id": task_id,
            "project_id": 1,
            "terminal": terminal,
            "address": address,
            "address_search_text": address_search_text,
            "name": f"终端 {terminal}",
            "status": "published",
            "claimed_by": None,
            "claimed_at": None,
            "released_at": None,
            "total_groups": len(terminal_groups),
            "completed_groups": sum(1 for group in terminal_groups if is_reviewed_group(group)),
            "exception_groups": sum(1 for group in terminal_groups if is_problem_group(group)),
            "incomplete_groups": count_incomplete_scanned_groups(terminal_groups),
            "unconstructed_groups": count_unconstructed_groups(terminal_groups),
            "pending_groups": sum(1 for group in terminal_groups if is_unreviewed_group(group)),
            "scan_rows": scan_rows,
            "groups_with_scan": sum(1 for group in terminal_groups if group["photo_count"] > 0),
            **metrics,
            "complete_groups": count_complete_groups(terminal_groups),
            "partial_groups": count_partial_groups(terminal_groups),
            "has_scan_info": has_scan_info,
            "can_claim": has_scan_info,
            "claim_block_reason": "" if has_scan_info else "该终端暂无扫码信息，不能领取",
            "progress": calculate_progress(terminal_groups),
            "completeness_rate": metrics["upload_rate"],
            **construction_defaults(),
        }
        tasks.append(task)
    return tasks


def first_task_address(groups: list[dict[str, Any]]) -> str:
    for group in groups:
        address = str(group.get("address") or "").strip()
        if address:
            return address
    return ""


def task_address_search_text(groups: list[dict[str, Any]]) -> str:
    addresses: list[str] = []
    seen: set[str] = set()
    for group in groups:
        address = str(group.get("address") or "").strip()
        if not address or address in seen:
            continue
        seen.add(address)
        addresses.append(address)
    return " ".join(addresses)


def build_summary(
    total_rows: list[dict[str, Any]],
    stage_rows: list[dict[str, Any]],
    scan_rows: list[dict[str, Any]],
    groups: list[dict[str, Any]],
    stage_unmatched: list[dict[str, Any]],
    scan_unmatched: list[dict[str, Any]],
) -> dict[str, Any]:
    photo_rows_linked = sum(item["photo_count"] for item in groups)
    scanned_groups = sum(1 for item in groups if item.get("photo_count", 0) > 0)
    reviewed_groups = sum(1 for item in groups if is_reviewed_group(item))
    exception_groups = sum(1 for item in groups if is_problem_group(item))
    return {
        "total_catalog_rows": len(total_rows),
        "stage_catalog_rows": len(stage_rows),
        "scan_rows": len(scan_rows),
        "groups": len(groups),
        "matched_groups": sum(1 for item in groups if item["status"] != "unmatched"),
        "incomplete_groups": count_incomplete_scanned_groups(groups),
        "unconstructed_groups": count_unconstructed_groups(groups),
        "approved_groups": sum(1 for item in groups if item["status"] == "approved"),
        "exception_groups": exception_groups,
        "reviewed_groups": reviewed_groups,
        "unreviewed_groups": sum(1 for item in groups if is_unreviewed_group(item)),
        "stage_unmatched": len(stage_unmatched),
        "scan_unmatched": len(scan_unmatched),
        "photo_rows_linked": photo_rows_linked,
        "scanned_groups": scanned_groups,
        "installer_distribution": summarize_installers_by_group(groups),
        "downloaded_photos": sum(
            1 for group in groups for photo in group.get("photos", []) if photo.get("download_status") == "downloaded"
        ),
        "unclassified_photos": sum(
            1 for group in groups for photo in group.get("photos", []) if photo.get("category") == "unclassified"
        ),
        "review_progress": calculate_progress(groups),
    }


def summarize_installers_by_group(groups: list[dict[str, Any]]) -> list[dict[str, Any]]:
    counts: dict[str, int] = defaultdict(int)
    for group in groups:
        installer = ""
        for photo in group.get("photos", []):
            installer = str(photo.get("creator") or "").strip()
            if installer:
                break
        if installer:
            counts[installer] += 1
    total = sum(counts.values())
    return [
        {"installer": installer, "group_count": count, "share": round(count / total, 4) if total else 0.0}
        for installer, count in sorted(counts.items(), key=lambda item: (-item[1], item[0]))[:8]
    ]


def _date_key_from_value(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    if len(text) >= 10 and re.match(r"^\d{4}-\d{2}-\d{2}", text):
        return text[:10]
    try:
        normalized = text.replace("Z", "+00:00")
        return datetime.fromisoformat(normalized).date().isoformat()
    except ValueError:
        return ""


def _datetime_from_value(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        result = value
    else:
        text = str(value or "").strip()
        if not text:
            return None
        if re.fullmatch(r"\d+(?:\.\d+)?", text):
            number = float(text)
            if 25569 <= number <= 60000:
                result = datetime(1899, 12, 30) + timedelta(days=number)
            elif number > 10_000_000_000:
                result = datetime.fromtimestamp(number / 1000, tz=UTC)
            elif number > 1_000_000_000:
                result = datetime.fromtimestamp(number, tz=UTC)
            else:
                return None
        else:
            normalized = text.replace("Z", "+00:00").replace("/", "-")
            if re.match(r"^\d{4}-\d{2}-\d{2} \d{2}:\d{2}", normalized):
                normalized = normalized.replace(" ", "T", 1)
            try:
                result = datetime.fromisoformat(normalized)
            except ValueError:
                return None
    if result.tzinfo is not None:
        return result.astimezone(LOCAL_WORK_TZ).replace(tzinfo=None)
    return result


def _format_time_of_day(value: datetime | None) -> str:
    return value.strftime("%H:%M") if value else ""


def _format_duration_label(minutes: int) -> str:
    minutes = max(0, int(minutes or 0))
    hours, rest = divmod(minutes, 60)
    if hours and rest:
        return f"{hours}小时{rest}分钟"
    if hours:
        return f"{hours}小时"
    return f"{rest}分钟"


def build_work_time_summary(timestamps: list[datetime | None]) -> dict[str, Any]:
    valid = sorted({value.replace(second=0, microsecond=0) for value in timestamps if value})
    start = valid[0] if valid else None
    end = valid[-1] if valid else None
    span_minutes = int(round((end - start).total_seconds() / 60)) if start and end and end > start else 0
    effective_minutes = 0
    buckets = {hour: 0 for hour in range(24)}
    if len(valid) >= 2:
        for left, right in zip(valid, valid[1:]):
            if right <= left:
                continue
            gap_minutes = int(round((right - left).total_seconds() / 60))
            if gap_minutes > WORK_SESSION_BREAK_MINUTES:
                continue
            effective_minutes += gap_minutes
            cursor = left
            while cursor < right:
                next_hour = (cursor.replace(minute=0, second=0, microsecond=0) + timedelta(hours=1))
                segment_end = min(right, next_hour)
                buckets[cursor.hour] += int(round((segment_end - cursor).total_seconds() / 60))
                cursor = segment_end
    return {
        "start_at": start.isoformat(timespec="minutes") if start else "",
        "end_at": end.isoformat(timespec="minutes") if end else "",
        "start_time": _format_time_of_day(start),
        "end_time": _format_time_of_day(end),
        "work_duration_minutes": effective_minutes,
        "work_duration_hours": round(effective_minutes / 60, 2) if effective_minutes else 0,
        "work_duration_label": _format_duration_label(effective_minutes),
        "work_span_minutes": span_minutes,
        "work_span_label": _format_duration_label(span_minutes),
        "break_threshold_minutes": WORK_SESSION_BREAK_MINUTES,
        "timepoint_count": len(valid),
        "hourly_segments": [
            {
                "hour": hour,
                "label": f"{hour:02d}:00",
                "minutes": minutes,
                "duration_label": _format_duration_label(minutes),
            }
            for hour, minutes in buckets.items()
        ],
    }


def _photo_is_construction_upload(photo: dict[str, Any]) -> bool:
    source_text = " ".join(
        str(photo.get(key) or "")
        for key in ("source", "upload_source", "storage_source", "source_file")
    ).lower()
    return "construction" in source_text


def _photo_work_date_key(photo: dict[str, Any]) -> str:
    if _photo_is_construction_upload(photo):
        return _date_key_from_value(photo.get("created_at")) or _date_key_from_value(photo.get("downloaded_at"))
    for key in (
        "scan_created_at",
        "source_created_at",
        "created_at",
        "\u521b\u5efa\u65f6\u95f4",
        "scan_time",
        "scanned_at",
        "taken_at",
        "classified_at",
    ):
        date_key = _date_key_from_value(photo.get(key))
        if date_key:
            return date_key
    return ""


def _photo_work_datetime(photo: dict[str, Any]) -> datetime | None:
    if _photo_is_construction_upload(photo):
        return _datetime_from_value(photo.get("created_at")) or _datetime_from_value(photo.get("downloaded_at"))
    for key in (
        "scan_created_at",
        "source_created_at",
        "created_at",
        "\u521b\u5efa\u65f6\u95f4",
        "scan_time",
        "scanned_at",
        "taken_at",
        "classified_at",
    ):
        value = _datetime_from_value(photo.get(key))
        if value:
            return value
    return None


def _installer_exception_group_payload(group: dict[str, Any]) -> dict[str, Any]:
    reasons = build_exception_reasons(group)
    return {
        "group_id": str(group.get("id") or ""),
        "meter_no": str(group.get("meter_no") or group.get("barcode") or ""),
        "terminal": str(group.get("terminal") or ""),
        "address": str(group.get("address") or ""),
        "status": str(group.get("status") or ""),
        "exception_note": str(group.get("exception_note") or group.get("review_note") or ""),
        "exception_reasons": reasons,
        "photo_count": int(group.get("photo_count") or len(group.get("photos", []) or [])),
    }


def installer_daily_workload(installer: str) -> dict[str, Any]:
    target = str(installer or "").strip()
    rows: dict[str, dict[str, Any]] = {}
    if not target:
        return {"installer": target, "items": []}
    for group in get_state()["groups"]:
        matched_photos = [
            photo
            for photo in group.get("photos", [])
            if str(photo.get("creator") or "").strip() == target and photo.get("is_active", True) is not False
        ]
        if not matched_photos:
            continue
        construction_dates = [
            _photo_work_date_key(photo)
            for photo in matched_photos
            if _photo_is_construction_upload(photo)
        ]
        date_key = max([date for date in construction_dates if date], default="")
        if not date_key:
            for photo in matched_photos:
                date_key = _photo_work_date_key(photo)
                if date_key:
                    break
        date_key = date_key or _date_key_from_value(group.get("last_photo_imported_at")) or "未记录日期"
        row = rows.setdefault(
            date_key,
            {
                "date": date_key,
                "group_count": 0,
                "photo_count": 0,
                "archived_count": 0,
                "exception_count": 0,
                "unreviewed_count": 0,
                "exception_groups": [],
                "_work_timestamps": [],
            },
        )
        photo_times = [value for value in (_photo_work_datetime(photo) for photo in matched_photos) if value]
        same_day_times = [value for value in photo_times if value.date().isoformat() == date_key]
        row["_work_timestamps"].extend(same_day_times or photo_times)
        row["group_count"] += 1
        row["photo_count"] += len(matched_photos)
        if is_reviewed_group(group):
            row["archived_count"] += 1
        elif is_problem_group(group):
            row["exception_count"] += 1
            row["exception_groups"].append(_installer_exception_group_payload(group))
        else:
            row["unreviewed_count"] += 1
    for row in rows.values():
        timestamps = row.pop("_work_timestamps", [])
        row.update(build_work_time_summary(timestamps))
    items = sorted(rows.values(), key=lambda item: str(item["date"]), reverse=True)
    return {"installer": target, "items": items}


def refresh_summary() -> None:
    state = get_state()
    refresh_group_exceptions()
    summary = build_summary([], [], [], state["groups"], state["stage_unmatched"], state["scan_unmatched"])
    summary["total_catalog_rows"] = state["summary"].get("total_catalog_rows", 0)
    summary["stage_catalog_rows"] = state["summary"].get("stage_catalog_rows", 0)
    summary["scan_rows"] = state["summary"].get("scan_rows", 0)
    state["summary"] = summary
    if state["projects"]:
        state["projects"][0]["summary"] = summary
    for task in state["tasks"]:
        ensure_construction_task_fields(task)
        task_groups = [group for group in state["groups"] if group["task_id"] == task["id"]]
        metrics = calculate_task_metrics(task_groups)
        task["address"] = first_task_address(task_groups)
        task["address_search_text"] = task_address_search_text(task_groups)
        task["total_groups"] = len(task_groups)
        task["completed_groups"] = sum(1 for group in task_groups if is_reviewed_group(group))
        task["exception_groups"] = sum(1 for group in task_groups if is_problem_group(group))
        task["incomplete_groups"] = count_incomplete_scanned_groups(task_groups)
        task["unconstructed_groups"] = count_unconstructed_groups(task_groups)
        task["pending_groups"] = sum(1 for group in task_groups if is_unreviewed_group(group))
        task["exception_order_count"] = sum(
            1
            for order in state.get("construction_exception_orders", [])
            if int(order.get("task_id") or 0) == int(task["id"]) and order.get("status") in {"open", "assigned", "submitted"}
        )
        task["scan_rows"] = sum(group["photo_count"] for group in task_groups)
        task["groups_with_scan"] = sum(1 for group in task_groups if group["photo_count"] > 0)
        task.update(metrics)
        task["complete_groups"] = count_complete_groups(task_groups)
        task["partial_groups"] = count_partial_groups(task_groups)
        task["has_scan_info"] = task["scan_rows"] > 0
        task["can_claim"] = task["has_scan_info"]
        task["claim_block_reason"] = "" if task["can_claim"] else "该终端暂无扫码信息，不能领取"
        task["progress"] = calculate_progress(task_groups)
        task["completeness_rate"] = metrics["upload_rate"]


def group_summary_counts(group: dict[str, Any]) -> dict[str, int]:
    return {
        "approved_groups": 1 if group.get("status") == "approved" else 0,
        "exception_groups": 1 if is_problem_group(group) else 0,
        "reviewed_groups": 1 if is_reviewed_group(group) else 0,
        "unreviewed_groups": 1 if is_unreviewed_group(group) else 0,
        "incomplete_groups": 1 if count_incomplete_scanned_groups([group]) else 0,
    }


def refresh_task_summary(task_id: int) -> None:
    state = get_state()
    task = find_task(task_id)
    ensure_construction_task_fields(task)
    task_groups = [group for group in state["groups"] if group["task_id"] == task_id]
    metrics = calculate_task_metrics(task_groups)
    task["address"] = first_task_address(task_groups)
    task["address_search_text"] = task_address_search_text(task_groups)
    task["total_groups"] = len(task_groups)
    task["completed_groups"] = sum(1 for group in task_groups if is_reviewed_group(group))
    task["exception_groups"] = sum(1 for group in task_groups if is_problem_group(group))
    task["incomplete_groups"] = count_incomplete_scanned_groups(task_groups)
    task["unconstructed_groups"] = count_unconstructed_groups(task_groups)
    task["pending_groups"] = sum(1 for group in task_groups if is_unreviewed_group(group))
    task["exception_order_count"] = sum(
        1
        for order in state.get("construction_exception_orders", [])
        if int(order.get("task_id") or 0) == int(task_id) and order.get("status") in {"open", "assigned", "submitted"}
    )
    task["scan_rows"] = sum(group["photo_count"] for group in task_groups)
    task["groups_with_scan"] = sum(1 for group in task_groups if group["photo_count"] > 0)
    task.update(metrics)
    task["complete_groups"] = count_complete_groups(task_groups)
    task["partial_groups"] = count_partial_groups(task_groups)
    task["has_scan_info"] = task["scan_rows"] > 0
    task["can_claim"] = task["has_scan_info"]
    task["claim_block_reason"] = "" if task["can_claim"] else "该终端暂无扫码信息，不能领取"
    task["progress"] = calculate_progress(task_groups)
    task["completeness_rate"] = metrics["upload_rate"]


def refresh_after_photo_classification(
    before_group: dict[str, Any],
    after_group: dict[str, Any],
    previous_category: str,
    next_category: str,
) -> None:
    state = get_state()
    summary = state["summary"]
    before_counts = group_summary_counts(before_group)
    after_counts = group_summary_counts(after_group)
    for key, before_value in before_counts.items():
        summary[key] = max(0, int(summary.get(key, 0)) + after_counts[key] - before_value)
    if previous_category == "unclassified" and next_category != "unclassified":
        summary["unclassified_photos"] = max(0, int(summary.get("unclassified_photos", 0)) - 1)
    elif previous_category != "unclassified" and next_category == "unclassified":
        summary["unclassified_photos"] = int(summary.get("unclassified_photos", 0)) + 1
    group_count = int(summary.get("groups", 0))
    summary["review_progress"] = round(int(summary.get("reviewed_groups", 0)) / group_count, 4) if group_count else 0.0
    if state["projects"]:
        state["projects"][0]["summary"] = summary
    refresh_task_summary(after_group["task_id"])


def calculate_progress(groups: list[dict[str, Any]]) -> float:
    if not groups:
        return 0.0
    reviewed = sum(1 for item in groups if is_reviewed_group(item))
    return round(reviewed / len(groups), 4)


def calculate_task_metrics(groups: list[dict[str, Any]]) -> dict[str, Any]:
    renovation_count = len(groups)
    uploaded_count = sum(1 for group in groups if group["photo_count"] > 0)
    reviewed_count = sum(1 for group in groups if is_reviewed_group(group))
    unreviewed_count = sum(1 for group in groups if group["photo_count"] > 0 and is_unreviewed_group(group))
    return {
        "renovation_count": renovation_count,
        "uploaded_count": uploaded_count,
        "reviewed_count": reviewed_count,
        "unreviewed_count": unreviewed_count,
        "upload_rate": round(uploaded_count / renovation_count, 4) if renovation_count else 0.0,
        "review_rate": round(reviewed_count / renovation_count, 4) if renovation_count else 0.0,
    }


def count_groups(groups: list[dict[str, Any]], statuses: set[str]) -> int:
    return sum(1 for item in groups if item["status"] in statuses)


def is_problem_group(group: dict[str, Any]) -> bool:
    return group.get("status") == "exception" or (
        group.get("photo_count", 0) > 0
        and (group.get("status") == "incomplete" or bool(group.get("has_archive_blocker")))
    )


def is_reviewed_group(group: dict[str, Any]) -> bool:
    return group.get("status") == "approved"


def is_unreviewed_group(group: dict[str, Any]) -> bool:
    return group.get("status") in OPEN_STATUSES and not is_problem_group(group)


def count_incomplete_scanned_groups(groups: list[dict[str, Any]]) -> int:
    return sum(
        1
        for item in groups
        if item.get("status") == "incomplete"
        and item.get("photo_count", 0) > 0
        and not is_problem_group(item)
    )


def count_unconstructed_groups(groups: list[dict[str, Any]]) -> int:
    return sum(1 for item in groups if item.get("photo_count", 0) == 0 and item.get("status") != "unmatched")


def count_complete_groups(groups: list[dict[str, Any]]) -> int:
    return sum(1 for item in groups if item["photo_count"] >= 4)


def count_partial_groups(groups: list[dict[str, Any]]) -> int:
    return sum(1 for item in groups if 0 < item["photo_count"] < 4)


def calculate_completeness_rate(groups: list[dict[str, Any]], scan_only: bool = False) -> float:
    scoped_groups = [item for item in groups if item["photo_count"] > 0] if scan_only else groups
    if not scoped_groups:
        return 0.0
    collected_slots = sum(min(item["photo_count"], 4) for item in scoped_groups)
    required_slots = len(scoped_groups) * 4
    return round(collected_slots / required_slots, 4)


def collect_module_group_map() -> dict[str, set[str]]:
    module_groups: dict[str, set[str]] = defaultdict(set)
    for group in get_state()["groups"]:
        for photo in group.get("photos", []):
            asset_no = str(photo.get("asset_no") or "").strip()
            if asset_no:
                module_groups[asset_no].add(group["id"])
    return module_groups


def validate_group_archive(group: dict[str, Any]) -> list[str]:
    photos = group.get("photos", [])
    reasons: list[str] = []
    if not photos:
        return reasons
    if len(photos) < 4:
        reasons.append("资料组照片不足 4 张")
    if photos and not any(str(photo.get("collector") or "").strip() for photo in photos):
        reasons.append("缺少采集器信息")
    if photos and not any(str(photo.get("asset_no") or "").strip() for photo in photos):
        reasons.append("缺少模块资产编号")
    module_groups = collect_module_group_map()
    duplicate_modules = sorted(
        {
            str(photo.get("asset_no") or "").strip()
            for photo in photos
            if str(photo.get("asset_no") or "").strip()
            and len(module_groups.get(str(photo.get("asset_no") or "").strip(), set()) - {group["id"]}) > 0
        }
    )
    if duplicate_modules:
        reasons.append(f"模块号重复: {', '.join(duplicate_modules[:3])}")
    return reasons


def set_group_exception_flags(group: dict[str, Any], reasons: list[str]) -> None:
    group["exception_flags"] = reasons
    group["exception_reasons"] = reasons
    group["has_archive_blocker"] = bool(reasons)


def refresh_group_exceptions() -> None:
    state = state_for_team()
    if not state.get("loaded"):
        return
    for group in state.get("groups", []):
        set_group_exception_flags(group, validate_group_archive(group))


def read_catalog_rows(path: Path, source: str) -> list[dict[str, Any]]:
    assert_file_exists(path)
    load_workbook = get_workbook_loader()
    workbook = load_workbook(path, read_only=True, data_only=True)
    return read_catalog_workbook_rows(workbook, source=source)


def read_catalog_xlsx_rows(content: bytes, source: str) -> list[dict[str, Any]]:
    load_workbook = get_workbook_loader()
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    return read_catalog_workbook_rows(workbook, source=source)


def read_catalog_workbook_rows(workbook, source: str) -> list[dict[str, Any]]:
    sheet = workbook.worksheets[0]
    rows = []
    blank_run = 0
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        terminal, meter_no, address = normalize_cell(row[0]), normalize_cell(row[1]), normalize_cell(row[2])
        if not terminal and not meter_no and not address:
            blank_run += 1
            if blank_run >= 20:
                break
            continue
        blank_run = 0
        if not meter_no:
            continue
        rows.append(
            {
                "source": source,
                "row_number": row_number,
                "terminal": terminal,
                "meter_no": meter_no,
                "address": address,
                "meter_match_key": build_total_catalog_match_key(meter_no),
            }
        )
    return rows


def read_scan_rows(path: Path) -> list[dict[str, Any]]:
    assert_file_exists(path)
    load_workbook = get_workbook_loader()
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    headers = [normalize_cell(value) for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    rows = []
    blank_run = 0
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        values = {headers[index]: normalize_cell(value) for index, value in enumerate(row) if index < len(headers)}
        barcode = values.get("\u626b\u7801\u5185\u5bb9", "")
        if not barcode:
            blank_run += 1
            if blank_run >= 20:
                break
            continue
        blank_run = 0
        try:
            meter_match_key = build_long_scan_match_key(barcode)
        except ValueError:
            meter_match_key = ""
        rows.append(
            {
                "row_number": row_number,
                "barcode": barcode,
                "meter_match_key": meter_match_key,
                "source_file": values.get("\u6765\u81ea\u6587\u4ef6", ""),
                "collector": values.get("\u91c7\u96c6\u5668", ""),
                "asset_no": values.get("\u6a21\u5757\u8d44\u4ea7\u7f16\u53f7", ""),
                "asset_type": values.get("\u8d44\u4ea7\u7c7b\u578b", ""),
                "creator": values.get("\u521b\u5efa\u8005", ""),
                "created_at": values.get("\u521b\u5efa\u65f6\u95f4", ""),
                "has_image": bool(values.get("\u56fe\u7247 (\u7535\u8111\u67e5\u770b)", "")),
            }
        )
    return rows


def list_groups(limit: int = 100, offset: int = 0, status: str | None = None) -> dict[str, Any]:
    state = get_state()
    groups = state["groups"]
    if status:
        groups = [item for item in groups if item["status"] == status]
    for group in groups[offset : offset + limit]:
        ensure_group_photo_storage_fields(group)
    return {"total": len(groups), "items": groups[offset : offset + limit]}


def ensure_group_photo_storage_fields(group: dict[str, Any]) -> dict[str, Any]:
    for photo in group.get("photos", []):
        ensure_photo_storage_fields(photo)
    return group


def group_target_text(group: dict[str, Any]) -> str:
    values = [
        group.get("id"),
        group.get("terminal"),
        group.get("meter_no"),
        group.get("meter_match_key"),
        group.get("address"),
        group.get("status"),
    ]
    for photo in group.get("photos", []):
        values.extend(
            [
                photo.get("barcode"),
                photo.get("collector"),
                photo.get("asset_no"),
                photo.get("creator"),
                photo.get("source_file"),
            ]
        )
    return " ".join(str(value or "") for value in values).lower()


def group_target_summary(group: dict[str, Any]) -> dict[str, Any]:
    photo_count = group.get("photo_count", 0)
    return {
        "id": group["id"],
        "task_id": group.get("task_id"),
        "terminal": group.get("terminal", ""),
        "meter_no": group.get("meter_no", ""),
        "meter_match_key": group.get("meter_match_key", ""),
        "address": group.get("address", ""),
        "status": group.get("status", ""),
        "reviewer": group.get("reviewer", ""),
        "review_note": group.get("review_note", ""),
        "photo_count": photo_count,
        "construction_status": "unconstructed" if photo_count == 0 else "scanned",
        "has_archive_blocker": group.get("has_archive_blocker", False),
        "exception_reasons": group.get("exception_reasons", []),
    }


def apply_construction_status(group: dict[str, Any]) -> dict[str, Any]:
    group["construction_status"] = "unconstructed" if group.get("photo_count", 0) == 0 else "scanned"
    return group


def review_queue_rank(group: dict[str, Any]) -> int:
    if is_group_fully_archived(group) or group.get("status") == "approved":
        return 3
    if group.get("status") == "exception" or group.get("has_archive_blocker"):
        return 1
    if group.get("photo_count", 0) == 0 and group.get("status") != "unmatched":
        return 2
    return 0


def search_group_targets(query: str = "", terminal: str = "", limit: int = 30, offset: int = 0) -> dict[str, Any]:
    groups = list(get_state()["groups"])
    if terminal:
        groups = [item for item in groups if str(item.get("terminal") or "") == terminal]
    q = query.strip().lower()
    if q:
        terms = [item for item in re.split(r"\s+", q) if item]
        groups = [item for item in groups if all(term in group_target_text(item) for term in terms)]
    groups = sorted(groups, key=lambda item: (str(item.get("terminal") or ""), str(item.get("meter_no") or ""), item["id"]))
    terminals = sorted({str(item.get("terminal") or "") for item in get_state()["groups"] if item.get("terminal")})
    return {
        "total": len(groups),
        "terminals": terminals,
        "items": [group_target_summary(item) for item in groups[offset : offset + limit]],
    }


def list_catalog_rows(
    catalog_type: str,
    query: str = "",
    terminal: str = "",
    limit: int = 100,
    offset: int = 0,
) -> dict[str, Any]:
    state = get_state()
    if catalog_type not in {"total", "stage"}:
        raise ValueError("Unsupported catalog type")
    rows = list(state["total_catalog"] if catalog_type == "total" else state["stage_catalog"])
    if terminal:
        rows = [item for item in rows if str(item.get("terminal") or "") == terminal]
    q = query.strip().lower()
    if q:
        terms = [item for item in re.split(r"\s+", q) if item]
        rows = [
            item
            for item in rows
            if all(
                term
                in " ".join(
                    str(item.get(field) or "")
                    for field in ["terminal", "meter_no", "address", "meter_match_key", "source"]
                ).lower()
                for term in terms
            )
        ]
    terminals = sorted({str(item.get("terminal") or "") for item in rows if item.get("terminal")})
    return {"total": len(rows), "terminals": terminals, "items": rows[offset : offset + limit]}


def list_task_groups(
    task_id: int,
    limit: int = 100,
    offset: int = 0,
    status: str | None = None,
    scan_only: bool = False,
    summary_only: bool = False,
) -> dict[str, Any]:
    find_task(task_id)
    state = get_state()
    groups = [item for item in state["groups"] if item["task_id"] == task_id]
    if scan_only:
        groups = [
            item
            for item in groups
            if item["photo_count"] >= 4 and item["status"] not in {"incomplete", "exception", "unmatched"}
        ]
    if status:
        groups = [item for item in groups if item["status"] == status]
    groups = sorted(
        groups,
        key=lambda group: (
            review_queue_rank(group),
            str(group.get("meter_no", "")),
            group["id"],
        ),
    )
    page = groups[offset : offset + limit]
    if summary_only:
        return {"total": len(groups), "items": [group_target_summary(item) for item in page]}
    return {"total": len(groups), "items": [apply_construction_status(item) for item in page]}


def list_exception_groups(
    reviewer: str = "",
    limit: int = 100,
    offset: int = 0,
) -> dict[str, Any]:
    groups = collect_exception_groups(reviewer=reviewer)
    return {"total": len(groups), "items": [group_target_summary(item) for item in groups[offset : offset + limit]]}


def collect_exception_groups(reviewer: str = "") -> list[dict[str, Any]]:
    refresh_group_exceptions()
    state = get_state()
    claimed_task_ids = {
        task["id"]
        for task in state["tasks"]
        if reviewer and task.get("claimed_by") == reviewer
    }
    groups = [
        item
        for item in state["groups"]
        if item.get("photo_count", 0) > 0
        and item.get("status") != "unmatched"
        and (item["status"] in {"incomplete", "exception"} or item.get("has_archive_blocker"))
    ]
    if reviewer:
        groups = [
            item
            for item in groups
            if item.get("task_id") in claimed_task_ids or not item.get("terminal")
        ]
    groups = sorted(
        groups,
        key=lambda group: (
            str(group.get("terminal") or "未关联终端"),
            str(group.get("meter_no", "")),
            group["id"],
        ),
    )
    return groups


def build_task_detail_export(task_id: int) -> bytes:
    find_task(task_id)
    groups = [item for item in get_state()["groups"] if item["task_id"] == task_id]
    return build_groups_export_workbook(groups, f"task-{task_id}")


def build_final_delivery_export(task_id: int | None = None, terminal: str = "", review_scope: str = "reviewed") -> bytes:
    groups = filter_delivery_groups(task_id=task_id, terminal=terminal, review_scope=review_scope)
    return build_groups_export_workbook(groups, "final-delivery")


def build_exception_meter_export(reviewer: str = "") -> bytes:
    groups = collect_exception_groups(reviewer=reviewer)
    return build_exception_meter_workbook(groups)


def build_project_outside_export() -> bytes:
    records = [
        ensure_unmatched_record(item)
        for item in get_state().get("scan_unmatched", [])
        if ensure_unmatched_record(item).get("project_outside")
    ]
    return build_project_outside_workbook(records)


def build_project_outside_workbook(records: list[dict[str, Any]]) -> bytes:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to export Excel files") from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "项目外施工"
    headers = [
        "记录ID",
        "表号/扫码内容",
        "短表号",
        "终端",
        "地址",
        "采集器",
        "模块资产编号",
        "安装人员",
        "照片数量",
        "记录人",
        "记录时间",
        "说明",
        "指派施工员",
        "来源文件",
    ]
    sheet.append(headers)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = "A1:N1"
    for record in records:
        sheet.append(
            [
                record.get("unmatched_id", ""),
                record.get("barcode") or record.get("meter_no") or "",
                record.get("meter_match_key") or "",
                record.get("terminal") or "",
                record.get("address") or "",
                record.get("collector") or "",
                record.get("module_asset_no") or record.get("asset_no") or "",
                record.get("creator") or "",
                len(split_urls(str(record.get("photo_urls") or ""))) or len(record.get("image_urls") or []),
                record.get("project_outside_by") or "",
                record.get("project_outside_at") or "",
                record.get("project_outside_note") or record.get("note") or "",
                record.get("assigned_to") or "",
                record.get("source_file") or "",
            ]
        )
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 10), 52)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_exception_meter_workbook(groups: list[dict[str, Any]]) -> bytes:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to export Excel files") from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "异常表计"
    headers = [
        "终端",
        "表号",
        "安装地址",
        "资料组ID",
        "异常类型",
        "异常原因",
        "现场处理建议",
        "照片数量",
        "已分类照片",
        "审阅人",
        "采集器",
        "模块资产编号",
        "安装人员",
        "第一张照片URL",
    ]
    sheet.append(headers)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = "A1:N1"
    for group in groups:
        photos = group.get("photos") or []
        first_photo = next((photo for photo in photos if photo.get("image_url")), photos[0] if photos else {})
        reasons = build_exception_reasons(group)
        reason_text = "；".join(reasons)
        sheet.append(
            [
                group.get("terminal", ""),
                group.get("meter_no", ""),
                group.get("address", ""),
                group.get("id", ""),
                build_exception_type(group),
                reason_text,
                build_field_work_suggestion(reason_text, group),
                group.get("photo_count", len(photos)),
                sum(1 for photo in photos if photo.get("category") not in {"", "unclassified", None}),
                group.get("reviewer") or "",
                first_photo.get("collector") or "",
                first_photo.get("asset_no") or first_photo.get("module_asset_no") or "",
                first_photo.get("creator") or "",
                first_photo.get("image_url") or "",
            ]
        )
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 10), 52)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_exception_type(group: dict[str, Any]) -> str:
    labels: list[str] = []
    if group.get("status") == "exception":
        labels.append("人工异常")
    if group.get("status") == "incomplete":
        labels.append("缺照片")
    if group.get("has_archive_blocker"):
        labels.append("归档问题")
    return "；".join(dict.fromkeys(labels)) or str(group.get("status") or "待处理")


def build_exception_reasons(group: dict[str, Any]) -> list[str]:
    reasons = [str(item).strip() for item in group.get("exception_reasons", []) if str(item).strip()]
    for note in (group.get("exception_note"), group.get("review_note")):
        note_text = str(note or "").strip()
        if note_text and note_text not in reasons:
            reasons.append(note_text)
    if not reasons:
        reasons.append(build_exception_type(group))
    return reasons


def build_field_work_suggestion(reason_text: str, group: dict[str, Any]) -> str:
    if "照片不足" in reason_text or group.get("status") == "incomplete":
        return "现场补拍或补传缺失照片"
    if "模块号重复" in reason_text:
        return "核对模块资产编号并更正重复记录"
    if "模块" in reason_text:
        return "核对模块资产编号并补录"
    if "采集器" in reason_text:
        return "核对采集器条码并补录"
    if group.get("has_archive_blocker"):
        return "核对照片分类和归档状态后重新归档"
    return "现场核对表号、地址、模块和照片后回填处理结果"


def build_final_delivery_manifest(task_id: int | None = None, terminal: str = "", review_scope: str = "reviewed") -> dict[str, Any]:
    groups = filter_delivery_groups(task_id=task_id, terminal=terminal, review_scope=review_scope)
    return {
        "generated_at": now_iso(),
        "photo_limit_per_group": 4,
        "scope": {"task_id": task_id, "terminal": terminal, "review_scope": review_scope},
        "groups": [build_delivery_group_manifest(group) for group in groups],
    }


def filter_delivery_groups(task_id: int | None = None, terminal: str = "", review_scope: str = "reviewed") -> list[dict[str, Any]]:
    terminal = terminal.strip()
    if task_id is None and not terminal:
        raise ValueError("Final delivery export must be scoped to one terminal")
    if review_scope not in {"reviewed", "all"}:
        raise ValueError("Unsupported delivery export scope")
    groups = list(get_state()["groups"])
    if task_id is not None:
        find_task(task_id)
        groups = [item for item in groups if item.get("task_id") == task_id]
    if terminal:
        groups = [item for item in groups if str(item.get("terminal") or "") == terminal]
    if review_scope == "reviewed":
        groups = [item for item in groups if is_reviewed_group(item)]
    return groups


def build_delivery_group_manifest(group: dict[str, Any]) -> dict[str, Any]:
    photos = group.get("photos") or []
    if is_reviewed_group(group) and group.get("delivery_cache_status") != "building" and not group_delivery_cache_ready(group):
        schedule_delivery_cache_build(str(group.get("id") or ""), current_team_id())
    return {
        "id": group.get("id", ""),
        "task_id": group.get("task_id"),
        "terminal": group.get("terminal", ""),
        "meter_no": group.get("meter_no", ""),
        "meter_match_key": group.get("meter_match_key", ""),
        "address": group.get("address", ""),
        "status": group.get("status", ""),
        "reviewer": group.get("reviewer") or "",
        "review_note": group.get("review_note") or "",
        "exception_note": group.get("exception_note") or "",
        "has_archive_blocker": group.get("has_archive_blocker", False),
        "exception_reasons": group.get("exception_reasons", []),
        "photo_count": len(photos),
        "delivery_cache_status": group.get("delivery_cache_status", "none"),
        "delivery_cache_built_at": group.get("delivery_cache_built_at", ""),
        "photos": [build_delivery_photo_manifest(group, photo, index) for index, photo in enumerate(photos, start=1)],
    }


def build_delivery_photo_manifest(group: dict[str, Any], photo: dict[str, Any], index: int) -> dict[str, Any]:
    ensure_photo_storage_fields(photo)
    category_label = photo.get("category_label") or PHOTO_CATEGORIES.get(photo.get("category"), PHOTO_CATEGORIES["unclassified"])
    return {
        "id": photo.get("id", ""),
        "index": index,
        "barcode": photo.get("barcode", ""),
        "collector": photo.get("collector", ""),
        "asset_no": photo.get("asset_no", ""),
        "creator": photo.get("creator", ""),
        "category": photo.get("category", "unclassified"),
        "category_label": category_label,
        "archive_filename": photo.get("archive_filename") or build_archive_filename(category_label, photo.get("image_url", "")),
        "image_url": photo.get("image_url", ""),
        "storage_type": photo.get("storage_type", ""),
        "storage_key": photo.get("storage_key", ""),
        "storage_bucket": photo.get("storage_bucket", ""),
        "sha256": photo.get("sha256", ""),
        "source_file": photo.get("source_file", ""),
        "delivery_cache_url": delivery_cache_url_for_photo(group, photo),
        "delivery_cache_status": photo.get("delivery_cache_status", "none"),
    }


def build_groups_export_workbook(groups: list[dict[str, Any]], sheet_title: str) -> bytes:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to export Excel files") from exc
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_title[:31]
    headers = [
        "终端",
        "表号",
        "地址",
        "资料组ID",
        "资料组状态",
        "审阅人",
        "照片数量",
        "照片序号",
        "照片分类",
        "归档文件名",
        "采集器",
        "模块",
        "安装人员",
        "照片URL",
    ]
    sheet.append(headers)
    for group in groups:
        photos = group.get("photos") or []
        if not photos:
            sheet.append(
                [
                    group.get("terminal", ""),
                    group.get("meter_no", ""),
                    group.get("address", ""),
                    group.get("id", ""),
                    group.get("status", ""),
                    group.get("reviewer") or "",
                    group.get("photo_count", 0),
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                ]
            )
            continue
        for index, photo in enumerate(photos, start=1):
            sheet.append(
                [
                    group.get("terminal", ""),
                    group.get("meter_no", ""),
                    group.get("address", ""),
                    group.get("id", ""),
                    group.get("status", ""),
                    group.get("reviewer") or "",
                    group.get("photo_count", 0),
                    index,
                    photo.get("category_label") or "",
                    photo.get("archive_filename") or "",
                    photo.get("collector") or "",
                    photo.get("asset_no") or "",
                    photo.get("creator") or "",
                    photo.get("image_url") or "",
                ]
            )
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 10), 48)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def get_group(group_id: str) -> dict[str, Any] | None:
    state = get_state()
    group = next((item for item in state["groups"] if item["id"] == group_id), None)
    return ensure_group_photo_storage_fields(group) if group is not None else None


def list_tasks() -> list[dict[str, Any]]:
    state = get_state()
    groups_by_task: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for group in state["groups"]:
        groups_by_task[int(group.get("task_id") or 0)].append(group)
    for task in state["tasks"]:
        ensure_construction_task_fields(task)
        task_groups = groups_by_task.get(int(task.get("id") or 0), [])
        task["address"] = first_task_address(task_groups)
        task["address_search_text"] = task_address_search_text(task_groups)
    return sorted(
        state["tasks"],
        key=lambda task: (
            not task.get("can_claim", False),
            str(task.get("terminal", "")),
            task["id"],
        ),
    )


def get_task_progress(task_id: int) -> dict[str, Any]:
    task = find_task(task_id)
    groups = [item for item in get_state()["groups"] if item["task_id"] == task_id]
    by_status = {status: 0 for status in sorted(REVIEWABLE_STATUSES)}
    for group in groups:
        by_status[group["status"]] = by_status.get(group["status"], 0) + 1
    return {
        "task_id": task_id,
        "status": task["status"],
        "claimed_by": task.get("claimed_by"),
        "total_groups": len(groups),
        "reviewed_groups": sum(1 for group in groups if is_reviewed_group(group)),
        "pending_groups": sum(1 for group in groups if is_unreviewed_group(group)),
        "approved_groups": by_status.get("approved", 0),
        "exception_groups": sum(1 for group in groups if is_problem_group(group)),
        "incomplete_groups": count_incomplete_scanned_groups(groups),
        "unconstructed_groups": count_unconstructed_groups(groups),
        "complete_groups": count_complete_groups(groups),
        "partial_groups": count_partial_groups(groups),
        "by_status": by_status,
        "progress": calculate_progress(groups),
        "completeness_rate": calculate_completeness_rate(groups, scan_only=True),
    }


def claim_task(task_id: int, reviewer: str) -> dict[str, Any]:
    task = find_task(task_id)
    if not task.get("can_claim", False):
        raise ValueError(task.get("claim_block_reason") or "Task has no scan information")
    if task["status"] not in {"published", "released", "in_review"}:
        raise ValueError(f"Task cannot be claimed from status {task['status']}")
    if task.get("claimed_by") and task["claimed_by"] != reviewer:
        raise ValueError("Task is already claimed by another reviewer")
    task["status"] = "in_review"
    task["claimed_by"] = reviewer
    task["claimed_at"] = now_iso()
    task["released_at"] = None
    return task


def release_task(task_id: int, reviewer: str, force: bool = False) -> dict[str, Any]:
    task = find_task(task_id)
    if not force and task.get("claimed_by") not in {None, reviewer}:
        raise ValueError("Only the current reviewer can release this task")
    task["status"] = "released"
    task["claimed_by"] = None
    task["released_at"] = now_iso()
    return task


def release_all_claimed_tasks(actor: str) -> dict[str, Any]:
    released_ids: list[int] = []
    for task in get_state()["tasks"]:
        if not task.get("claimed_by"):
            continue
        task["status"] = "released"
        task["claimed_by"] = None
        task["released_at"] = now_iso()
        released_ids.append(task["id"])
    return {"released": len(released_ids), "task_ids": released_ids, "actor": actor}


def open_construction_task(task_id: int, actor: str) -> dict[str, Any]:
    task = ensure_construction_task_fields(find_task(task_id))
    task["construction_enabled"] = True
    task["construction_opened_by"] = actor
    task["construction_opened_at"] = now_iso()
    task["construction_closed_at"] = None
    ensure_construction_task_fields(task)
    append_audit_event("construction_task_open", actor, {"task_id": task_id, "terminal": task.get("terminal", "")})
    return task


def close_construction_task(task_id: int, actor: str) -> dict[str, Any]:
    task = ensure_construction_task_fields(find_task(task_id))
    task["construction_enabled"] = False
    task["construction_closed_at"] = now_iso()
    ensure_construction_task_fields(task)
    append_audit_event("construction_task_close", actor, {"task_id": task_id, "terminal": task.get("terminal", "")})
    return task


def active_construction_tasks_for(actor: str, excluding_task_id: int | None = None) -> list[dict[str, Any]]:
    actor = actor.strip()
    if not actor:
        return []
    tasks: list[dict[str, Any]] = []
    for item in get_state()["tasks"]:
        task = ensure_construction_task_fields(item)
        if excluding_task_id is not None and int(task.get("id") or 0) == int(excluding_task_id):
            continue
        if task.get("construction_enabled") and task.get("construction_claimed_by") == actor:
            tasks.append(task)
    return tasks


def active_construction_task_for(actor: str, excluding_task_id: int | None = None) -> dict[str, Any] | None:
    tasks = active_construction_tasks_for(actor, excluding_task_id=excluding_task_id)
    return tasks[0] if tasks else None


def assign_construction_task(
    task_id: int,
    actor: str,
    constructor: str,
    note: str = "",
    due_date: str = "",
) -> dict[str, Any]:
    actor = actor.strip() or "admin"
    constructor = constructor.strip()
    if not constructor:
        raise ValueError("Constructor is required")
    task = ensure_construction_task_fields(find_task(task_id))
    active_tasks = active_construction_tasks_for(constructor, excluding_task_id=task_id)
    if len(active_tasks) >= MAX_ACTIVE_CONSTRUCTION_TASKS_PER_CONSTRUCTOR:
        terminals = ", ".join(str(item.get("terminal") or item.get("id") or "") for item in active_tasks[:3])
        raise ValueError(
            f"Current constructor already has {MAX_ACTIVE_CONSTRUCTION_TASKS_PER_CONSTRUCTOR} active terminals"
            f"{': ' + terminals if terminals else ''}"
        )
    previous = task.get("construction_claimed_by") or ""
    task["construction_enabled"] = True
    task["construction_claimed_by"] = constructor
    task["construction_claimed_at"] = now_iso()
    task["construction_released_at"] = None
    task["construction_opened_by"] = actor
    task["construction_opened_at"] = task.get("construction_opened_at") or now_iso()
    task["construction_assignment_note"] = note.strip()
    task["construction_due_date"] = due_date.strip()
    ensure_construction_task_fields(task)
    for order in construction_exception_orders():
        if int(order.get("task_id") or 0) == int(task_id) and order.get("status") in {"open", "assigned"}:
            order["assigned_to"] = constructor
            order["status"] = "assigned"
            order["updated_at"] = now_iso()
    append_audit_event(
        "construction_task_assign",
        actor,
        {
            "task_id": task_id,
            "terminal": task.get("terminal", ""),
            "constructor": constructor,
            "previous_constructor": previous,
            "note": note,
            "due_date": due_date,
        },
    )
    return task


def unassign_construction_task(task_id: int, actor: str, reason: str = "") -> dict[str, Any]:
    actor = actor.strip() or "admin"
    task = ensure_construction_task_fields(find_task(task_id))
    previous = task.get("construction_claimed_by") or ""
    task["construction_claimed_by"] = None
    task["construction_released_at"] = now_iso()
    task["construction_assignment_note"] = ""
    ensure_construction_task_fields(task)
    for order in construction_exception_orders():
        if int(order.get("task_id") or 0) == int(task_id) and order.get("status") == "assigned":
            order["assigned_to"] = ""
            order["status"] = "open"
            order["updated_at"] = now_iso()
    append_audit_event(
        "construction_task_unassign",
        actor,
        {"task_id": task_id, "terminal": task.get("terminal", ""), "constructor": previous, "reason": reason},
    )
    return task


def claim_construction_task(task_id: int, actor: str) -> dict[str, Any]:
    actor = actor.strip() or "constructor"
    task = ensure_construction_task_fields(find_task(task_id))
    if not task.get("construction_enabled"):
        raise ValueError("该终端尚未开放施工")
    if task.get("construction_claimed_by") != actor:
        raise ValueError("Construction task must be assigned by an administrator before entry")
    ensure_construction_task_fields(task)
    append_audit_event("construction_task_enter", actor, {"task_id": task_id, "terminal": task.get("terminal", "")})
    return task


def release_construction_task(task_id: int, actor: str, force: bool = False) -> dict[str, Any]:
    actor = actor.strip() or "constructor"
    task = ensure_construction_task_fields(find_task(task_id))
    if not force and task.get("construction_claimed_by") not in {None, actor}:
        raise ValueError("只有当前施工员可以释放该终端")
    task["construction_claimed_by"] = None
    task["construction_released_at"] = now_iso()
    ensure_construction_task_fields(task)
    append_audit_event("construction_task_release", actor, {"task_id": task_id, "terminal": task.get("terminal", "")})
    return task


def list_construction_tasks(actor: str = "", include_closed: bool = False) -> list[dict[str, Any]]:
    actor = actor.strip()
    all_tasks = [ensure_construction_task_fields(task) for task in list_tasks()]
    assigned_to_actor = [task for task in all_tasks if actor and task.get("construction_claimed_by") == actor]
    if actor and not include_closed:
        return sorted(assigned_to_actor, key=lambda task: (str(task.get("terminal", "")), task["id"]))
    tasks = []
    for task in all_tasks:
        if include_closed:
            tasks.append(task)
            continue
        if task.get("construction_enabled") and task.get("construction_claimed_by"):
            tasks.append(task)
    return sorted(
        tasks,
        key=lambda task: (
            -int(task.get("uploaded_count") or 0) if include_closed else 0,
            -int(task.get("unconstructed_groups") or 0) if include_closed else 0,
            -int(task.get("exception_order_count") or 0) if include_closed else 0,
            not bool(task.get("construction_claimed_by") == actor),
            str(task.get("terminal", "")),
            task["id"],
        ),
    )


def list_construction_task_groups(
    task_id: int,
    limit: int = 100,
    offset: int = 0,
    status: str | None = None,
    summary_only: bool = False,
) -> dict[str, Any]:
    find_task(task_id)
    state = get_state()
    groups = [
        item
        for item in state["groups"]
        if item["task_id"] == task_id
        and item.get("photo_count", 0) == 0
        and item.get("status") not in {"unmatched", "exception"}
    ]
    if status:
        groups = [item for item in groups if item.get("status") == status]
    groups = sorted(
        groups,
        key=lambda group: (
            str(group.get("meter_no", "")),
            group["id"],
        ),
    )
    page = groups[offset : offset + limit]
    if summary_only:
        return {"total": len(groups), "items": [group_target_summary(item) for item in page]}
    return {"total": len(groups), "items": [apply_construction_status(item) for item in page]}


def upload_construction_group_batch(
    group_id: str,
    actor: str,
    client_batch_id: str,
    collector: str,
    module_asset_no: str,
    photos: list[dict[str, Any]],
    creator: str = "",
) -> dict[str, Any]:
    actor = actor.strip() or "constructor"
    creator = creator.strip() or actor
    client_batch_id = client_batch_id.strip()
    if not client_batch_id:
        raise ValueError("Client batch id is required")
    group = get_group(group_id)
    if group is None:
        raise KeyError(group_id)
    task = ensure_construction_task_fields(find_task(group["task_id"]))
    if task.get("construction_claimed_by") != actor:
        raise ValueError("Construction task must be claimed by the current constructor before upload")
    existing_composite = {
        make_construction_photo_unique_key(photo)
        for photo in group.get("photos", [])
        if photo.get("client_batch_id") or photo.get("client_photo_id") or photo.get("sha256")
    }
    existing_sha = {
        str(photo.get("sha256") or "")
        for photo in group.get("photos", [])
        if str(photo.get("sha256") or "")
    }
    added = 0
    skipped_duplicates = 0
    for item in photos:
        url = str(item.get("url") or "").strip()
        sha256 = str(item.get("sha256") or "").strip()
        client_photo_id = str(item.get("client_photo_id") or "").strip()
        slot = str(item.get("slot") or "other").strip() or "other"
        if not url:
            continue
        composite = (client_batch_id, client_photo_id, sha256)
        if composite in existing_composite or (sha256 and sha256 in existing_sha):
            skipped_duplicates += 1
            continue
        row = {
            "row_number": f"construction-{group_id}-{len(group['photos']) + 1}",
            "barcode": group.get("meter_no", ""),
            "meter_match_key": group.get("meter_match_key", ""),
            "source_file": "construction-mobile",
            "collector": collector,
            "asset_no": module_asset_no,
            "asset_type": "",
            "creator": creator,
            "created_at": now_iso(),
            "has_image": True,
            "image_file_id": "",
            "image_url": url,
        }
        photo = build_photo_record(len(group["photos"]) + 1, row)
        slot_category = CONSTRUCTION_SLOT_CATEGORIES.get(slot, "other")
        photo.update(
            {
                "client_batch_id": client_batch_id,
                "client_photo_id": client_photo_id,
                "sha256": sha256,
                "construction_slot": slot,
                "construction_slot_label": PHOTO_CATEGORIES.get(slot_category, PHOTO_CATEGORIES["other"]),
                "upload_source": "construction-mobile",
                "storage_type": item.get("storage_type") or "local_upload",
                "storage_key": item.get("storage_key") or str(url).removeprefix("/static/uploads/"),
                "storage_bucket": item.get("storage_bucket", ""),
                "storage_source": item.get("storage_source") or "construction-mobile",
                "original_filename": item.get("filename", ""),
                "category": "unclassified",
                "category_label": PHOTO_CATEGORIES["unclassified"],
                "archive_status": "pending",
                "archive_filename": "",
                "archived_at": None,
            }
        )
        group["photos"].append(photo)
        existing_composite.add(composite)
        if sha256:
            existing_sha.add(sha256)
        added += 1
    group["photo_count"] = len(group["photos"])
    if collector:
        group["construction_collector"] = collector
    if module_asset_no:
        group["construction_module_asset_no"] = module_asset_no
    group["constructor"] = actor
    group["construction_updated_at"] = now_iso()
    if group["status"] == "exception":
        group["status"] = "pending"
        group["has_archive_blocker"] = False
        for order in construction_exception_orders():
            if order.get("group_id") == group_id and order.get("status") in {"open", "assigned"}:
                order["status"] = "submitted"
                order["submitted_by"] = actor
                order["submitted_at"] = now_iso()
                order["updated_at"] = now_iso()
    elif group["status"] in DONE_STATUSES or group["status"] in OPEN_STATUSES:
        group["status"] = "incomplete" if group["photo_count"] < 4 else "pending"
    group["reviewer"] = None
    group["review_note"] = ""
    group["exception_note"] = ""
    group["reviewed_at"] = None
    mark_delivery_cache_stale(group, "construction upload changed photos")
    append_audit_event(
        "construction_upload_batch",
        actor,
        {
            "group_id": group_id,
            "task_id": group.get("task_id"),
            "client_batch_id": client_batch_id,
            "added": added,
            "skipped_duplicates": skipped_duplicates,
        },
    )
    refresh_summary()
    return {"group": group, "task": task, "added": added, "skipped_duplicates": skipped_duplicates}


def review_group(
    group_id: str,
    status: str,
    reviewer: str,
    note: str = "",
    exception_note: str = "",
) -> dict[str, Any]:
    if status not in REVIEWABLE_STATUSES:
        raise ValueError(f"Unsupported group status: {status}")
    group = get_group(group_id)
    if group is None:
        raise KeyError(group_id)
    ensure_task_claimed_by(group, reviewer)
    if status == "exception" and not (note or exception_note):
        raise ValueError("Exception review requires a note")
    previous = group["status"]
    group["status"] = status
    group["reviewer"] = reviewer
    group["review_note"] = note
    group["exception_note"] = exception_note if status == "exception" else ""
    group["reviewed_at"] = now_iso() if status in DONE_STATUSES else None
    state = get_state()
    state["review_events"].append(
        {
            "group_id": group_id,
            "task_id": group["task_id"],
            "previous_status": previous,
            "next_status": status,
            "reviewer": reviewer,
            "note": note,
            "exception_note": group["exception_note"],
            "created_at": now_iso(),
        }
    )
    if status == "approved":
        schedule_delivery_cache_build(group_id, current_team_id())
    else:
        mark_delivery_cache_stale(group, f"review status changed to {status}")
    refresh_summary()
    return group


def save_exception_note(group_id: str, reviewer: str, note: str) -> dict[str, Any]:
    return review_group(group_id, status="exception", reviewer=reviewer, exception_note=note)


def exception_category_label(category: str) -> str:
    value = str(category or "").strip()
    if value in CONSTRUCTION_EXCEPTION_CATEGORIES:
        return CONSTRUCTION_EXCEPTION_CATEGORIES[value]
    if value in CONSTRUCTION_EXCEPTION_CATEGORIES.values():
        return value
    return CONSTRUCTION_EXCEPTION_CATEGORIES["other"]


def construction_exception_orders() -> list[dict[str, Any]]:
    return state_for_team().setdefault("construction_exception_orders", [])


def create_construction_exception_order(
    group: dict[str, Any],
    *,
    actor: str,
    category: str,
    note: str,
) -> dict[str, Any]:
    orders = construction_exception_orders()
    task = find_task(group["task_id"])
    label = exception_category_label(category)
    existing = next(
        (
            item
            for item in orders
            if item.get("group_id") == group.get("id")
            and item.get("status") in {"open", "assigned", "submitted"}
        ),
        None,
    )
    payload = {
        "meter_no": group.get("meter_no", ""),
        "collector": group.get("construction_collector") or groupPhotoValue(group, "collector"),
        "module_asset_no": group.get("construction_module_asset_no") or groupPhotoValue(group, "asset_no"),
        "address": group.get("address", ""),
    }
    assigned_to = str(task.get("construction_claimed_by") or "")
    if existing:
        existing.update(
            {
                "status": "assigned" if assigned_to else "open",
                "category": label,
                "note": note,
                "assigned_to": assigned_to,
                "payload": {**existing.get("payload", {}), **payload},
                "updated_at": now_iso(),
            }
        )
        return existing
    order = {
        "id": f"ceo-{len(orders) + 1:06d}",
        "team_id": current_team_id(),
        "task_id": group.get("task_id"),
        "group_id": group.get("id"),
        "terminal": group.get("terminal", ""),
        "status": "assigned" if assigned_to else "open",
        "category": label,
        "note": note,
        "assigned_to": assigned_to,
        "created_by": actor,
        "submitted_by": "",
        "created_at": now_iso(),
        "updated_at": now_iso(),
        "submitted_at": None,
        "resolved_at": None,
        "payload": payload,
    }
    orders.append(order)
    return order


def list_construction_exception_orders(actor: str = "", task_id: int | None = None) -> list[dict[str, Any]]:
    actor = actor.strip()
    orders = list(construction_exception_orders())
    if task_id is not None:
        orders = [item for item in orders if int(item.get("task_id") or 0) == int(task_id)]
    if actor:
        orders = [item for item in orders if item.get("assigned_to") == actor]
    return sorted(orders, key=lambda item: (str(item.get("terminal") or ""), str(item.get("created_at") or ""), str(item.get("id") or "")))


def assign_construction_exception_order(
    order_id: str,
    *,
    actor: str,
    constructor: str,
    note: str = "",
    due_date: str = "",
) -> dict[str, Any]:
    constructor = constructor.strip()
    if not constructor:
        raise ValueError("Constructor is required")
    orders = construction_exception_orders()
    for index, item in enumerate(orders):
        order = ensure_construction_exception_order(item)
        if order.get("id") != order_id:
            continue
        task_id = int(order.get("task_id") or 0)
        if task_id:
            assign_construction_task(task_id, actor=actor, constructor=constructor, note=note, due_date=due_date)
        order["assigned_to"] = constructor
        order["assigned_by"] = actor
        order["assigned_at"] = now_iso()
        order["assignment_note"] = note.strip()
        order["due_date"] = due_date.strip()
        if order.get("status") == "open":
            order["status"] = "assigned"
        orders[index] = order
        append_audit_event(
            "assign_construction_exception_order",
            actor,
            {"order_id": order_id, "constructor": constructor, "note": note, "due_date": due_date},
        )
        return order
    raise KeyError(order_id)


def unassign_construction_exception_order(order_id: str, *, actor: str, reason: str = "") -> dict[str, Any]:
    orders = construction_exception_orders()
    for index, item in enumerate(orders):
        order = ensure_construction_exception_order(item)
        if order.get("id") != order_id:
            continue
        previous = order.get("assigned_to") or ""
        order["assigned_to"] = ""
        order["unassigned_by"] = actor
        order["unassigned_at"] = now_iso()
        order["unassign_reason"] = reason.strip()
        if order.get("status") == "assigned":
            order["status"] = "open"
        orders[index] = order
        append_audit_event(
            "unassign_construction_exception_order",
            actor,
            {"order_id": order_id, "previous_constructor": previous, "reason": reason},
        )
        return order
    raise KeyError(order_id)


def submit_construction_exception_order(
    order_id: str,
    actor: str,
    updates: dict[str, Any] | None = None,
    note: str = "",
) -> dict[str, Any]:
    actor = actor.strip() or "constructor"
    order = next((item for item in construction_exception_orders() if item.get("id") == order_id), None)
    if order is None:
        raise KeyError(order_id)
    if order.get("assigned_to") and order.get("assigned_to") != actor:
        raise ValueError("Exception order is assigned to another constructor")
    group = get_group(str(order.get("group_id") or ""))
    if group is None:
        raise KeyError(str(order.get("group_id") or ""))
    task = ensure_construction_task_fields(find_task(group["task_id"]))
    if task.get("construction_claimed_by") != actor:
        raise ValueError("Construction task must be assigned to the current constructor")
    updates = updates or {}
    meter_no = str(updates.get("meter_no") or updates.get("barcode") or "").strip()
    collector = str(updates.get("collector") or "").strip()
    module_asset_no = str(updates.get("module_asset_no") or updates.get("asset_no") or "").strip()
    if meter_no:
        group["meter_no"] = meter_no
    if collector:
        group["construction_collector"] = collector
    if module_asset_no:
        group["construction_module_asset_no"] = module_asset_no
    order["status"] = "submitted"
    order["submitted_by"] = actor
    order["submitted_at"] = now_iso()
    order["updated_at"] = now_iso()
    order["payload"] = {
        **(order.get("payload") or {}),
        "meter_no": group.get("meter_no", ""),
        "collector": collector or group.get("construction_collector", ""),
        "module_asset_no": module_asset_no or group.get("construction_module_asset_no", ""),
        "note": note.strip(),
    }
    group["status"] = "pending"
    group["reviewer"] = None
    group["review_note"] = ""
    group["exception_note"] = note.strip()
    group["has_archive_blocker"] = False
    group["reviewed_at"] = None
    mark_delivery_cache_stale(group, "construction exception order submitted")
    append_audit_event(
        "construction_exception_order_submit",
        actor,
        {"order_id": order_id, "group_id": group["id"], "task_id": group["task_id"], "updates": updates},
    )
    refresh_summary()
    return {"order": order, "group": group}


def groupPhotoValue(group: dict[str, Any], field: str) -> str:
    for photo in group.get("photos", []):
        value = str(photo.get(field) or "").strip()
        if value:
            return value
    return ""


def reset_group_to_unconstructed(group_id: str, actor: str, reason: str = "", force: bool = False) -> dict[str, Any]:
    group = get_group(group_id)
    if group is None:
        raise KeyError(group_id)
    if not force:
        ensure_task_claimed_by(group, actor)
    moved = []
    for photo in group.get("photos", []):
        ensure_photo_storage_fields(photo)
        photo["is_active"] = False
        photo["deleted_at"] = now_iso()
        photo["deleted_by"] = actor
        photo["delete_reason"] = reason or "reset_to_unconstructed"
        moved.append(photo)
    group.setdefault("deleted_photos", []).extend(moved)
    group["photos"] = []
    group["photo_count"] = 0
    for key in ("construction_collector", "construction_module_asset_no", "constructor"):
        group.pop(key, None)
    group["status"] = "pending"
    group["reviewer"] = None
    group["review_note"] = ""
    group["exception_note"] = ""
    group["exception_reasons"] = []
    group["has_archive_blocker"] = False
    group["reviewed_at"] = None
    group["reset_to_unconstructed_at"] = now_iso()
    mark_delivery_cache_stale(group, "reset to unconstructed")
    append_audit_event(
        "group_reset_to_unconstructed",
        actor,
        {"group_id": group_id, "soft_deleted_photos": len(moved), "reason": reason},
    )
    refresh_summary()
    return {"group": group, "soft_deleted_photos": len(moved)}


def return_group_to_exception_order(
    group_id: str,
    actor: str,
    category: str,
    note: str,
    force: bool = False,
) -> dict[str, Any]:
    group = get_group(group_id)
    if group is None:
        raise KeyError(group_id)
    if not force:
        ensure_task_claimed_by(group, actor)
    note = note.strip()
    if not note:
        raise ValueError("Exception reason is required")
    label = exception_category_label(category)
    group["status"] = "exception"
    group["reviewer"] = actor
    group["review_note"] = ""
    group["exception_note"] = note
    reasons = [label]
    if note and note not in reasons:
        reasons.append(note)
    group["exception_reasons"] = reasons
    group["has_archive_blocker"] = True
    group["reviewed_at"] = None
    mark_delivery_cache_stale(group, "returned to exception order")
    order = create_construction_exception_order(group, actor=actor, category=label, note=note)
    group["exception_work_order_id"] = order["id"]
    append_audit_event(
        "group_returned_to_exception_order",
        actor,
        {"group_id": group_id, "order_id": order["id"], "category": label, "note": note},
    )
    refresh_summary()
    return {"group": group, "order": order}


def classify_photo(group_id: str, photo_id: str, category: str, reviewer: str) -> dict[str, Any]:
    if category not in PHOTO_CATEGORIES:
        raise ValueError(f"Unsupported photo category: {category}")
    group = get_group(group_id)
    if group is None:
        raise KeyError(group_id)
    ensure_task_claimed_by(group, reviewer)
    before_group = group.copy()
    photo = next((item for item in group["photos"] if item["id"] == photo_id), None)
    if photo is None:
        raise KeyError(photo_id)
    if photo.get("download_status") != "downloaded":
        has_previewable_source = bool(
            str(photo.get("image_url") or "").strip()
            or str(photo.get("storage_key") or "").strip()
            or str(photo.get("storage_type") or "").strip() in {"oss", "local_upload", "external_url"}
        )
        if not has_previewable_source:
            raise ValueError("Photo must have an image URL before classification")
        photo["download_status"] = "downloaded"
    previous = photo["category"]
    photo["category"] = category
    photo["category_label"] = PHOTO_CATEGORIES[category]
    photo["classified_by"] = reviewer
    photo["classified_at"] = now_iso()
    photo["archive_status"] = "archived"
    photo["archive_filename"] = build_archive_filename(photo["category_label"], photo.get("image_url", ""))
    photo["archived_at"] = now_iso()
    state = get_state()
    state["photo_events"].append(
        {
            "group_id": group_id,
            "photo_id": photo_id,
            "previous_category": previous,
            "next_category": category,
            "reviewer": reviewer,
            "created_at": now_iso(),
        }
    )
    update_group_archive_status(group, reviewer)
    refresh_after_photo_classification(before_group, group, previous, category)
    return photo


def delete_group_photo(group_id: str, photo_id: str, reviewer: str) -> dict[str, Any]:
    group = get_group(group_id)
    if group is None:
        raise KeyError(group_id)
    ensure_task_claimed_by(group, reviewer)
    photos = group.get("photos", [])
    photo = next((item for item in photos if item["id"] == photo_id), None)
    if photo is None:
        raise KeyError(photo_id)
    group["photos"] = [item for item in photos if item["id"] != photo_id]
    group["photo_count"] = len(group["photos"])
    for index, item in enumerate(group["photos"], start=1):
        item["index"] = index
    group["status"] = "incomplete" if group["photo_count"] < 4 else "pending"
    group["reviewer"] = None
    group["review_note"] = ""
    group["exception_note"] = ""
    group["reviewed_at"] = None
    set_group_exception_flags(group, validate_group_archive(group))
    state = get_state()
    state["photo_events"].append(
        {
            "group_id": group_id,
            "photo_id": photo_id,
            "previous_category": photo.get("category", "unclassified"),
            "next_category": "deleted",
            "reviewer": reviewer,
            "created_at": now_iso(),
        }
    )
    append_audit_event(
        "delete_group_photo",
        reviewer,
        {
            "group_id": group_id,
            "photo_id": photo_id,
            "photo_count": group["photo_count"],
            "image_url": photo.get("image_url", ""),
        },
    )
    mark_delivery_cache_stale(group, "photo deleted")
    refresh_summary()
    return {"group": group, "deleted_photo": photo}


def ensure_task_claimed_by(group: dict[str, Any], reviewer: str) -> None:
    task = find_task(group["task_id"])
    if task.get("claimed_by") != reviewer:
        raise ValueError("Task must be claimed by the current reviewer before review or classification")


def is_group_fully_archived(group: dict[str, Any]) -> bool:
    photos = group.get("photos", [])
    return bool(photos) and all(photo.get("archive_status") == "archived" for photo in photos)


def update_group_archive_status(group: dict[str, Any], reviewer: str) -> None:
    if not is_group_fully_archived(group):
        return
    reasons = validate_group_archive(group)
    set_group_exception_flags(group, reasons)
    if reasons:
        group["status"] = "exception"
        group["exception_note"] = "; ".join(reasons)
        mark_delivery_cache_stale(group, "archive blocked")
        append_audit_event("archive_blocked", reviewer, {"group_id": group["id"], "reasons": reasons})
        return
    group["status"] = "approved"
    group["reviewer"] = reviewer
    group["review_note"] = "分类完成"
    group["exception_note"] = ""
    group["reviewed_at"] = now_iso()
    schedule_delivery_cache_build(group["id"], current_team_id())


def find_task(task_id: int) -> dict[str, Any]:
    state = get_state()
    task = next((item for item in state["tasks"] if item["id"] == task_id), None)
    if task is None:
        raise KeyError(task_id)
    return task


def normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return repair_mojibake(str(value).strip())


def repair_mojibake(value: str) -> str:
    if not value:
        return ""
    try:
        repaired = value.encode("latin1").decode("utf-8")
    except (UnicodeEncodeError, UnicodeDecodeError):
        return value
    if repaired != value and any("\u4e00" <= char <= "\u9fff" for char in repaired):
        return repaired
    return value


def assert_file_exists(path: Path) -> None:
    if not path.exists():
        raise FileNotFoundError(f"Test workbook not found: {path}")


def now_iso() -> str:
    return datetime.now(UTC).isoformat()


def build_archive_filename(category_label: str, image_url: str) -> str:
    suffix = Path(image_url.split("?", 1)[0]).suffix.lower() if image_url else ".jpg"
    if suffix not in {".jpg", ".jpeg", ".png", ".webp", ".bmp"}:
        suffix = ".jpg"
    return f"{sanitize_filename(category_label)}{suffix}"


def sanitize_filename(value: str) -> str:
    return "".join("_" if char in '<>:"/\\|?*' else char for char in value).strip() or "photo"


def get_workbook_loader():
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise ModuleNotFoundError(
            "openpyxl is required to read local simulation workbooks. "
            "Install v2-api requirements before bootstrapping from Excel files."
        ) from exc
    return load_workbook
